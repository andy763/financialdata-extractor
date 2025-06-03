import os
import re
import sys
import logging
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from urllib.parse import urlparse
from collections import defaultdict

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def is_blue_cell(cell):
    """Check if cell has blue background color"""
    # Simple approach to detect blue-ish cells
    try:
        if cell.fill and cell.fill.start_color and cell.fill.start_color.type == 'rgb':
            # Get RGB components
            color = cell.fill.start_color.rgb
            if isinstance(color, str) and len(color) >= 6:
                # Check for ARGB format (common in newer Excel files)
                if len(color) == 8:  # ARGB format
                    color = color[2:]  # Skip alpha channel
                
                # Extract RGB components
                try:
                    # Blue component is strong in blue cells
                    r = int(color[0:2], 16)
                    g = int(color[2:4], 16)
                    b = int(color[4:6], 16)
                    
                    # Check if this is a blue-ish cell (blue component dominates)
                    if b > 180 and b > (r + g) / 2:
                        return True
                except ValueError:
                    pass
    except Exception as e:
        # Silently handle any exceptions to avoid crashing
        pass
    
    return False

def is_valid_share_price(price_value, context_text="", url=""):
    """Check if the value looks like a valid share price"""
    # If it's not a numeric type, it's not valid
    if not isinstance(price_value, (int, float)):
        return False
    
    # Most share prices should not end in .0 exactly (though some legitimately might)
    # This helps catch dates (2023.0) and other numbers that aren't prices
    is_round_number = price_value == round(price_value, 0)
    
    if is_round_number:
        # Values below 100 that end in .0 are more likely to be valid
        if price_value < 100:
            return True
        else:
            # Values of 100.0 and above are suspicious
            return False
    
    return True

def find_processable_rows_and_get_urls(ws, url_column_letter):
    """Find all rows with URLs to process"""
    rows_with_urls = {}
    
    # Skip header row (row 1)
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{url_column_letter}{row}"]
        url = cell.value
        
        # Only process cells with valid URLs
        if url and isinstance(url, str) and url.strip().lower().startswith(("http://", "https://")):
            rows_with_urls[row] = url.strip()
    
    return rows_with_urls

def track_error_domain(url):
    """Helper function to track error domains with improved logic."""
    try:
        if not url or not isinstance(url, str):
            return "invalid://"
        
        # Clean the URL
        cleaned_url = url.strip()
        if not cleaned_url.lower().startswith(("http://", "https://")):
            return "invalid://"
        
        parsed_url = urlparse(cleaned_url)
        if parsed_url.netloc:
            return f"{parsed_url.scheme}://{parsed_url.netloc}/"
        else:
            # If no domain can be extracted, use the full URL
            return cleaned_url[:100] + ("..." if len(cleaned_url) > 100 else "")
    except Exception:
        # If all parsing fails, use the original URL (truncated if too long)
        url_str = str(url) if url else "unknown"
        return url_str[:100] + ("..." if len(url_str) > 100 else "")

def main():
    try:
        # Find Excel file
        possible_locations = [
            os.path.join("data", "Custodians.xlsx"),  # data folder, capital C (most likely)
            "Custodians.xlsx",                         # Current directory, capital C
            "custodians.xlsx",                         # Current directory, lowercase
            os.path.join("data", "custodians.xlsx"),   # data folder, lowercase
            os.path.join("..", "data", "Custodians.xlsx"), # Parent folder's data folder
            os.path.join("..", "data", "custodians.xlsx"), # Parent folder's data folder
        ]
        
        # Allow override from command line
        if len(sys.argv) > 1:
            excel_file = sys.argv[1]
        else:
            excel_file = None
            for location in possible_locations:
                if os.path.exists(location):
                    excel_file = location
                    print(f"Found Excel file: {location}")
                    break
        
        if not excel_file or not os.path.exists(excel_file):
            logging.error("Could not find Custodians.xlsx file")
            print("Error: Could not find Custodians.xlsx in standard locations.")
            print("Please specify the path as a command-line argument.")
            return
        
        # Constants (same as in excel_stock_updater.py)
        sheet_name = "Non-derivative exposures"
        dest_col = "L"  # Price column
        url_col = "P"   # URL column
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            if sheet_name not in wb.sheetnames:
                logging.error(f"Sheet '{sheet_name}' not found")
                print(f"Error: Sheet '{sheet_name}' not found.")
                return
            ws = wb[sheet_name]
        except Exception as e:
            logging.error(f"Error loading workbook: {e}")
            print(f"Error loading workbook: {e}")
            return
        
        # Initialize error tracking - exactly as in excel_stock_updater.py
        error_domains = defaultdict(int)
        error_urls = defaultdict(list)
        total_processed = 0
        blue_cells_count = 0
        blue_cells_data = []
        normal_cells_data = []
        invalid_price_domains = defaultdict(int)
        invalid_price_urls = defaultdict(list)
        invalid_price_count = 0
        
        # Get rows with URLs to process
        rows_urls = find_processable_rows_and_get_urls(ws, url_col)
        if not rows_urls:
            logging.info("No URLs found to process.")
            print("No URLs found to process.")
            return
        
        # Process each row (extract data for logging)
        for row, url in rows_urls.items():
            total_processed += 1
            
            # Skip invalid URLs
            if not isinstance(url, str) or not url.strip().lower().startswith(("http://","https://")):
                cell = ws[f"{dest_col}{row}"]
                is_blue = is_blue_cell(cell)
                
                cell_info = {
                    'row': row,
                    'url': url,
                    'success': False,
                    'has_invalid_price': False
                }
                
                if not is_blue:
                    domain = track_error_domain(url)
                    error_domains[domain] += 1
                    error_urls[domain].append(url)
                
                if is_blue:
                    blue_cells_data.append(cell_info)
                else:
                    normal_cells_data.append(cell_info)
                continue
            
            # Process the actual cell data
            price_cell = ws[f"{dest_col}{row}"]
            url_cell = ws[f"{url_col}{row}"]
            is_blue = is_blue_cell(url_cell)
            
            # Get the actual price value
            price_value = price_cell.value
            
            cell_info = {
                'row': row,
                'url': url,
                'success': False,
                'has_invalid_price': False
            }
            
            # Check if we have a price
            if price_value:
                # Determine if the price looks valid
                if isinstance(price_value, (int, float)) and not is_valid_share_price(price_value, url=url):
                    if not is_blue:
                        invalid_price_count += 1
                        domain = track_error_domain(url)
                        invalid_price_domains[domain] += 1
                        invalid_price_urls[domain].append(url)
                    cell_info["has_invalid_price"] = True
                else:
                    cell_info["success"] = True
            else:
                # No price = error
                if not is_blue:
                    domain = track_error_domain(url)
                    error_domains[domain] += 1
                    error_urls[domain].append(url)
            
            if is_blue:
                blue_cells_count += 1
                blue_cells_data.append(cell_info)
            else:
                normal_cells_data.append(cell_info)
        
        # Create improved log file with ranked errors and separated sections
        # THIS IS THE EXACT CODE FROM excel_stock_updater.py
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_filename = f"stock_prices_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        try:
            with open(log_filename, 'w', encoding='utf-8') as log_file:
                log_file.write("STOCK PRICES EXTRACTION LOG\n")
                log_file.write("=" * 50 + "\n")
                log_file.write(f"Run Date: {current_time}\n")
                log_file.write(f"Script: excel_stock_updater.py\n\n")
                
                # RANKED ERROR DOMAINS (most errors first) - ENHANCED: Show individual URLs
                log_file.write("SHARE PRICE ERRORS BY DOMAIN (RANKED - MOST ERRORS FIRST):\n")
                log_file.write("-" * 60 + "\n")
                if error_domains:
                    # Sort by error count in descending order
                    sorted_errors = sorted(error_domains.items(), key=lambda x: x[1], reverse=True)
                    for domain, count in sorted_errors:
                        log_file.write(f"{count:3d} errors: {domain}\n")
                        # Show the specific URLs for this domain
                        urls_for_domain = error_urls.get(domain, [])
                        for url in urls_for_domain[:5]:  # Show first 5 URLs
                            log_file.write(f"    └─ {url}\n")
                        if len(urls_for_domain) > 5:
                            log_file.write(f"    └─ ... and {len(urls_for_domain) - 5} more URLs\n")
                        log_file.write("\n")
                else:
                    log_file.write("No errors encountered!\n")
                
                # ENHANCED: INVALID PRICE VALUES SECTION (includes .0 values, percentages, years, etc.)
                log_file.write(f"\nINVALID PRICE VALUES BY DOMAIN (RANKED - MOST INVALID PRICES FIRST):\n")
                log_file.write("-" * 75 + "\n")
                log_file.write(f"Total invalid prices found: {invalid_price_count}\n")
                log_file.write("(These are likely incorrect - .0 values, dates, percentages, fund metadata, etc. instead of actual prices)\n")
                log_file.write("(Includes ALL .0 values: 24.00, 30.0, etc. - Excludes blue cells)\n\n")
                
                if invalid_price_domains:
                    # Sort by invalid price count in descending order
                    sorted_invalid_errors = sorted(invalid_price_domains.items(), key=lambda x: x[1], reverse=True)
                    for domain, count in sorted_invalid_errors:
                        log_file.write(f"{count:3d} invalid prices: {domain}\n")
                        # Show the specific URLs for this domain
                        urls_for_domain = invalid_price_urls.get(domain, [])
                        for url in urls_for_domain[:5]:  # Show first 5 URLs
                            log_file.write(f"    └─ {url}\n")
                        if len(urls_for_domain) > 5:
                            log_file.write(f"    └─ ... and {len(urls_for_domain) - 5} more URLs\n")
                        log_file.write("\n")
                else:
                    log_file.write("No invalid prices found!\n")
                
                # BLUE CELLS SECTION
                log_file.write(f"\nBLUE CELLS SECTION (EXCLUDED FROM SUCCESS RATE):\n")
                log_file.write("-" * 50 + "\n")
                log_file.write(f"Total blue cells: {blue_cells_count}\n")
                if blue_cells_data:
                    blue_success = sum(1 for cell in blue_cells_data if cell['success'])
                    blue_errors = blue_cells_count - blue_success
                    blue_invalid_prices = sum(1 for cell in blue_cells_data if cell.get('has_invalid_price', False))
                    log_file.write(f"Blue cells successful: {blue_success}\n")
                    log_file.write(f"Blue cells with errors: {blue_errors}\n")
                    log_file.write(f"Blue cells with invalid prices: {blue_invalid_prices}\n")
                    if blue_cells_count > 0:
                        blue_success_rate = blue_success / blue_cells_count
                        log_file.write(f"Blue cells success rate: {blue_success}/{blue_cells_count} = {blue_success_rate:.2%}\n")
                    log_file.write("\nBlue cell details:\n")
                    for cell in blue_cells_data[:10]:  # Show first 10
                        status = "✅ Success" if cell['success'] else ("🔸 Invalid Price" if cell.get('has_invalid_price') else "❌ Error")
                        log_file.write(f"  Row {cell['row']}: {status} - {cell['url'][:60]}{'...' if len(cell['url']) > 60 else ''}\n")
                    if len(blue_cells_data) > 10:
                        log_file.write(f"  ... and {len(blue_cells_data) - 10} more blue cells\n")
                else:
                    log_file.write("No blue cells found.\n")
                
                # NORMAL CELLS SECTION (Updated calculations)
                normal_cells_count = len(normal_cells_data)
                normal_success = sum(1 for cell in normal_cells_data if cell['success'])
                normal_errors = sum(1 for cell in normal_cells_data if not cell['success'] and not cell.get('has_invalid_price', False))
                normal_invalid_prices = sum(1 for cell in normal_cells_data if cell.get('has_invalid_price', False))
                
                log_file.write(f"\nNORMAL CELLS SECTION (USED FOR SUCCESS RATE):\n")
                log_file.write("-" * 50 + "\n")
                log_file.write(f"Total normal cells: {normal_cells_count}\n")
                log_file.write(f"Normal cells successful: {normal_success}\n")
                log_file.write(f"Normal cells with errors: {normal_errors}\n")
                log_file.write(f"Normal cells with invalid prices (treated as errors): {normal_invalid_prices}\n")
                if normal_cells_count > 0:
                    normal_success_rate = normal_success / normal_cells_count
                    log_file.write(f"Normal cells success rate: {normal_success}/{normal_cells_count} = {normal_success_rate:.2%}\n")
                    log_file.write(f"(Success rate excludes both errors and invalid prices)\n")
                
                log_file.write(f"\nSUMMARY:\n")
                log_file.write("-" * 10 + "\n")
                log_file.write(f"Total URLs processed: {total_processed}\n")
                log_file.write(f"Blue cells (excluded from success rate): {blue_cells_count}\n")
                log_file.write(f"Normal cells (used for success rate): {normal_cells_count}\n")
                log_file.write(f"Successful extractions (normal cells only): {normal_success}\n")
                log_file.write(f"URLs with invalid prices: {invalid_price_count}\n")
                log_file.write(f"Traditional errors: {sum(error_domains.values())}\n")
                if normal_cells_count > 0:
                    final_success_rate = normal_success / normal_cells_count
                    log_file.write(f"FINAL SUCCESS RATE: {normal_success}/{normal_cells_count} = {final_success_rate:.2%}\n")
                    log_file.write(f"(Excludes blue cells, traditional errors, AND invalid prices)\n")
                else:
                    log_file.write("FINAL SUCCESS RATE: N/A (no normal cells to process)\n")
            
            logging.info(f"Log file created: {log_filename}")
            print(f"Log file created: {log_filename}")
        except Exception as e:
            logging.error(f"Error creating log file: {e}")
            print(f"Error creating log file: {e}")
    
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        print(f"Error: {e}")

if __name__ == "__main__":
    main() 