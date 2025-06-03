import sys
import os
import re
import logging
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Color

# Add the project root to sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from excel_stock_updater import get_aminagroup_price, is_valid_share_price

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def test_price_extraction(url):
    """Test the Amina Group price extraction directly"""
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    )
    
    driver = None
    try:
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        
        # Extract price using our function
        price = get_aminagroup_price(driver, url)
        logging.info(f"Extracted price: {price}")
        
        # Test validation
        is_valid = is_valid_share_price(price, f"current price {price}", url)
        logging.info(f"Is price valid: {is_valid}")
        
        # Test with context
        context_text = f"Current price {price} USD"
        is_valid_with_context = is_valid_share_price(price, context_text, url)
        logging.info(f"Is price valid with better context: {is_valid_with_context}")
        
        return {
            "price": price,
            "is_valid": is_valid,
            "is_valid_with_context": is_valid_with_context,
            "context_text": context_text
        }
    except Exception as e:
        logging.error(f"Error in test: {e}", exc_info=True)
        return {"error": str(e)}
    finally:
        if driver:
            driver.quit()

def check_excel_handling(workbook_path="data/Custodians.xlsx", sheet_name="Non-derivative exposures"):
    """Check how the Excel file handles Amina Group entries"""
    try:
        wb = load_workbook(workbook_path)
        if sheet_name not in wb.sheetnames:
            logging.error(f"Sheet '{sheet_name}' not found")
            return False
            
        ws = wb[sheet_name]
        
        # Find Amina Group URLs
        amina_rows = []
        for row in range(1, ws.max_row + 1):
            url_cell = ws.cell(row=row, column=16)  # Column P
            if url_cell.value and "aminagroup.com" in str(url_cell.value).lower():
                price_cell = ws.cell(row=row, column=12)  # Column L
                amina_rows.append({
                    "row": row,
                    "url": url_cell.value,
                    "price_value": price_cell.value,
                    "price_cell_color": price_cell.fill.start_color.index if price_cell.fill.start_color else None
                })
        
        logging.info(f"Found {len(amina_rows)} Amina Group rows in Excel")
        for row_data in amina_rows:
            logging.info(f"Row {row_data['row']}: URL={row_data['url']}, Price={row_data['price_value']}, Cell color={row_data['price_cell_color']}")
        
        return amina_rows
    except Exception as e:
        logging.error(f"Error checking Excel: {e}", exc_info=True)
        return []

def main():
    """Main test function"""
    logging.info("Starting Amina Group test suite")
    
    # Test URLs from Amina Group
    test_urls = [
        "https://aminagroup.com/individuals/investments/btc-usd-tracker-certificate/",
        "https://aminagroup.com/individuals/investments/eth-usd-tracker-certificate/",
        "https://aminagroup.com/individuals/investments/dot-usd-tracker-certificate/",
        "https://aminagroup.com/individuals/investments/sol-usd-tracker-certificate/",
        "https://aminagroup.com/individuals/investments/ada-usd-tracker-certificate/"
    ]
    
    # Check Excel handling
    excel_rows = check_excel_handling()
    
    # Run extraction tests
    results = {}
    for url in test_urls:
        logging.info(f"Testing URL: {url}")
        result = test_price_extraction(url)
        results[url] = result
        logging.info(f"Result: {result}")
        
        # Small delay between requests
        time.sleep(2)
    
    # Print summary
    logging.info("\n==== TEST RESULTS SUMMARY ====")
    for url, result in results.items():
        if "error" in result:
            logging.info(f"❌ {url}: Error - {result['error']}")
        else:
            valid_status = "✓" if result['is_valid'] else "✗" 
            valid_context_status = "✓" if result['is_valid_with_context'] else "✗"
            logging.info(f"URL: {url}")
            logging.info(f"  Price: {result['price']}")
            logging.info(f"  Valid: {valid_status} (without context) | {valid_context_status} (with context)")
            logging.info(f"  Context text: '{result['context_text']}'")
    
    # Analyze the issue
    logging.info("\n==== ISSUE ANALYSIS ====")
    if any("error" in result for result in results.values()):
        logging.info("❌ Price extraction is failing for some URLs")
    elif all(result["is_valid"] for result in results.values()):
        logging.info("✅ All prices are valid - no validation issues")
    elif all(result["is_valid_with_context"] for result in results.values()):
        logging.info("⚠️ Prices are only valid with proper context - is_valid_share_price() may be too strict")
    else:
        logging.info("❌ Price validation is failing even with context - potential issue in is_valid_share_price()")
    
    # Fix recommendation
    logging.info("\n==== FIX RECOMMENDATION ====")
    
    price_values = [result["price"] for result in results.values() if "price" in result]
    if price_values:
        if all(price == int(price) for price in price_values):
            logging.info("All prices are whole numbers (.0) - issue is likely in the .0 validation rules")
            logging.info("Suggested fix: Update is_valid_share_price() to recognize Amina Group prices")
        else:
            logging.info("Some prices have decimals - issue could be elsewhere")
    
    # Additional Excel issue analysis
    if excel_rows:
        red_cells = [row for row in excel_rows if row["price_cell_color"] == "00FF0000"]
        if red_cells:
            logging.info(f"Found {len(red_cells)} red cells - indicates an error state")
        
        empty_values = [row for row in excel_rows if not row["price_value"]]
        if empty_values:
            logging.info(f"Found {len(empty_values)} empty price cells - data is not being saved")
        
        error_values = [row for row in excel_rows if isinstance(row["price_value"], str) and "error" in row["price_value"].lower()]
        if error_values:
            logging.info(f"Found {len(error_values)} error messages in price cells")
            
    logging.info("Test completed")

if __name__ == "__main__":
    main() 