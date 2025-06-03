import sys
import os
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import time

# Add the project root to sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from excel_stock_updater import get_aminagroup_price, preserve_cell_color_and_set_value

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def find_aminagroup_rows(wb, sheet_name="Non-derivative exposures"):
    """Find all rows with Amina Group URLs"""
    if sheet_name not in wb.sheetnames:
        logging.error(f"Sheet '{sheet_name}' not found")
        return []
        
    ws = wb[sheet_name]
    
    # Find Amina Group URLs in column P (16)
    amina_rows = []
    for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
        url_cell = ws.cell(row=row, column=16)  # Column P
        if url_cell.value and isinstance(url_cell.value, str) and "aminagroup.com" in url_cell.value.lower():
            amina_rows.append({
                "row": row,
                "url": url_cell.value
            })
    
    return amina_rows, ws

def update_aminagroup_prices(workbook_path="data/Custodians.xlsx", dest_col="L"):
    """Extract and update Amina Group prices in the Excel file"""
    try:
        # Load workbook
        wb = load_workbook(workbook_path)
        amina_rows, ws = find_aminagroup_rows(wb)
        
        if not amina_rows:
            logging.info("No Amina Group URLs found in the Excel file")
            return False
        
        logging.info(f"Found {len(amina_rows)} Amina Group URLs to process")
        
        # Initialize WebDriver
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
        )
        
        driver = None
        try:
            driver = webdriver.Chrome(
                service=ChromeService(ChromeDriverManager().install()),
                options=options
            )
            
            # Process each Amina Group URL
            success_count = 0
            for row_data in amina_rows:
                row = row_data["row"]
                url = row_data["url"]
                
                logging.info(f"Processing row {row}: {url}")
                
                try:
                    # Extract price
                    price = get_aminagroup_price(driver, url)
                    logging.info(f"Extracted price: {price}")
                    
                    # Update cell
                    cell = ws[f"{dest_col}{row}"]
                    preserve_cell_color_and_set_value(cell, f"current price: {price}")
                    success_count += 1
                    
                    # Save after each update
                    wb.save(workbook_path)
                    logging.info(f"Updated row {row} with price {price}")
                    
                except Exception as e:
                    logging.error(f"Error processing row {row}: {e}")
                
                # Small delay between requests
                time.sleep(2)
            
            # Final save
            wb.save(workbook_path)
            logging.info(f"Successfully updated {success_count} out of {len(amina_rows)} Amina Group prices")
            
            # Also update the results file if it exists
            results_path = workbook_path.replace("Custodians.xlsx", "Custodians_Results.xlsx")
            if os.path.exists(results_path):
                logging.info(f"Also updating results file: {results_path}")
                wb_results = load_workbook(results_path)
                amina_rows_results, ws_results = find_aminagroup_rows(wb_results)
                
                if amina_rows_results:
                    # Copy the values from the original file
                    for row_data in amina_rows_results:
                        row = row_data["row"]
                        original_cell = ws[f"{dest_col}{row}"]
                        results_cell = ws_results[f"{dest_col}{row}"]
                        preserve_cell_color_and_set_value(results_cell, original_cell.value)
                    
                    wb_results.save(results_path)
                    logging.info(f"Updated {len(amina_rows_results)} rows in results file")
            
            return True
                
        finally:
            if driver:
                driver.quit()
                
    except Exception as e:
        logging.error(f"Error updating Amina Group prices: {e}")
        return False

def main():
    """Main function"""
    logging.info("Starting Amina Group price update")
    
    # Make a backup of the Excel file first
    import shutil
    from datetime import datetime
    
    original_path = "data/Custodians.xlsx"
    backup_path = f"data/Custodians_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        shutil.copy2(original_path, backup_path)
        logging.info(f"Created backup at {backup_path}")
    except Exception as e:
        logging.error(f"Could not create backup: {e}")
        if input("Continue without backup? (y/n): ").lower() != 'y':
            logging.info("Exiting without making changes")
            return
    
    # Update the prices
    success = update_aminagroup_prices()
    
    if success:
        logging.info("✅ Amina Group prices successfully updated")
    else:
        logging.error("❌ Failed to update Amina Group prices")
        # Restore from backup if requested
        if input("Restore from backup? (y/n): ").lower() == 'y':
            try:
                shutil.copy2(backup_path, original_path)
                logging.info("Restored from backup")
            except Exception as e:
                logging.error(f"Error restoring from backup: {e}")

if __name__ == "__main__":
    main() 