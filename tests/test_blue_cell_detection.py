"""
Test script to verify blue cell detection functionality
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color

def is_blue_cell(cell):
    """
    Check if a cell has any blue-ish color (simplified detection)
    Returns True if the cell appears to be blue, False otherwise
    """
    try:
        if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
            color = cell.fill.fgColor
            # Check different color representations
            if hasattr(color, 'rgb') and color.rgb:
                # Handle both string and RGB object types
                rgb_value = color.rgb
                if hasattr(rgb_value, 'value'):  # RGB object case
                    rgb_value = rgb_value.value if rgb_value.value else str(rgb_value)
                elif not isinstance(rgb_value, str):  # Other object types
                    rgb_value = str(rgb_value)
                
                # Remove alpha channel if present (ARGB format)
                if isinstance(rgb_value, str) and len(rgb_value) == 8:  # ARGB format
                    rgb_value = rgb_value[2:]  # Remove first 2 characters (alpha)
                
                # Simplified blue detection - check if it contains blue-ish values
                # Look for various shades of blue (not just the exact #00B0F0)
                if isinstance(rgb_value, str) and len(rgb_value) >= 6:
                    # Convert to uppercase for consistent checking
                    rgb_upper = rgb_value.upper()
                    # Check for common blue patterns
                    blue_patterns = [
                        '00B0F0',  # Original target blue
                        '0000FF',  # Pure blue
                        '4472C4',  # Another common blue
                        '5B9BD5',  # Light blue
                        '2F75B5',  # Dark blue
                    ]
                    
                    # Check if it matches any blue pattern or if blue component is dominant
                    for pattern in blue_patterns:
                        if pattern in rgb_upper:
                            return True
                    
                    # Check if the blue component (last 2 digits) is significantly higher than red/green
                    try:
                        if len(rgb_upper) == 6:
                            red = int(rgb_upper[0:2], 16)
                            green = int(rgb_upper[2:4], 16)
                            blue = int(rgb_upper[4:6], 16)
                            # Consider it blue if blue component is much higher than red and green
                            if blue > 150 and blue > (red + green):
                                return True
                    except ValueError:
                        pass
                        
            elif hasattr(color, 'indexed') and color.indexed is not None:
                # Handle indexed colors - some indexed colors are blue
                # Common blue indexed colors: 5 (blue), 41 (light blue), etc.
                blue_indexed = [5, 41, 44, 49]
                return color.indexed in blue_indexed
    except Exception:
        # If anything fails, just return False
        pass
    return False

def test_blue_cell_detection():
    """Test the blue cell detection functionality"""
    wb = Workbook()
    ws = wb.active
    
    # Test case 1: Create a cell with the exact blue color we want to detect
    test_cell_1 = ws['A1']
    test_cell_1.value = "Test Blue Cell"
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    test_cell_1.fill = blue_fill
    
    # Test case 2: Create a cell with a different color
    test_cell_2 = ws['A2']
    test_cell_2.value = "Test Red Cell"
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    test_cell_2.fill = red_fill
    
    # Test case 3: Create a cell with no fill
    test_cell_3 = ws['A3']
    test_cell_3.value = "Test No Fill"
    
    # Test case 4: Create a cell with the blue color in ARGB format
    test_cell_4 = ws['A4']
    test_cell_4.value = "Test Blue Cell ARGB"
    blue_fill_argb = PatternFill(start_color='FF00B0F0', end_color='FF00B0F0', fill_type='solid')
    test_cell_4.fill = blue_fill_argb
    
    # Test case 5: Create a cell using Color object directly
    test_cell_5 = ws['A5']
    test_cell_5.value = "Test Blue Cell via Color Object"
    blue_color = Color(rgb='00B0F0')
    blue_fill_color = PatternFill(start_color=blue_color, end_color=blue_color, fill_type='solid')
    test_cell_5.fill = blue_fill_color
    
    # Test case 6: Create a cell with a different shade of blue
    test_cell_6 = ws['A6']
    test_cell_6.value = "Test Light Blue"
    light_blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    test_cell_6.fill = light_blue_fill
    
    # Test case 7: Create a cell with pure blue
    test_cell_7 = ws['A7']
    test_cell_7.value = "Test Pure Blue"
    pure_blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    test_cell_7.fill = pure_blue_fill
    
    # Test case 8: Create a cell with a greenish color (should NOT be detected as blue)
    test_cell_8 = ws['A8']
    test_cell_8.value = "Test Green"
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    test_cell_8.fill = green_fill
    
    # Save the test file
    wb.save("test_blue_detection.xlsx")
    print("Test file saved as 'test_blue_detection.xlsx'")
    
    # Test the detection
    print("\nTesting blue cell detection:")
    print(f"A1 (Blue #00B0F0): {is_blue_cell(test_cell_1)}")
    print(f"A2 (Red #FF0000): {is_blue_cell(test_cell_2)}")
    print(f"A3 (No fill): {is_blue_cell(test_cell_3)}")
    print(f"A4 (Blue ARGB): {is_blue_cell(test_cell_4)}")
    print(f"A5 (Blue via Color object): {is_blue_cell(test_cell_5)}")
    print(f"A6 (Light Blue #5B9BD5): {is_blue_cell(test_cell_6)}")
    print(f"A7 (Pure Blue #0000FF): {is_blue_cell(test_cell_7)}")
    print(f"A8 (Green #00FF00): {is_blue_cell(test_cell_8)}")
    
    # Test the specific color objects to see their types
    print(f"\nDebugging color object types:")
    if test_cell_1.fill and test_cell_1.fill.fgColor:
        color = test_cell_1.fill.fgColor
        print(f"A1 color.rgb type: {type(color.rgb)}")
        print(f"A1 color.rgb value: {color.rgb}")
    
    if test_cell_5.fill and test_cell_5.fill.fgColor:
        color = test_cell_5.fill.fgColor
        print(f"A5 color.rgb type: {type(color.rgb)}")
        print(f"A5 color.rgb value: {color.rgb}")
    
    print("Blue cell detection test completed!")

if __name__ == "__main__":
    test_blue_cell_detection() 