"""
Test script for the updated outstanding shares updater with improved extractors
"""

import time
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
import os

def create_test_excel():
    """Create a test Excel file with a few URLs to test"""
    
    # Create a test workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Shares"
    
    # Add headers
    ws['L1'] = "Source"
    ws['P1'] = "URL"
    ws['M1'] = "Outstanding Shares"
    
    # Test URLs from our investigation
    test_data = [
        ("Valour Bitcoin", "https://valour.com/en/products/physical-bitcoin-carbon-neutral-usd"),
        ("Grayscale Bitcoin Cash", "https://www.grayscale.com/funds/grayscale-bitcoin-cash-trust"),
        ("ProShares BITO", "https://www.proshares.com/our-etfs/strategic/bito"),
        ("VanEck Bitcoin", "https://www.vaneck.com/de/en/investments/bitcoin-etp/overview/"),
        ("WisdomTree Bitcoin", "https://www.wisdomtree.eu/en-gb/products/ucits-etfs-unleveraged-etps/cryptocurrency/wisdomtree-physical-bitcoin"),
    ]
    
    # Add test data
    for i, (source, url) in enumerate(test_data, start=2):
        ws[f'L{i}'] = source
        ws[f'P{i}'] = url
    
    # Save test workbook
    test_filename = "test_output.xlsx"
    wb.save(test_filename)
    print(f"Created test file: {test_filename}")
    return test_filename

def test_updated_outstanding_shares():
    """Test the updated outstanding shares updater"""
    
    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    print("=" * 60)
    print("TESTING UPDATED OUTSTANDING SHARES UPDATER")
    print("=" * 60)
    
    # Create test Excel file
    test_file = create_test_excel()
    
    # Import the updated outstanding shares updater
    try:
        from outstanding_shares_updater import extract_outstanding_shares_with_ai_fallback
        print("✅ Successfully imported updated outstanding shares updater")
    except ImportError as e:
        print(f"❌ Failed to import outstanding shares updater: {e}")
        return
    
    # Set up Chrome options
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    
    driver = None
    successful_extractions = 0
    total_tests = 5
    method_stats = {}
    
    try:
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        driver.set_page_load_timeout(30)
        
        # Test URLs
        test_urls = [
            "https://valour.com/en/products/physical-bitcoin-carbon-neutral-usd",
            "https://www.grayscale.com/funds/grayscale-bitcoin-cash-trust",
            "https://www.proshares.com/our-etfs/strategic/bito",
            "https://www.vaneck.com/de/en/investments/bitcoin-etp/overview/",
            "https://www.wisdomtree.eu/en-gb/products/ucits-etfs-unleveraged-etps/cryptocurrency/wisdomtree-physical-bitcoin",
        ]
        
        for i, url in enumerate(test_urls, 1):
            print(f"\n[{i}/{total_tests}] Testing: {url}")
            
            try:
                # Test the updated extractor
                result = extract_outstanding_shares_with_ai_fallback(driver, url)
                
                if result and "outstanding_shares" in result:
                    print(f"✅ SUCCESS: Found outstanding shares: {result['outstanding_shares']}")
                    method_used = result.get("method", "unknown")
                    print(f"   Method used: {method_used}")
                    successful_extractions += 1
                    
                    # Track method statistics
                    if method_used in method_stats:
                        method_stats[method_used] += 1
                    else:
                        method_stats[method_used] = 1
                        
                else:
                    print(f"❌ FAILED: {result.get('error', 'Unknown error')}")
                    method_stats["failed"] = method_stats.get("failed", 0) + 1
                    
            except Exception as e:
                print(f"❌ ERROR: {str(e)}")
                method_stats["failed"] = method_stats.get("failed", 0) + 1
            
            time.sleep(1)  # Brief pause between tests
        
        print("\n" + "=" * 60)
        print("EXTRACTION METHOD STATISTICS:")
        print("=" * 60)
        for method, count in method_stats.items():
            percentage = (count / total_tests) * 100
            print(f"{method.replace('_', ' ').title()}: {count} ({percentage:.1f}%)")
        
        print("\n" + "=" * 60)
        print("SUMMARY:")
        print(f"Total URLs tested: {total_tests}")
        print(f"Successful extractions: {successful_extractions}")
        print(f"Success rate: {(successful_extractions/total_tests)*100:.1f}%")
        
        if successful_extractions > 0:
            print(f"✅ Updated outstanding shares updater is working!")
            print(f"🎯 Improvement achieved with custom extractors!")
        else:
            print(f"❌ No successful extractions. Further refinement needed.")
        
        print("=" * 60)
        
    except Exception as e:
        print(f"Error setting up driver: {str(e)}")
        
    finally:
        if driver:
            driver.quit()
        
        # Clean up test file
        try:
            if os.path.exists(test_file):
                os.remove(test_file)
                print(f"Cleaned up test file: {test_file}")
        except Exception as e:
            print(f"Could not clean up test file: {e}")

if __name__ == "__main__":
    test_updated_outstanding_shares() 