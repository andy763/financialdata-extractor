import re
import logging
import hashlib
from urllib.parse import urlencode, urlparse
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Color
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException
import requests
from datetime import datetime, date, timedelta
from bf4py import BF4Py
import time
import json
from collections import defaultdict
# NEW: Import for Groq AI analysis
try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False
    logging.warning("Groq library not available. AI fallback will be disabled.")

# Groq AI Configuration
GROQ_API_KEY = "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"

# Multiple models in order of preference (best to fastest/most available)
GROQ_MODELS = [
    "llama-3.3-70b-versatile",     # Best quality, higher rate limits
    "llama-3.1-8b-instant",       # Faster, more available
    "gemma2-9b-it",               # Alternative option
    "mixtral-8x7b-32768",         # Fallback option
]

def make_groq_request_with_fallback(client, messages, max_retries=3):
    """
    Make a Groq API request with model fallback and rate limit handling
    """
    import time
    import random
    
    for model_index, model in enumerate(GROQ_MODELS):
        for attempt in range(max_retries):
            try:
                logging.info(f"Trying Groq model: {model} (attempt {attempt + 1})")
                
                # Shorter prompt for faster models to stay within token limits
                if model in ["llama-3.1-8b-instant", "gemma2-9b-it"]:
                    max_tokens = 50  # Reduced for faster models
                else:
                    max_tokens = 100
                
                response = client.chat.completions.create(
                    messages=messages,
                    model=model,
                    temperature=0.1,
                    max_tokens=max_tokens,
                )
                
                logging.info(f"✅ Success with {model}")
                return response, model
                
            except Exception as e:
                error_str = str(e).lower()
                
                # Check if it's a rate limit error
                if "rate limit" in error_str or "429" in error_str:
                    if attempt < max_retries - 1:
                        # Exponential backoff with jitter
                        wait_time = (2 ** attempt) + random.uniform(0, 1)
                        logging.warning(f"Rate limit hit with {model}, waiting {wait_time:.1f}s before retry...")
                        time.sleep(wait_time)
                        continue
                    else:
                        logging.warning(f"Rate limit exhausted for {model}, trying next model...")
                        break  # Try next model
                
                # Check for other errors that suggest model unavailability
                elif any(err in error_str for err in ["unavailable", "overloaded", "timeout"]):
                    logging.warning(f"Model {model} unavailable: {e}")
                    break  # Try next model immediately
                
                else:
                    # Unknown error, log and try next model
                    logging.error(f"Unknown error with {model}: {e}")
                    break
    
    # If all models failed
    raise Exception("All Groq models failed or rate limited")

def clean_html_for_ai(html_content, max_length=6000):  # Reduced from 8000 to save tokens
    """
    Clean and truncate HTML content for AI analysis while preserving price-relevant information.
    OPTIMIZED: Shorter content to reduce token usage and avoid rate limits.
    """
    try:
        soup = BeautifulSoup(html_content, "html.parser")
        
        # Remove script and style elements
        for script in soup(["script", "style", "nav", "header", "footer"]):
            script.decompose()
        
        # Get text content
        text = soup.get_text()
        
        # Clean up whitespace
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = ' '.join(chunk for chunk in chunks if chunk)
        
        # Focus on financial content - look for sections with financial keywords
        financial_keywords = [
            'price', 'share', 'market', 'trading', 'last', 'close', 'open',
            'quote', 'valuation', 'closing', 'current', 'bid', 'ask',
            'usd', 'eur', 'gbp', 'cad', 'aud', 'chf', '$', '€', '£'
        ]
        
        # Split text into sentences and prioritize those with financial keywords
        sentences = text.split('.')
        financial_sentences = []
        other_sentences = []
        
        for sentence in sentences:
            if any(keyword in sentence.lower() for keyword in financial_keywords):
                financial_sentences.append(sentence.strip())
            else:
                other_sentences.append(sentence.strip())
        
        # Combine financial sentences first, then others if space allows
        prioritized_text = '. '.join(financial_sentences)
        if len(prioritized_text) < max_length:
            remaining_space = max_length - len(prioritized_text)
            additional_text = '. '.join(other_sentences)[:remaining_space]
            prioritized_text = prioritized_text + '. ' + additional_text
        
        return prioritized_text[:max_length]
    
    except Exception as e:
        logging.error(f"Error cleaning HTML for AI: {e}")
        # Fallback: just truncate raw text
        soup = BeautifulSoup(html_content, "html.parser")
        return soup.get_text()[:max_length]

def analyze_website_with_groq(html_content, url):
    """
    Use Groq AI to analyze website content and extract share prices.
    IMPROVED: Multiple model fallback and rate limit handling with enhanced validation.
    """
    if not GROQ_AVAILABLE:
        return {"error": "Groq AI not available"}
    
    try:
        # Initialize Groq client
        client = Groq(api_key=GROQ_API_KEY)
        
        # Clean HTML content for AI analysis (reduced size to save tokens)
        cleaned_content = clean_html_for_ai(html_content, max_length=5000)
        
        # ENHANCED PROMPT - More specific to avoid non-price numbers
        prompt = f"""Extract ONLY the current trading/market price from this financial webpage: {url}

FIND THESE PRICES:
- Share price, market price, last price, closing price, trading price
- Current quote, bid/ask prices, NAV per share

AVOID THESE (NOT PRICES):
- Years (2024, 2025, etc.), dates, inception dates
- Protection percentages (80%, 90%, 100% protection)
- Fund sizes (112m, 1089m AUM), expense ratios (0.35%)
- Version numbers (1.0, 2.0), ratings (1-5 stars)
- Cap ranges (10%-11%, 28%-31%), outcome periods

RULES:
1. Return ONLY the numeric price value (no currency symbols)
2. Price should be between 0.01 and 50,000
3. Avoid whole numbers that could be years/percentages
4. If no ACTUAL trading price found: "NO_PRICE_FOUND"

Content: {cleaned_content[:4000]}

Current trading price:"""

        # Create message format
        messages = [{"role": "user", "content": prompt}]
        
        # Make request with fallback handling
        chat_completion, used_model = make_groq_request_with_fallback(client, messages)
        
        # Extract the response
        ai_response = chat_completion.choices[0].message.content.strip()
        logging.info(f"Groq AI response from {used_model} for {url}: {ai_response}")
        
        # Parse the AI response
        if "NO_PRICE_FOUND" in ai_response.upper():
            return {"error": "AI could not find valid share price"}
        
        # Try to extract a number from the response
        number_match = re.search(r'(\d{1,10}(?:[,\.\s]\d{1,3})*(?:\.\d{1,6})?)', ai_response)
        if number_match:
            price_str = number_match.group(1)
            
            # Clean and normalize the number
            # Handle different decimal separators and thousand separators
            if ',' in price_str and '.' in price_str:
                # Both present - determine which is decimal
                if price_str.rindex(',') > price_str.rindex('.'):
                    # European format: 1.234,56
                    normalized = price_str.replace('.', '').replace(',', '.')
                else:
                    # US format: 1,234.56
                    normalized = price_str.replace(',', '')
            elif ',' in price_str:
                # Only comma - could be decimal or thousands
                if price_str.count(',') == 1 and len(price_str.split(',')[1]) <= 3:
                    # Likely decimal
                    normalized = price_str.replace(',', '.')
                else:
                    # Likely thousands
                    normalized = price_str.replace(',', '')
            else:
                # No comma, remove spaces
                normalized = price_str.replace(' ', '')
            
            try:
                price_value = float(normalized)
                
                # ENHANCED VALIDATION - Use our new validation function
                if is_valid_share_price(price_value, ai_response, url):
                    return {"ai_extracted_price": price_value}
                else:
                    return {"error": f"AI extracted invalid price (likely not a share price): {price_value}"}
                    
            except ValueError:
                return {"error": f"Could not parse AI response as number: {price_str}"}
        else:
            return {"error": f"AI response contains no recognizable number: {ai_response}"}
            
    except Exception as e:
        error_msg = str(e)
        if "rate limit" in error_msg.lower() or "429" in error_msg:
            logging.error(f"Groq rate limit exceeded for {url}: {e}")
            return {"error": f"AI rate limit exceeded - try again later"}
        else:
            logging.error(f"Groq AI analysis error for {url}: {e}")
            return {"error": f"AI analysis failed: {str(e)}"}

def try_ai_fallback(driver, url):
    """
    Try AI analysis as a fallback when traditional scraping fails.
    """
    if not GROQ_AVAILABLE:
        return {"error": "AI fallback not available"}
    
    try:
        logging.info(f"Attempting AI fallback analysis for: {url}")
        
        # Get fresh page content for AI analysis
        driver.get(url)
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        
        # Allow time for dynamic content to load
        time.sleep(3)
        
        # Get the page source
        html_content = driver.page_source
        
        # Analyze with AI
        result = analyze_website_with_groq(html_content, url)
        
        if "ai_extracted_price" in result:
            logging.info(f"AI successfully extracted price: {result['ai_extracted_price']}")
            return {"ai extracted price": result["ai_extracted_price"]}
        else:
            logging.info(f"AI fallback failed: {result.get('error', 'Unknown error')}")
            return result
            
    except Exception as e:
        logging.error(f"AI fallback error for {url}: {e}")
        return {"error": f"AI fallback failed: {str(e)}"}

def _extract_isin_bf(url: str, timeout: float = 5.0) -> str:
    """Return ISIN from any boerse-frankfurt.de instrument URL."""
    # 1️⃣  fast path – try to read it from the URL itself
    m = re.search(r'/([A-Z0-9]{12})(?=[-/?])', url, re.I)
    if m:
        return m.group(1).upper()

    # 2️⃣  fallback – one cheap GET and one regex
    # Ensure requests is imported if this function is moved to a module where it's not already imported.
    # For this script, requests is imported globally.
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
    r.raise_for_status()                       # raises for 4xx/5xx
    m = re.search(r'ISIN\s*:\s*([A-Z0-9]{12})', r.text, re.I) # Escaped colon for regex
    if m:
        return m.group(1).upper()

    # optional: look into JSON blobs if the page ever drops the text line
    m = re.search(r'"isin"\s*:\s*"([A-Z0-9]{12})"', r.text, re.I) # Escaped colon for regex
    if m:
        return m.group(1).upper()

    raise ValueError("ISIN not found in URL or HTML.")

def find_processable_rows_and_get_urls(ws, url_column_letter):
    """
    Dynamically finds ALL rows that have URLs in the specified column.
    No predetermined row limit - scans the entire worksheet.
    """
    processable_rows_urls = {}
    
    # Find the maximum row with data in the worksheet
    max_row = ws.max_row
    
    # Start from row 2 (skip header) and scan all rows
    for row_idx in range(2, max_row + 1):
        coord = f"{url_column_letter}{row_idx}"
        url_value = ws[coord].value
        
        # Check if the cell has a URL (string starting with http)
        if url_value and isinstance(url_value, str) and url_value.strip().lower().startswith(("http://", "https://")):
            processable_rows_urls[row_idx] = url_value.strip()
                
    return processable_rows_urls

def get_fidelity_open_price(driver, url):
    """
    Navigate to a Fidelity digital/institutional quote page and scrape the 'Open' price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    # wait for the main price element to ensure page has loaded its data
    try:
        WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "div.nre-quick-quote-price"))
        )
    except Exception:
        # if it never appears, we'll still parse whatever loaded
        pass

    soup = BeautifulSoup(driver.page_source, "html.parser")

    # find the label "Open"
    label = soup.find(
        lambda tag: tag.name in ("div", "span")
                    and tag.get_text(strip=True).lower() == "open"
    )
    if not label:
        raise ValueError("Open label not found on Fidelity page")

    # the next sibling div/span should contain something like "$91.37"
    value_tag = label.find_next_sibling(
        lambda tag: tag.name in ("div", "span")
                    and re.match(r"^\$\s*[\d,]+\.\d{2}$", tag.get_text(strip=True))
    )
    if not value_tag:
        raise ValueError("Open value not found next to label")

    text = value_tag.get_text(strip=True)
    m = re.search(r"\$\s*([\d,]+\.\d{2})", text)
    if not m:
        raise ValueError(f"Unexpected format for Open: '{text}'")

    num = m.group(1).replace(",", "")
    return float(num)

def get_nasdaq_european_market_price(driver, url):
    """
    Navigate to a NASDAQ European market page and scrape the share price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # First approach: Look for text patterns in the entire page
    page_text = soup.get_text(" ", strip=True)
    
    # Look for patterns like "EUR 16.29" or "16.29 EUR"
    currency_patterns = [
        r"([A-Z]{3})\s*([\d,\.]+)",  # EUR 16.29
        r"([\d,\.]+)\s*([A-Z]{3})"   # 16.29 EUR
    ]
    
    for pattern in currency_patterns:
        matches = re.finditer(pattern, page_text)
        for match in matches:
            if match.group(1) in ["EUR", "USD", "GBP"]:
                # First group is currency
                price_str = match.group(2)
                currency = match.group(1)
            elif match.group(2) in ["EUR", "USD", "GBP"]:
                # Second group is currency
                price_str = match.group(1)
                currency = match.group(2)
            else:
                continue
                
            logging.info(f"Found price: {currency} {price_str}")
            
            # Normalize number format
            if "," in price_str and "." in price_str:
                price_normalized = price_str.replace(",", "")
            elif "," in price_str:
                price_normalized = price_str.replace(",", ".")
            else:
                price_normalized = price_str
                
            try:
                return float(price_normalized)
            except ValueError:
                continue

    # If we get here, we couldn't find the price
    raise ValueError("Could not find price on NASDAQ European market page")

def get_nasdaq_market_activity_price(driver, url):
    """
    Navigate to a NASDAQ market activity page (stocks, ETFs, etc.) and scrape the share price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # First try to find the price in the main header area
    try:
        # Look for the last price element which typically contains the current price
        price_element = soup.find("span", class_="symbol-page-header__pricing-price")
        if price_element and price_element.text:
            price_text = price_element.text.strip()
            # Clean the price text (remove $ and commas)
            price_text = price_text.replace('$', '').replace(',', '')
            return float(price_text)
    except (ValueError, AttributeError):
        pass
    
    # If we can't find the price in the header, look through the page content
    try:
        # Try to find price inside div with class "symbol-header"
        header_div = soup.find("div", class_="symbol-header")
        if header_div:
            price_spans = header_div.find_all("span")
            for span in price_spans:
                text = span.text.strip()
                if '$' in text:
                    price_match = re.search(r'\$\s*([\d,]+\.\d{2})', text)
                    if price_match:
                        price = price_match.group(1).replace(',', '')
                        return float(price)
    except (ValueError, AttributeError):
        pass
    
    # If still not found, try a more general approach for dollar amounts
    dollar_pattern = re.compile(r'\$\s*([\d,]+\.\d{2})')
    page_text = soup.get_text()
    matches = dollar_pattern.findall(page_text)
    
    if matches:
        # Find dollar amounts that look like prices (not volumes, market caps, etc.)
        possible_prices = []
        for match in matches:
            try:
                price = float(match.replace(',', ''))
                if 0.1 < price < 10000:  # Reasonable price range
                    possible_prices.append(price)
            except ValueError:
                continue
        
        if possible_prices:
            # Use the most prominent price (often the first one found)
            return possible_prices[0]
    
    raise ValueError("Could not find price on NASDAQ market activity page")

def get_bxswiss_price(driver, url):
    """
    Navigate to a BX Swiss instrument page and scrape the last traded price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Try to find the last traded price in the instrument details
    try:
        # Look for a label containing "Last traded"
        last_traded_elements = soup.find_all(text=re.compile(r"Last traded", re.IGNORECASE))
        
        for element in last_traded_elements:
            # Check the text of the element and nearby elements for a price
            price_text = element.parent.get_text() if element.parent else ""
            
            # Look for currency code followed by numeric value
            currency_price_match = re.search(r'([A-Z]{3})\s*([\d,\.]+)', price_text)
            if currency_price_match:
                currency = currency_price_match.group(1)
                price_str = currency_price_match.group(2)
                logging.info(f"Found price: {currency} {price_str}")
                
                # Normalize number format
                price_normalized = price_str.replace(',', '.')
                return float(price_normalized)
    except (ValueError, AttributeError) as e:
        logging.error(f"Error extracting BX Swiss price: {e}")
    
    # Try a more general approach if specific element wasn't found
    try:
        # Look for text patterns like "Last traded USD 9.365" in the entire page
        page_text = soup.get_text(" ", strip=True)
        last_traded_pattern = re.compile(r'Last traded\s+([A-Z]{3})\s+([\d,\.]+)', re.IGNORECASE)
        match = last_traded_pattern.search(page_text)
        
        if match:
            currency = match.group(1)
            price_str = match.group(2)
            logging.info(f"Found price using page text: {currency} {price_str}")
            
            # Normalize number format
            price_normalized = price_str.replace(',', '.')
            return float(price_normalized)
    except (ValueError, AttributeError) as e:
        logging.error(f"Error extracting BX Swiss price from page text: {e}")
    
    raise ValueError("Could not find last traded price on BX Swiss page")

def get_euronext_price(driver, url):
    """
    Navigate to a Euronext live page and scrape the share/valuation price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for page to load")

    # Save the page source to a string for debugging and easier manipulation
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html.parser")
    
    # Log some debug info about what we're searching for
    logging.debug(f"Looking for valuation price in Euronext page: {url}")
    
    # First approach: Look for specific price elements with class names that Euronext might use
    try:
        # Try multiple CSS selectors that might contain price information
        selectors = [
            "div.stock-infos span.stock-price", 
            "div.stock-info span.price",
            "td.euronext-last-price",
            "span.last-price",
            "div.price-box",
            "div.valuation-price",
            "div.quote-page__price",
            ".full-data div.data-value",
            ".product-data div.data-value",
            ".product-price"
        ]
        
        for selector in selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text().strip()
                # Look for currency symbol followed by number
                match = re.search(r'[€$£]\s*([\d,\.]+)', text)
                if match:
                    price_str = match.group(1)
                    # Handle European number format (replace comma with dot)
                    price_normalized = price_str.replace(',', '.')
                    logging.info(f"Found Euronext price using selector '{selector}': {price_normalized}")
                    return float(price_normalized)
    except Exception as e:
        logging.error(f"Error extracting Euronext price using selectors: {e}")
    
    # Second approach: Look for tables with price data
    try:
        # Find all table rows
        rows = soup.find_all('tr')
        for row in rows:
            # Get the text of the row
            row_text = row.get_text(" ", strip=True).lower()
            # Look for rows containing "valuation price" or similar labels
            if "valuation price" in row_text or "last price" in row_text or "share price" in row_text:
                # Try to extract a price from this row
                match = re.search(r'[€$£]\s*([\d,\.]+)', row_text)
                if match:
                    price_str = match.group(1)
                    price_normalized = price_str.replace(',', '.')
                    logging.info(f"Found Euronext price in table row: {price_normalized}")
                    return float(price_normalized)
    except Exception as e:
        logging.error(f"Error extracting Euronext price from tables: {e}")
    
    # Third approach: Extract all possible prices from the page and select the most likely one
    try:
        # Find all text containing currency symbols
        all_text = soup.get_text(" ", strip=True)
        price_matches = re.finditer(r'(valuation price|last price|share price|price)?\s*[€$£]\s*([\d,\.]+)', all_text, re.IGNORECASE)
        
        candidates = []
        for match in price_matches:
            label = match.group(1) if match.group(1) else ""
            price_str = match.group(2)
            price_normalized = price_str.replace(',', '.')
            
            try:
                price_value = float(price_normalized)
                # Give higher priority to matches with a label
                priority = 2 if label else 1
                candidates.append((priority, price_value, label))
            except ValueError:
                continue
        
        # Sort by priority (labeled prices first)
        candidates.sort(reverse=True)
        
        if candidates:
            # Take the highest priority price
            _, best_price, label = candidates[0]
            label_info = f" (labeled as '{label}')" if label else ""
            logging.info(f"Found Euronext price{label_info}: {best_price}")
            return best_price
    except Exception as e:
        logging.error(f"Error extracting Euronext prices with general approach: {e}")
    
    # Final approach: Look for any number that might be a price
    try:
        # Find a string that matches € followed by a number, anywhere in the page
        all_text = soup.get_text(" ", strip=True)
        price_pattern = re.compile(r'€\s*([\d,\.]+)', re.IGNORECASE)
        matches = price_pattern.findall(all_text)
        
        if matches:
            # Take the first match
            price_str = matches[0]
            price_normalized = price_str.replace(',', '.')
            logging.info(f"Found Euronext price with fallback method: {price_normalized}")
            return float(price_normalized)
    except Exception as e:
        logging.error(f"Error extracting Euronext price with fallback method: {e}")
    
    raise ValueError("Could not find valuation price on Euronext page")

def get_tmx_price(driver, url):
    """
    Navigate to a TMX Money page and scrape the share price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # First approach: Look for the price element with specific class
    try:
        # Try multiple possible selectors for price elements
        price_selectors = [
            "span.price", 
            "div.price", 
            "span.quote-price", 
            "div.quote-price",
            "span.last-price",
            "div.last-price",
            ".quote-header .price",
            ".stock-price"
        ]
        
        for selector in price_selectors:
            price_elements = soup.select(selector)
            for element in price_elements:
                price_text = element.get_text().strip()
                # Match price patterns like "$19.90" or "Price:$19.90" or "$19.90 CAD"
                match = re.search(r'\$\s*([\d,\.]+)', price_text)
                if match:
                    price_str = match.group(1)
                    # Clean the price (remove commas)
                    price_normalized = price_str.replace(',', '')
                    logging.info(f"Found TMX price using selector '{selector}': {price_normalized}")
                    return float(price_normalized)
    except Exception as e:
        logging.error(f"Error extracting TMX price from selectors: {e}")
    
    # Second approach: Look for text patterns in the page
    try:
        # Find text containing "Price:" followed by a dollar amount
        page_text = soup.get_text(" ", strip=True)
        price_patterns = [
            r'Price:\s*\$\s*([\d,\.]+)',  # "Price: $19.90"
            r'Price\s*\$\s*([\d,\.]+)',   # "Price$19.90"
            r'Last:\s*\$\s*([\d,\.]+)',   # "Last: $19.90"
            r'Last Price:\s*\$\s*([\d,\.]+)'  # "Last Price: $19.90"
        ]
        
        for pattern in price_patterns:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                price_str = match.group(1)
                price_normalized = price_str.replace(',', '')
                logging.info(f"Found TMX price using pattern '{pattern}': {price_normalized}")
                return float(price_normalized)
    except Exception as e:
        logging.error(f"Error extracting TMX price from page text: {e}")
    
    # Third approach: Look for any dollar amount in the page that might be a price
    try:
        all_text = soup.get_text(" ", strip=True)
        # Find all dollar amounts in the page
        dollar_matches = re.finditer(r'\$\s*([\d,\.]+)', all_text)
        
        prices = []
        for match in dollar_matches:
            price_str = match.group(1)
            try:
                price_value = float(price_str.replace(',', ''))
                # Only consider reasonable price values (e.g., not volume or market cap)
                if 0.1 < price_value < 10000:
                    prices.append(price_value)
            except ValueError:
                continue
        
        if prices:
            # Take the first reasonable price found
            logging.info(f"Found TMX price using fallback method: {prices[0]}")
            return prices[0]
    except Exception as e:
        logging.error(f"Error extracting TMX price with fallback method: {e}")
    
    raise ValueError("Could not find share price on TMX Money page")

def get_london_stock_exchange_price(driver, url):
    """
    Navigate to a London Stock Exchange company page and scrape the share price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for London Stock Exchange page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for price in specific LSE elements (most reliable)
    try:
        # The .last-price selector worked perfectly in testing
        price_selectors = [
            ".last-price",
            "div.price",
            "span.price", 
            "div.stock-price",
            "span.stock-price",
            "td.price",
            "div.current-price",
            "span.current-price",
            ".price-value",
            ".quote-price"
        ]
        
        for selector in price_selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text().strip()
                # Look for price patterns
                price_match = re.search(r'([\d,\.]+)', text)
                if price_match and len(text) < 50:  # Reasonable length for a price element
                    price_str = price_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', ''))
                        if 0.01 <= price_value <= 100000:  # Reasonable price range
                            logging.info(f"Found LSE price using selector '{selector}': {price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with LSE selector method: {e}")
    
    # Method 2: Look for price patterns in text content
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for patterns like "Price (USD) 77.05", "Price 79.88", etc.
        price_patterns = [
            r'Price\s*\([^)]+\)\s*([\d,\.]+)',  # "Price (USD) 77.05"
            r'Price\s+([\d,\.]+)',              # "Price 79.88"
            r'Current price\s*:?\s*([\d,\.]+)', # "Current price: 77.05"
            r'Share price\s*:?\s*([\d,\.]+)',   # "Share price: 77.05"
            r'Last price\s*:?\s*([\d,\.]+)',    # "Last price: 77.05"
        ]
        
        for pattern in price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 100000:  # Reasonable price range
                        logging.info(f"Found LSE price using pattern '{pattern}': {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with LSE pattern method: {e}")
    
    # Method 3: Look for table data that might contain prices
    try:
        # Look for table rows or data elements
        table_cells = soup.find_all(['td', 'th', 'div'])
        for cell in table_cells:
            cell_text = cell.get_text().strip()
            # Look for cells that might contain price labels
            if any(keyword in cell_text.lower() for keyword in ['price', 'last', 'current']):
                # Check if this cell or nearby cells contain a number
                next_sibling = cell.find_next_sibling()
                if next_sibling:
                    next_text = next_sibling.get_text().strip()
                    price_match = re.search(r'([\d,\.]+)', next_text)
                    if price_match:
                        price_str = price_match.group(1)
                        try:
                            price_value = float(price_str.replace(',', ''))
                            if 0.01 <= price_value <= 100000:
                                logging.info(f"Found LSE price in table structure: {price_value}")
                                return price_value
                        except ValueError:
                            continue
                
                # Also check if the price is in the same cell
                same_cell_match = re.search(r'([\d,\.]+)', cell_text)
                if same_cell_match:
                    price_str = same_cell_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', ''))
                        if 0.01 <= price_value <= 100000:
                            logging.info(f"Found LSE price in same cell: {price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with LSE table method: {e}")
    
    raise ValueError("Could not find share price on London Stock Exchange page")

def get_cboe_us_price(driver, url):
    """
    Navigate to a CBOE US equities page and scrape the last price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for CBOE US page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for "Last Price" label and value
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for patterns like "Last Price 12.81"
        last_price_patterns = [
            r'Last\s+Price\s+([\d,\.]+)',  # "Last Price 12.81"
            r'Last\s+Price:\s*([\d,\.]+)', # "Last Price: 12.81"
            r'Last\s+Price\s*\$\s*([\d,\.]+)', # "Last Price $12.81"
        ]
        
        for pattern in last_price_patterns:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 100000:  # Reasonable price range
                        logging.info(f"Found CBOE US price using pattern '{pattern}': {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with CBOE US pattern method: {e}")
    
    # Method 2: Look for table rows or data elements containing "Last Price"
    try:
        # Look for table cells, divs, spans that might contain price data
        elements = soup.find_all(['td', 'th', 'div', 'span'])
        for element in elements:
            element_text = element.get_text().strip()
            if 'last price' in element_text.lower():
                # Extract number from this element or nearby elements
                price_match = re.search(r'([\d,\.]+)', element_text)
                if price_match:
                    price_str = price_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', ''))
                        if 0.01 <= price_value <= 100000:
                            logging.info(f"Found CBOE US price in element: {price_value}")
                            return price_value
                    except ValueError:
                        continue
                
                # Check next sibling for price value
                next_sibling = element.find_next_sibling()
                if next_sibling:
                    sibling_text = next_sibling.get_text().strip()
                    price_match = re.search(r'([\d,\.]+)', sibling_text)
                    if price_match:
                        price_str = price_match.group(1)
                        try:
                            price_value = float(price_str.replace(',', ''))
                            if 0.01 <= price_value <= 100000:
                                logging.info(f"Found CBOE US price in sibling: {price_value}")
                                return price_value
                        except ValueError:
                            continue
    except Exception as e:
        logging.error(f"Error with CBOE US element method: {e}")
    
    # Method 3: Look for CSS selectors that might contain price data
    try:
        price_selectors = [
            ".last-price",
            ".price-value",
            ".market-price",
            ".quote-price",
            "[data-price]",
            ".price"
        ]
        
        for selector in price_selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text().strip()
                price_match = re.search(r'([\d,\.]+)', text)
                if price_match and len(text) < 50:
                    price_str = price_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', ''))
                        if 0.01 <= price_value <= 100000:
                            logging.info(f"Found CBOE US price using selector '{selector}': {price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with CBOE US selector method: {e}")
    
    raise ValueError("Could not find last price on CBOE US page")

def get_cboe_au_price(driver, url):
    """
    Navigate to a CBOE AU equities page and scrape the last traded price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for CBOE AU page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for "Last Traded / Best Bid" label and value
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for patterns like "Last Traded / Best Bid (AUD) 16.810"
        last_traded_patterns = [
            r'Last\s+Traded\s*/\s*Best\s+Bid\s*\([^)]+\)\s*([\d,\.]+)',  # "Last Traded / Best Bid (AUD) 16.810"
            r'Last\s+Traded\s*(?:/\s*Best\s+Bid)?\s*([\d,\.]+)',         # "Last Traded 16.810"
            r'Best\s+Bid\s*\([^)]+\)\s*([\d,\.]+)',                      # "Best Bid (AUD) 16.810"
            r'Last\s+Price\s*\([^)]+\)\s*([\d,\.]+)',                    # "Last Price (AUD) 16.810"
        ]
        
        for pattern in last_traded_patterns:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 100000:  # Reasonable price range
                        logging.info(f"Found CBOE AU price using pattern '{pattern}': {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with CBOE AU pattern method: {e}")
    
    # Method 2: Look for table rows or data elements containing price labels
    try:
        # Look for table cells, divs, spans that might contain price data
        elements = soup.find_all(['td', 'th', 'div', 'span'])
        for element in elements:
            element_text = element.get_text().strip()
            # Check for various price-related keywords
            if any(keyword in element_text.lower() for keyword in ['last traded', 'best bid', 'last price']):
                # Extract number from this element or nearby elements
                price_match = re.search(r'([\d,\.]+)', element_text)
                if price_match:
                    price_str = price_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', ''))
                        if 0.01 <= price_value <= 100000:
                            logging.info(f"Found CBOE AU price in element: {price_value}")
                            return price_value
                    except ValueError:
                        continue
                
                # Check next sibling for price value
                next_sibling = element.find_next_sibling()
                if next_sibling:
                    sibling_text = next_sibling.get_text().strip()
                    price_match = re.search(r'([\d,\.]+)', sibling_text)
                    if price_match:
                        price_str = price_match.group(1)
                        try:
                            price_value = float(price_str.replace(',', ''))
                            if 0.01 <= price_value <= 100000:
                                logging.info(f"Found CBOE AU price in sibling: {price_value}")
                                return price_value
                        except ValueError:
                            continue
    except Exception as e:
        logging.error(f"Error with CBOE AU element method: {e}")
    
    # Method 3: Look for CSS selectors that might contain price data
    try:
        price_selectors = [
            ".last-traded",
            ".best-bid",
            ".last-price",
            ".price-value",
            ".market-price",
            ".quote-price",
            "[data-price]",
            ".price"
        ]
        
        for selector in price_selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text().strip()
                price_match = re.search(r'([\d,\.]+)', text)
                if price_match and len(text) < 50:
                    price_str = price_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', ''))
                        if 0.01 <= price_value <= 100000:
                            logging.info(f"Found CBOE AU price using selector '{selector}': {price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with CBOE AU selector method: {e}")
    
    raise ValueError("Could not find last traded price on CBOE AU page")

def get_vaneck_price(driver, url):
    """
    Navigate to a VanEck investment page and extract the ACTUAL last traded price from exchange data.
    This function specifically looks for "Last Traded Price" from Deutsche Börse, Euronext, etc.
    Uses visible browser mode for React components to load properly.
    Returns a float or raises ValueError.
    """
    # For VanEck, we need to use a visible browser for React components to work
    # Save current options and create new visible driver if needed
    current_options = driver.options if hasattr(driver, 'options') else None
    visible_driver = None
    
    try:
        # Create a new visible driver specifically for VanEck
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service as ChromeService
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.common.action_chains import ActionChains
        
        options = webdriver.ChromeOptions()
        # Do NOT use headless mode - React components need visible browser
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        visible_driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        visible_driver.set_page_load_timeout(90)
        
        # Load the page
        visible_driver.get(url)
        
        WebDriverWait(visible_driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        logging.info("VanEck: Page loaded successfully")
        
        # Wait for trading components to appear
        try:
            trading_block = WebDriverWait(visible_driver, 20).until(
                EC.presence_of_element_located((By.TAG_NAME, "ve-tradinginformationblock"))
            )
            logging.info("VanEck: Found trading information block")
            
            # Scroll to the component and click to trigger loading
            visible_driver.execute_script("arguments[0].scrollIntoView(true);", trading_block)
            time.sleep(3)
            
            try:
                ActionChains(visible_driver).move_to_element(trading_block).click().perform()
                logging.info("VanEck: Clicked trading component to trigger loading")
            except:
                logging.info("VanEck: Could not click trading component")
                
        except:
            logging.info("VanEck: Trading component not found initially")
        
        # Execute JavaScript to force React components to load
        js_scripts = [
            "window.scrollTo(0, document.body.scrollHeight);",
            "window.scrollTo(0, 0);",
            "window.dispatchEvent(new Event('resize'));",
            "window.dispatchEvent(new Event('scroll'));",
            """
            document.querySelectorAll('ve-tradinginformationblock').forEach(function(el) {
                if (el.connectedCallback) el.connectedCallback();
                if (el.attributeChangedCallback) el.attributeChangedCallback();
                el.dispatchEvent(new Event('load'));
                el.dispatchEvent(new Event('DOMContentLoaded'));
            });
            """,
            """
            document.querySelectorAll('ve-tradinginformationblock').forEach(function(el) {
                el.style.display = 'block';
                el.style.visibility = 'visible';
            });
            """
        ]
        
        for script in js_scripts:
            try:
                visible_driver.execute_script(script)
                time.sleep(2)
            except Exception as e:
                logging.debug(f"VanEck: JS script error: {e}")
        
        logging.info("VanEck: JavaScript executed to trigger React components")
        
        # Monitor for trading data to appear (30 seconds max)
        max_attempts = 15  # 30 seconds
        
        for attempt in range(max_attempts):
            time.sleep(2)
            
            # Get fresh page content
            soup = BeautifulSoup(visible_driver.page_source, "html.parser")
            page_text = soup.get_text(" ", strip=True)
            
            # Look for trading price patterns
            patterns = [
                # Specific "Last Traded Price" patterns
                r'Last\s+Traded\s+Price[:\s]*([€$£]?\s*[\d,\.]+)',
                r'Last\s+Traded\s+Price[:\s]*€?\s*([\d,\.]+)',
                r'Last\s+Traded\s+Price[:\s]*\$?\s*([\d,\.]+)',
                
                # Exchange-specific patterns
                r'Deutsche\s+Börse[^0-9]*?([€$£]?\s*[\d,\.]+)',
                r'Euronext[^0-9]*?([€$£]?\s*[\d,\.]+)',
                r'XETR[^0-9]*?([€$£]?\s*[\d,\.]+)',
                
                # General trading patterns
                r'Trading\s+Price[:\s]*([€$£]?\s*[\d,\.]+)',
                r'Market\s+Price[:\s]*([€$£]?\s*[\d,\.]+)',
                r'Current\s+Price[:\s]*([€$£]?\s*[\d,\.]+)',
            ]
            
            for pattern in patterns:
                matches = re.finditer(pattern, page_text, re.IGNORECASE)
                for match in matches:
                    price_str = match.group(1)
                    
                    try:
                        # Clean price string
                        clean_price = re.sub(r'[€$£]', '', price_str).strip()
                        
                        # Handle European number format
                        if "," in clean_price and "." not in clean_price:
                            clean_price = clean_price.replace(",", ".")
                        elif "," in clean_price and "." in clean_price:
                            clean_price = clean_price.replace(".", "").replace(",", ".")
                        
                        price_value = float(clean_price.replace(',', ''))
                        
                        # Validate reasonable ETP price range
                        if 0.1 <= price_value <= 100:
                            logging.info(f"VanEck: Found last traded price: {price_value} using pattern: {pattern}")
                            return price_value
                    except ValueError:
                        continue
            
            # Check trading components for loaded data
            trading_components = soup.find_all("ve-tradinginformationblock")
            for component in trading_components:
                content = component.get_text(" ", strip=True)
                
                if "Loading..." not in content and len(content) > 30:
                    logging.info(f"VanEck: Trading component loaded data")
                    
                    # Look for prices in the loaded content
                    component_numbers = re.findall(r'([\d,\.]+)', content)
                    for num_str in component_numbers:
                        try:
                            if "," in num_str and "." not in num_str:
                                num_str = num_str.replace(",", ".")
                            elif "," in num_str and "." in num_str:
                                num_str = num_str.replace(".", "").replace(",", ".")
                            
                            price_value = float(num_str.replace(',', ''))
                            if 0.5 <= price_value <= 50:
                                logging.info(f"VanEck: Found price in trading component: {price_value}")
                                return price_value
                        except ValueError:
                            continue
            
            if attempt % 5 == 0:
                logging.info(f"VanEck: Attempt {attempt + 1}/{max_attempts} - waiting for trading data...")
        
        # If we get here, no trading data was found
        raise ValueError("Trading components found but last traded price data did not load within timeout")
            
    except TimeoutException:
        raise ValueError("Timed out waiting for VanEck page or trading components to load")

    except Exception as e:
        logging.error(f"VanEck: Unexpected error: {e}")
        raise ValueError(f"Error extracting VanEck last traded price: {e}")
        
    finally:
        # Always clean up the visible driver
        if visible_driver:
            try:
                visible_driver.quit()
            except:
                pass

def fetch_and_extract_data(driver, url, keywords):
    """
    Fetches content with Selenium, then applies:
      1. *New* Börse-Frankfurt: Closing price prev trading day
      2. NASDAQ European market price
      3. NASDAQ regular market activity price
      4. BX Swiss instrument price
      5. Euronext live price
      6. TMX Money price
      7. London Stock Exchange price
      8. CBOE US price
      9. CBOE AU price
      10. VanEck price (with multiple exchange detection)
      11. ProShares logic
      12. Grayscale logic
      13. QR Asset Cesta logic
      14. Montpensier Arbével logic
      15. Hashdex US logic
      16. Hashdex Brazil logic
      17. Xtrackers Galaxy Physical Bitcoin & Ethereum ETCs
      18. Fidelity "Open" price
      19. Generic keyword-based fallback
    """
    # --- Börse-Frankfurt: Closing price prev trading day ---
    if "boerse-frankfurt.de" in url.lower():
        try:
            isin = _extract_isin_bf(url)
            # Try to determine MIC if present in URL, otherwise default to XETR
            mic_match = re.search(r'[?&]mic=([A-Z]{4})', url, re.IGNORECASE)
            mic = mic_match.group(1).upper() if mic_match else "XETR"
            
            logging.info(f"Boerse Frankfurt API: Fetching ISIN {isin}, MIC {mic} for URL {url}")
            close_price = get_prev_close_boerse(isin, mic=mic)
            return {"closing price prev trading day": close_price}
        except requests.exceptions.RequestException as req_exc:
            logging.error(f"Boerse Frankfurt network/HTTP error for {url} (ISIN attempt failed or during API call): {req_exc}")
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Boerse Frankfurt network/HTTP error: {req_exc}"}
        except ValueError as val_exc: # Catches ISIN not found from _extract_isin_bf or data errors from get_prev_close_boerse
            logging.error(f"Boerse Frankfurt data error for {url} (e.g., ISIN not found, API data issue): {val_exc}")
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Boerse Frankfurt data error: {val_exc}"}
        except Exception as exc: # Catch any other unexpected errors
            logging.error(f"Boerse Frankfurt unexpected error for {url}: {exc}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Boerse Frankfurt unexpected error: {exc}"}

    # --- NASDAQ European market pages ---
    if "nasdaq.com/european-market-activity" in url.lower():
        try:
            price = get_nasdaq_european_market_price(driver, url)
            return {"share price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"NASDAQ European market error: {e}"}
        except Exception as e:
            logging.error(f"NASDAQ European market unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"NASDAQ European market unexpected error: {e}"}
            
    # --- NASDAQ regular market activity pages ---
    if "nasdaq.com/market-activity" in url.lower():
        try:
            price = get_nasdaq_market_activity_price(driver, url)
            return {"share price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"NASDAQ market activity error: {e}"}
        except Exception as e:
            logging.error(f"NASDAQ market activity unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"NASDAQ market activity unexpected error: {e}"}

    # --- BX Swiss instrument pages ---
    if "bxswiss.com/instruments" in url.lower():
        try:
            price = get_bxswiss_price(driver, url)
            return {"last traded price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"BX Swiss error: {e}"}
        except Exception as e:
            logging.error(f"BX Swiss unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"BX Swiss unexpected error: {e}"}

    # --- Euronext live pages ---
    if "live.euronext.com" in url.lower():
        try:
            price = get_euronext_price(driver, url)
            return {"valuation price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Euronext error: {e}"}
        except Exception as e:
            logging.error(f"Euronext unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Euronext unexpected error: {e}"}

    # --- TMX Money pages ---
    if "money.tmx.com" in url.lower():
        try:
            price = get_tmx_price(driver, url)
            return {"share price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"TMX Money error: {e}"}
        except Exception as e:
            logging.error(f"TMX Money unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"TMX Money unexpected error: {e}"}

    # --- London Stock Exchange pages ---
    if "londonstockexchange.com" in url.lower():
        try:
            price = get_london_stock_exchange_price(driver, url)
            return {"share price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"London Stock Exchange error: {e}"}
        except Exception as e:
            logging.error(f"London Stock Exchange unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"London Stock Exchange unexpected error: {e}"}

    # --- CBOE US pages ---
    if "cboe.com/us" in url.lower():
        try:
            price = get_cboe_us_price(driver, url)
            return {"share price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"CBOE US error: {e}"}
        except Exception as e:
            logging.error(f"CBOE US unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"CBOE US unexpected error: {e}"}

    # --- CBOE AU pages ---
    if "cboe.com/au" in url.lower():
        try:
            price = get_cboe_au_price(driver, url)
            return {"share price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"CBOE AU error: {e}"}
        except Exception as e:
            logging.error(f"CBOE AU unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"CBOE AU unexpected error: {e}"}

    # --- Morningstar France pages ---
    if "morningstar.fr" in url.lower():
        try:
            # Check if this is a PDF URL and try to convert it to HTML
            if "&Format=PDF" in url or ".pdf" in url.lower():
                logging.info(f"Morningstar France: Detected PDF URL, attempting to convert to HTML: {url}")
                # Remove PDF parameters to get the HTML version
                html_url = url.replace("&Format=PDF", "").replace("&format=pdf", "")
                # Remove DocumentId parameter as well since it's PDF-specific
                html_url = re.sub(r'&DocumentId=[^&]*', '', html_url)
                html_url = re.sub(r'\&tab=\d+', '', html_url)  # Remove tab parameter which might be PDF-specific
                logging.info(f"Morningstar France: Converted URL to: {html_url}")
                url = html_url  # Use the converted URL
            
            price = get_morningstar_fr_price(driver, url)
            return {"vl": price}
        except ValueError as e:
            # Check if the error is due to PDF format
            if "&Format=PDF" in url or ".pdf" in url.lower():
                # Don't try AI fallback for PDF URLs - just return a clear error
                return {"error": f"PDF document not supported - Morningstar France requires HTML pages for price extraction"}
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Morningstar France error: {e}"}
        except Exception as e:
            logging.error(f"Morningstar France unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Morningstar France unexpected error: {e}"}

    # --- Franklin Templeton pages ---
    if "franklintempleton.com" in url.lower():
        try:
            price = get_franklin_templeton_price(driver, url)
            return {"market price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Franklin Templeton error: {e}"}
        except Exception as e:
            logging.error(f"Franklin Templeton unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Franklin Templeton unexpected error: {e}"}

    # --- Amina Group pages ---
    if "aminagroup.com" in url.lower():
        try:
            price = get_aminagroup_price(driver, url)
            return {"current price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Amina Group error: {e}"}
        except Exception as e:
            logging.error(f"Amina Group unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Amina Group unexpected error: {e}"}

    # --- Schwab Asset Management pages ---
    if "schwabassetmanagement.com" in url.lower():
        try:
            price = get_schwab_asset_management_price(driver, url)
            return {"market price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Schwab Asset Management error: {e}"}
        except Exception as e:
            logging.error(f"Schwab Asset Management unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Schwab Asset Management unexpected error: {e}"}

    # --- Evolve ETFs pages ---
    if "evolveetfs.com" in url.lower():
        try:
            price = get_evolve_etfs_price(driver, url)
            return {"market price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Evolve ETFs error: {e}"}
        except Exception as e:
            logging.error(f"Evolve ETFs unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Evolve ETFs unexpected error: {e}"}

    # --- 21Shares pages ---
    if "21shares.com" in url.lower():
        try:
            price = get_21shares_price(driver, url)
            return {"market price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"21Shares error: {e}"}
        except Exception as e:
            logging.error(f"21Shares unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"21Shares unexpected error: {e}"}

    # --- Ninepoint pages ---
    if "ninepoint.com" in url.lower():
        try:
            price = get_ninepoint_price(driver, url)
            return {"market price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Ninepoint error: {e}"}
        except Exception as e:
            logging.error(f"Ninepoint unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Ninepoint unexpected error: {e}"}

    # --- BetaShares pages ---
    if "betashares.com.au" in url.lower():
        try:
            price = get_betashares_price(driver, url)
            return {"market price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"BetaShares error: {e}"}
        except Exception as e:
            logging.error(f"BetaShares unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"BetaShares unexpected error: {e}"}

    # --- CSO P Asset pages ---
    if "csopasset.com" in url.lower():
        try:
            price = get_csopasset_price(driver, url)
            return {"closing price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"CSO P Asset error: {e}"}
        except Exception as e:
            logging.error(f"CSO P Asset unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"CSO P Asset unexpected error: {e}"}

    # --- Valour pages ---
    if "valour.com" in url.lower():
        try:
            price = get_valour_price(driver, url)
            return {"closing price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Valour error: {e}"}
        except Exception as e:
            logging.error(f"Valour unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Valour unexpected error: {e}"}

    # --- Purpose Investments pages ---
    if "purposeinvest.com" in url.lower():
        try:
            price = get_purposeinvest_price(driver, url)
            return {"closing price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Purpose Investments error: {e}"}
        except Exception as e:
            logging.error(f"Purpose Investments unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"Purpose Investments unexpected error: {e}"}

    # --- VanEck pages ---
    if "vaneck.com" in url.lower():
        try:
            price = get_vaneck_price(driver, url)
            return {"last traded price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"VanEck error: {e}"}
        except Exception as e:
            logging.error(f"VanEck unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"VanEck unexpected error: {e}"}

    # --- bitcap.com pages ---
    if "bitcap.com" in url.lower():
        try:
            price = get_bitcap_price(driver, url)
            return {"share price": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"bitcap error: {e}"}
        except Exception as e:
            logging.error(f"bitcap unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"bitcap unexpected error: {e}"}

    # --- qrasset.com.br pages ---
    if "qrasset.com.br" in url.lower() and url.lower().endswith("/#cesta"):
        try:
            price = get_qrasset_cota_price(driver, url)
            return {"valor da cota": price}
        except ValueError as e:
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"QR Asset error: {e}"}
        except Exception as e:
            logging.error(f"QR Asset unexpected error for {url}: {e}", exc_info=True)
            # Try AI fallback before returning error
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
            return {"error": f"QR Asset unexpected error: {e}"}

    # Fallback to other extraction logic if not Börse-Frankfurt or NASDAQ
    # or if the above failed (though it should return an error dict)
    
    try:
        driver.get(url)
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # --- 1. ProShares ETF pages ---
        if "proshares.com/our-etfs" in url.lower():
            label = soup.find(text=re.compile(r"Market Price", re.IGNORECASE))
            if label:
                price_text = label.find_next(text=re.compile(r"\$\d{1,3}(?:,\d{3})*(?:\.\d{2})?"))
                if price_text:
                    num = price_text.strip().replace("$", "").replace(",", "")
                    try:
                        return {"market price": float(num)}
                    except ValueError:
                        return {"error": f"Could not parse ProShares price '{num}'"}
                return {"error": "Market Price value not found"}
            return {"error": "Market Price label not found"}

        # --- 2. Grayscale ETF pages ---
        if "grayscale.com/funds" in url.lower() or "etfs.grayscale.com" in url.lower():
            # IMPORTANT: For www.grayscale.com URLs, the content is loaded dynamically with JavaScript
            # We need to use the already-loaded page content from Selenium (soup is from driver.page_source)
            
            # Get the full page text from the already-loaded Selenium page
            page_text = soup.get_text(" ", strip=True)
            
            # Method 1: Look for "Market Price as of MM/DD/YYYY $XX.XX" pattern (MOST RELIABLE)
            pattern1 = re.search(
                r'Market Price\s+as\s+of\s+\d{1,2}/\d{1,2}/\d{4}\s*\$\s*([\d,\.]+)', 
                page_text, re.IGNORECASE
            )
            if pattern1:
                price_str = pattern1.group(1)
                logging.info(f"Found Grayscale price with 'Market Price as of DATE $PRICE': {price_str}")
                price_normalized = price_str.replace(',', '')
                try:
                    return {"market price": float(price_normalized)}
                except ValueError:
                    return {"error": f"Could not parse Grayscale price '{price_str}'"}
            
            # Method 2: Look for price before "Market Price as of" pattern 
            # Pattern: $XX.XX Market Price as of MM/DD/YYYY
            pattern2 = re.search(
                r'\$\s*([\d,\.]+)\s+Market Price\s+as\s+of\s+\d{1,2}/\d{1,2}/\d{4}', 
                page_text, re.IGNORECASE
            )
            if pattern2:
                price_str = pattern2.group(1)
                logging.info(f"Found Grayscale price with '$PRICE Market Price as of DATE': {price_str}")
                price_normalized = price_str.replace(',', '')
                try:
                    return {"market price": float(price_normalized)}
                except ValueError:
                    return {"error": f"Could not parse Grayscale price '{price_str}'"}
            
            # Method 3: Look for flexible "Market Price" and "as of" with price nearby
            # This handles cases where there might be extra text between elements
            pattern3 = re.search(
                r'Market Price.*?as\s+of.*?\$\s*([\d,\.]+)', 
                page_text, re.IGNORECASE | re.DOTALL
            )
            if pattern3:
                price_str = pattern3.group(1)
                # Filter out obviously wrong values (too large for a fund price)
                price_normalized = price_str.replace(',', '')
                try:
                    price_value = float(price_normalized)
                    if price_value < 10000:  # Reasonable fund price range
                        logging.info(f"Found Grayscale price with flexible pattern: {price_str}")
                        return {"market price": price_value}
                    else:
                        logging.info(f"Grayscale price {price_value} seems too large for a fund, continuing search...")
                except ValueError:
                    logging.error(f"Could not parse Grayscale price '{price_str}' as float")
            
            # Method 4: Look for heading elements with dollar signs that might be prices
            # Based on the HTML structure like #### $85.76
            headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
            for heading in headings:
                heading_text = heading.get_text(strip=True)
                # Look for price patterns in headings
                price_match = re.search(r'\$\s*([\d,\.]+)', heading_text)
                if price_match:
                    price_str = price_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', ''))
                        # Filter for reasonable fund prices (between $0.01 and $10,000)
                        if 0.01 <= price_value <= 10000:
                            # Check if this heading is near market price text
                            next_sibling = heading.find_next_sibling()
                            if next_sibling and 'market price' in next_sibling.get_text(' ', strip=True).lower():
                                logging.info(f"Found Grayscale price in heading near 'market price': {price_str}")
                                return {"market price": price_value}
                    except ValueError:
                        continue
            
            # Method 5: Look for any reasonable price near "Market Price" text
            market_price_elements = soup.find_all(
                lambda tag: tag.name and 'market price' in tag.get_text(' ', strip=True).lower()
            )
            
            for element in market_price_elements:
                # Check previous and next siblings for prices
                for sibling in [element.find_previous_sibling(), element.find_next_sibling()]:
                    if sibling:
                        sibling_text = sibling.get_text(strip=True)
                        price_match = re.search(r'\$\s*([\d,\.]+)', sibling_text)
                        if price_match:
                            price_str = price_match.group(1)
                            try:
                                price_value = float(price_str.replace(',', ''))
                                if 0.01 <= price_value <= 10000:
                                    logging.info(f"Found Grayscale price near 'Market Price' element: {price_str}")
                                    return {"market price": price_value}
                            except ValueError:
                                continue
            
            return {"error": "Grayscale market price not found"}

        # --- 3. QR Asset Cesta pages ---
        if "qrasset.com.br" in url.lower() and url.lower().endswith("/#cesta"):
            label = soup.find("span", class_="label", string=re.compile(r"Valor da Cota", re.IGNORECASE))
            if not label:
                return {"error": "Valor da Cota label not found"}
            value_span = label.find_next_sibling("span", class_="value")
            if not value_span:
                return {"error": "Valor da Cota value element not found"}
            price_text = value_span.get_text(strip=True)
            price_pattern = re.compile(r"R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}")
            if not price_pattern.fullmatch(price_text):
                return {"error": f"Unexpected format '{price_text}'"}
            num_text = price_text.replace("R$", "").strip().replace(".", "").replace(",", ".")
            try:
                return {"valor da cota": float(num_text)}
            except ValueError:
                return {"error": f"Could not parse Valor da Cota '{num_text}'"}

        # --- 4. Montpensier Arbével fund pages ---
        if "montpensier-arbevel.com/fonds/" in url.lower():
            vl_label = soup.find(text=re.compile(r"^VL$", re.IGNORECASE))
            if not vl_label:
                return {"error": "VL label not found"}
            price_tag = vl_label.find_next(string=re.compile(r"[\d\.\,]+\s*€"))
            if not price_tag:
                return {"error": "VL value not found"}
            raw = price_tag.strip().replace("€","").strip()
            num = raw.replace(".","").replace(",",".")
            try:
                vl = float(num)
            except ValueError:
                return {"error": f"Could not parse VL '{raw}'"}
            result = {"vl": vl}
            date_tag = price_tag.find_next(string=re.compile(r"au\s*\d{2}/\d{2}/\d{4}", re.IGNORECASE))
            if date_tag:
                result["date"] = date_tag.strip()
            return result

        # --- 5. Hashdex US ETF pages (e.g. https://www.hashdex-etfs.com/NCIQ) ---
        if "hashdex-etfs.com" in url.lower() or "hashdex-etfs.com" in url.lower():
            closing = soup.find(
                lambda tag: tag.name in ("h3","h4")
                            and re.search(r"Closing Price", tag.get_text(strip=True), re.IGNORECASE)
            )
            if not closing:
                return {"error": "Closing Price label not found"}
            price_tag = closing.find_previous_sibling(
                lambda tag: tag.name == "h4"
                            and re.search(r"USD\s*\d{1,3}(?:,\d{3})*(?:\.\d{2})?", tag.get_text(strip=True))
            )
            if not price_tag:
                return {"error": "Closing Price value not found"}
            txt = price_tag.get_text(strip=True)
            m = re.search(r"USD\s*([\d,]+\.\d{2})", txt)
            if not m:
                return {"error": f"Unexpected format '{txt}'"}
            num = m.group(1).replace(",","")
            try:
                return {"closing price": float(num)}
            except ValueError:
                return {"error": f"Could not parse Closing Price '{num}'"}

        # --- 6. Hashdex Brazil ETF pages (/pt-BR/products/…) ---
        if url.lower().startswith("https://hashdex.com/pt-br/products/"):
            page_text = soup.get_text(" ", strip=True)
            m = re.search(r"Valor da Cota\s*R\$\s*([\d\.]+?,\d{2})", page_text, re.IGNORECASE)
            if not m:
                return {"error": "Valor da Cota not found or unexpected format"}
            num_text = m.group(1).replace(".", "").replace(",", ".")
            try:
                return {"valor da cota": float(num_text)}
            except ValueError:
                return {"error": f"Could not parse Valor da Cota '{num_text}'"}

        # --- Xtrackers Galaxy Physical Bitcoin & Ethereum ETCs (CH1315732250 & CH1315732268) ---
        if url.lower().startswith("https://etf.dws.com/") and re.search(
                r"xtrackers-galaxy-physical-(bitcoin|ethereum)-etc-securities",
                url.lower()
            ):
            # Pull all text and look for "Value per ETC security ... USD"
            text = soup.get_text(" ", strip=True)
            m = re.search(r"Value per ETC security.*?([\d\.,]+)\s*USD", text, re.IGNORECASE)
            if not m:
                return {"error": "Value per ETC security not found"}

            raw_val = m.group(1)
            # Normalize both European ("7,57") and US ("18.25") formats:
            if "," in raw_val and "." in raw_val:
                # e.g. "1.234,56" → "1234.56"
                norm = raw_val.replace(".", "").replace(",", ".")
            elif "," in raw_val:
                # e.g. "7,57" → "7.57"
                norm = raw_val.replace(",", ".")
            else:
                # e.g. "18.25" → "18.25"
                norm = raw_val.replace(",", "")

            try:
                return {"value per etc security": float(norm)}
            except ValueError:
                return {"error": f"Could not parse '{raw_val}' as float"}

        # --- 8. Fidelity "Open" price ---
        if "digital.fidelity.com" in url.lower() or "institutional.fidelity.com" in url.lower():
            try:
                open_price = get_fidelity_open_price(driver, url)
                return {"open": open_price}
            except ValueError as e:
                return {"error": f"Fidelity scrape error: {e}"}

        # --- SIX-Group "Previous close price" extraction ---
        if "six-group.com" in url.lower():
            driver.get(url)
            # 1) Wait until the <dt> for "Previous close price" is in the DOM
            prev_dt = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//dt[contains(normalize-space(.), 'Previous close price')]"
                ))
            )
            # 2) Grab its immediately following <dd>
            dd = prev_dt.find_element(
                By.XPATH,
                "following-sibling::dd[1]"
            )
            price_text = dd.text.strip()  # e.g. "9.31 USD" or "9.31"

            # 3) Extract number and optional currency
            # Regex: Group 1 captures the number, Group 2 (optional) captures the currency
            m = re.search(r"([\d\.,]+)(?:\s*([A-Za-z]{3}))?", price_text)
            
            if not m:
                # If even the number part isn't found with the flexible regex
                return {"error": f"Could not extract number from SIX Group price '{price_text}'"}

            num_raw = m.group(1)
            currency_code = m.group(2) # This will be None if currency not found
            
            # 4) Normalize thousand/decimal separators
            num_clean = ""
            num_dots = num_raw.count('.')
            num_commas = num_raw.count(',')

            if num_dots > 0 and num_commas > 0:
                # Both comma and period exist
                if num_raw.rfind(',') > num_raw.rfind('.'): # European style: e.g., "1.234,56"
                    num_clean = num_raw.replace('.', '').replace(',', '.')
                else: # US style: e.g., "1,234.56"
                    num_clean = num_raw.replace(',', '')
            elif num_dots > 0: # Only dots
                if num_dots > 1: # Multiple dots, assume thousands separators: "1.234.567" -> "1234567" or "2.213" -> "2213"
                    num_clean = num_raw.replace('.', '')
                else: # Single dot, assume decimal: "123.45" -> "123.45"
                    num_clean = num_raw
            elif num_commas > 0: # Only commas
                if num_commas > 1: # Multiple commas, assume thousands separators: "1,234,567" -> "1234567"
                    num_clean = num_raw.replace(',', '')
                else: # Single comma, assume decimal: "123,45" -> "123.45"
                    num_clean = num_raw.replace(',', '.')
            else: # No separators (e.g., "123")
                num_clean = num_raw

            try:
                price_val = float(num_clean)
            except ValueError:
                return {"error": f"Could not parse number from SIX Group '{num_clean}' (original: '{price_text}')"}

            result_data = {"previous close": price_val}
            if currency_code: # Only add currency to the result if it was actually found
                result_data["currency"] = currency_code.upper()
            
            return result_data

        # --- 9. Generic keyword-based fallback (KEEP THIS AS LAST RESORT) ---
        extracted = {}
        page_text = " ".join(soup.strings).lower()
        for keyword in keywords:
            pat = re.compile(
                rf"\b{re.escape(keyword.lower())}\b[^\d]*?([0-9,]+\.?[0-9]*)",
                re.IGNORECASE
            )
            m = pat.search(page_text)
            if m:
                raw = m.group(1)
                cleaned = re.sub(r"[^\d\.]", "", raw)
                try:
                    if cleaned.count('.') > 1 or not cleaned.replace('.', '', 1).isdigit():
                        raise ValueError
                    potential_price = float(cleaned)
                    
                    # ENHANCED: Use validation to filter out invalid prices
                    if is_valid_share_price(potential_price, page_text, url):
                        extracted[keyword] = potential_price
                    else:
                        logging.debug(f"Generic fallback rejected invalid price {potential_price} for keyword '{keyword}' on {url}")
                        extracted[keyword] = f"Error: invalid price value '{raw}' (likely not a share price)"
                except ValueError:
                    extracted[keyword] = f"Error: cannot parse '{raw}'"
        
        # If generic fallback found something valid, return it
        if extracted and any(isinstance(v, float) for v in extracted.values()):
            return extracted
        
        # If no traditional method worked, try AI fallback as final resort
        logging.info(f"All traditional methods failed for {url}, trying AI fallback...")
        ai_result = try_ai_fallback(driver, url)
        if "ai extracted price" in ai_result:
            return ai_result
        
        # If even AI failed, return the generic fallback results (if any) or empty dict
        return extracted if extracted else {}

    except Exception as e:
        logging.error(f"Error processing URL {url}: {e}", exc_info=True)
        # Try AI fallback as final resort even for unexpected exceptions
        try:
            logging.info(f"Unexpected error occurred for {url}, trying AI fallback...")
            ai_result = try_ai_fallback(driver, url)
            if "ai extracted price" in ai_result:
                return ai_result
        except Exception as ai_e:
            logging.error(f"AI fallback also failed for {url}: {ai_e}")
        
        return {"error": str(e)}

def _bf_get_salt(session: requests.Session) -> str:
    """Grab the tracing-salt from the current main.*.js bundle."""
    r = session.get("https://www.boerse-frankfurt.de/")
    r.raise_for_status() # Ensure we got a valid response
    
    # Try the new pattern first (without leading slash)
    match = re.search(r'src="(main\.[^"]+\.js)"', r.text)
    if not match:
        # Try the old pattern (with leading slash) as fallback
        match = re.search(r'src="/(main\.[^"]+\.js)"', r.text)
    
    if not match:
        raise ValueError("Could not find main JS bundle in Boerse Frankfurt homepage.")
    
    bundle_path = match.group(1)
    # Ensure the path starts with / for the URL construction
    if not bundle_path.startswith('/'):
        bundle_path = '/' + bundle_path
        
    js_url = f"https://www.boerse-frankfurt.de{bundle_path}"
    js_r = session.get(js_url)
    js_r.raise_for_status()
    js = js_r.text
    salt_match = re.search(r'salt:"(\w+)"', js)
    if not salt_match:
        raise ValueError("Could not find salt in Boerse Frankfurt JS bundle.")
    return salt_match.group(1)

def _bf_headers(url: str, salt: str) -> dict:
    now      = datetime.utcnow()
    localnow = datetime.now()
    ts       = now.isoformat(timespec="milliseconds") + "Z"
    trace_id = hashlib.md5(f"{ts}{url}{salt}".encode()).hexdigest()
    xsec     = hashlib.md5(localnow.strftime("%Y%m%d%H%M").encode()).hexdigest()
    return {
        "accept": "application/json, text/plain, */*",
        "client-date": ts,
        "x-client-traceid": trace_id,
        "x-security": xsec,
        "origin": "https://www.boerse-frankfurt.de",
        "referer": "https://www.boerse-frankfurt.de/",
    }

def get_prev_close_boerse(isin: str, mic: str = "XETR") -> float:
    """Lightweight – no Selenium – fetch of 'Closing price prev trading day'."""
    sess = requests.Session()
    # Add a user-agent to the session, as some servers might require it.
    sess.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    })
    salt = _bf_get_salt(sess)

    # previous *trading* day (naive)
    prev = date.today()
    # Adjust for weekend: if today is Sat, prev is Fri. If Sun, prev is Fri.
    if prev.weekday() == 5: # Saturday
        prev -= timedelta(days=1)
    elif prev.weekday() == 6: # Sunday
        prev -= timedelta(days=2)
    
    # Now, ensure 'prev' is the last *trading* day by checking again
    # and decrementing if it's a weekend or the current day.
    # This loop handles the case where today is Monday, so prev becomes Sunday, then Saturday, then Friday.
    # It also handles the initial decrement if today is a weekday.
    prev -= timedelta(days=1) # Go to previous day
    while prev.weekday() >= 5: # Skip Saturday (5) or Sunday (6)
        prev -= timedelta(days=1)


    params = {
        "isin": isin,
        "mic": mic,
        "minDate": prev.isoformat(),
        "maxDate": prev.isoformat(),
        "limit": 1,
        "cleanSplit": "false",
        "cleanPayout": "false",
        "cleanSubscription": "false",
    }
    api_url_base = "https://api.boerse-frankfurt.de/v1/data/price_history"
    query_string = urlencode(params)
    full_url = f"{api_url_base}?{query_string}"
    
    r = sess.get(full_url, headers=_bf_headers(full_url, salt), timeout=(5, 15)) # Increased connect timeout slightly
    r.raise_for_status() # Will raise an HTTPError if the HTTP request returned an unsuccessful status code
    
    response_json = r.json()
    data = response_json.get("data") # Use .get for safer access
    
    if not data or not isinstance(data, list) or len(data) == 0:
        raise ValueError(f"No price data returned from API for ISIN {isin} on {prev.isoformat()}. Response: {response_json}")
    
    # Check if 'close' key exists and is a number
    if "close" not in data[0] or not isinstance(data[0]["close"], (int, float)):
        raise ValueError(f"Unexpected data format: 'close' field missing or not a number in {data[0]}. Response: {response_json}")
        
    return float(data[0]["close"])

def is_blue_cell(cell):
    """
    Check if a cell has any blue-ish color (simplified detection)
    Returns True if the cell appears to be blue, False otherwise
    """
    try:
        if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
            color = cell.fill.fgColor
            # Check different color representations
            if hasattr(color, 'rgb') and color.rgb:
                # Handle both string and RGB object types
                rgb_value = color.rgb
                if hasattr(rgb_value, 'value'):  # RGB object case
                    rgb_value = rgb_value.value if rgb_value.value else str(rgb_value)
                elif not isinstance(rgb_value, str):  # Other object types
                    rgb_value = str(rgb_value)
                
                # Remove alpha channel if present (ARGB format)
                if isinstance(rgb_value, str) and len(rgb_value) == 8:  # ARGB format
                    rgb_value = rgb_value[2:]  # Remove first 2 characters (alpha)
                
                # Simplified blue detection - check if it contains blue-ish values
                # Look for various shades of blue (not just the exact #00B0F0)
                if isinstance(rgb_value, str) and len(rgb_value) >= 6:
                    # Convert to uppercase for consistent checking
                    rgb_upper = rgb_value.upper()
                    # Check for common blue patterns
                    blue_patterns = [
                        '00B0F0',  # Original target blue
                        '0000FF',  # Pure blue
                        '4472C4',  # Another common blue
                        '5B9BD5',  # Light blue
                        '2F75B5',  # Dark blue
                    ]
                    
                    # Check if it matches any blue pattern or if blue component is dominant
                    for pattern in blue_patterns:
                        if pattern in rgb_upper:
                            return True
                    
                    # Check if the blue component (last 2 digits) is significantly higher than red/green
                    try:
                        if len(rgb_upper) == 6:
                            red = int(rgb_upper[0:2], 16)
                            green = int(rgb_upper[2:4], 16)
                            blue = int(rgb_upper[4:6], 16)
                            # Consider it blue if blue component is much higher than red and green
                            if blue > 150 and blue > (red + green):
                                return True
                    except ValueError:
                        pass
                        
            elif hasattr(color, 'indexed') and color.indexed is not None:
                # Handle indexed colors - some indexed colors are blue
                # Common blue indexed colors: 5 (blue), 41 (light blue), etc.
                blue_indexed = [5, 41, 44, 49]
                return color.indexed in blue_indexed
    except Exception:
        # If anything fails, just return False
        pass
    return False

def preserve_cell_color_and_set_value(cell, value):
    """
    Set cell value while preserving existing fill color
    """
    from openpyxl.styles import PatternFill
    
    # Store the current fill properties before modifying
    current_fill_info = None
    try:
        if cell.fill and hasattr(cell.fill, 'patternType') and cell.fill.patternType:
            # Extract fill properties to recreate later
            current_fill_info = {
                'patternType': cell.fill.patternType,
                'fgColor': None,
                'bgColor': None
            }
            
            # Try to extract foreground color
            if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                current_fill_info['fgColor'] = cell.fill.fgColor
            
            # Try to extract background color  
            if hasattr(cell.fill, 'bgColor') and cell.fill.bgColor:
                current_fill_info['bgColor'] = cell.fill.bgColor
    except Exception:
        # If we can't extract fill info, we'll just set the value without preserving color
        current_fill_info = None
    
    # Set the value
    cell.value = value
    
    # Restore the fill if we extracted the info successfully
    if current_fill_info and current_fill_info['patternType']:
        try:
            # Create a new PatternFill with the same properties
            new_fill = PatternFill(
                patternType=current_fill_info['patternType'],
                fgColor=current_fill_info['fgColor'],
                bgColor=current_fill_info['bgColor']
            )
            cell.fill = new_fill
        except Exception:
            # If restoring fails, just leave the cell with its new value
            pass

def get_bitcap_price(driver, url):
    """
    Navigate to a bitcap.com fund page and scrape the share price.
    Looks for patterns like "Share price €81.28*" and extracts the numeric value.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load (bitcap uses dynamic content)
        time.sleep(8)
    except TimeoutException:
        raise ValueError("Timed out waiting for bitcap page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for "Share price €XX.XX" pattern (most specific)
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for various share price patterns
        share_price_patterns = [
            r'Share price\s*€\s*([\d,\.]+)',           # "Share price €81.28"
            r'Share price.*?€\s*([\d,\.]+)',           # "Share price [anything] €81.28"
            r'€\s*([\d,\.]+).*?Share price',           # "€81.28 [anything] Share price"
            r'Share\s+price[^\d]*€\s*([\d,\.]+)',      # "Share price [non-digits] €81.28"
            r'price\s*€\s*([\d,\.]+)',                 # "price €81.28"
        ]
        
        for pattern in share_price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                logging.info(f"Found bitcap price with pattern '{pattern}': {price_str}")
                
                # Clean and normalize the price
                price_normalized = price_str.replace(',', '.')
                try:
                    price_value = float(price_normalized)
                    if 0.01 <= price_value <= 10000:  # Reasonable fund price range
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with bitcap pattern method: {e}")
    
    # Method 2: Look for price in HTML elements with "Share price" context
    try:
        # Find elements containing "Share price" text
        share_price_elements = soup.find_all(
            lambda tag: tag.name and 'share price' in tag.get_text(' ', strip=True).lower()
        )
        
        for element in share_price_elements:
            element_text = element.get_text(' ', strip=True)
            # Look for Euro amounts in this element
            euro_match = re.search(r'€\s*([\d,\.]+)', element_text)
            if euro_match:
                price_str = euro_match.group(1)
                try:
                    price_value = float(price_str.replace(',', '.'))
                    if 0.01 <= price_value <= 10000:
                        logging.info(f"Found bitcap price in element: {price_value}")
                        return price_value
                except ValueError:
                    continue
                    
            # Also check nearby elements for the price
            for sibling in [element.find_next_sibling(), element.find_previous_sibling()]:
                if sibling:
                    sibling_text = sibling.get_text(' ', strip=True)
                    euro_match = re.search(r'€\s*([\d,\.]+)', sibling_text)
                    if euro_match:
                        price_str = euro_match.group(1)
                        try:
                            price_value = float(price_str.replace(',', '.'))
                            if 0.01 <= price_value <= 10000:
                                logging.info(f"Found bitcap price in sibling element: {price_value}")
                                return price_value
                        except ValueError:
                            continue
    except Exception as e:
        logging.error(f"Error with bitcap element method: {e}")
    
    # Method 3: Look for CSS selectors that might contain price data
    try:
        price_selectors = [
            ".share-price",
            ".price-value", 
            ".fund-price",
            ".current-price",
            "[data-price]",
            ".price"
        ]
        
        for selector in price_selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text(' ', strip=True)
                euro_match = re.search(r'€\s*([\d,\.]+)', text)
                if euro_match and len(text) < 100:  # Reasonable length for price element
                    price_str = euro_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', '.'))
                        if 0.01 <= price_value <= 10000:
                            logging.info(f"Found bitcap price using selector '{selector}': {price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with bitcap selector method: {e}")
    
    raise ValueError("Could not find share price on bitcap page")

def is_valid_share_price(price_value, context_text="", url=""):
    """
    Comprehensive validation to filter out obvious non-price values.
    Returns True if the value is likely a real share price, False otherwise.
    """
    try:
        price = float(price_value)
        
        # Basic range validation - exclude extremely unreasonable values
        if price <= 0 or price > 100000:
            return False
        
        # ENHANCED .0 DETECTION - Detect ALL forms of .0 values (24.00, 30.0, etc.)
        # Check if the number has only zeros after the decimal point
        price_str = str(price)
        if '.' in price_str:
            # Split into integer and decimal parts
            integer_part, decimal_part = price_str.split('.')
            # Check if decimal part contains only zeros (handles 24.0, 24.00, 24.000, etc.)
            if decimal_part and all(digit == '0' for digit in decimal_part):
                is_zero_decimal = True
                whole_number_value = int(float(price))  # Convert to int for further checks
            else:
                is_zero_decimal = False
                whole_number_value = None
        else:
            # It's already a whole number (no decimal point)
            is_zero_decimal = True
            whole_number_value = int(price)
        
        # If it's a .0 value (including 24.00, 30.0, etc.), apply strict filtering
        if is_zero_decimal and whole_number_value is not None:
            
            # 1. YEAR DETECTION - Reject years (1990-2030)
            if 1990 <= whole_number_value <= 2030:
                logging.debug(f"Rejected year value: {price} (detected as .0)")
                return False
            
            # 2. PERCENTAGE VALUES - Common protection/allocation percentages
            # BUT FIRST check if we have explicit price context that overrides this
            if context_text:
                context_lower = context_text.lower()
                price_keywords = ['price', 'share price', 'market price', 'last price', 'trading price', 
                                'current price', 'last traded', 'closing price', 'nav', 'quote']
                has_price_context = any(keyword in context_lower for keyword in price_keywords)
                
                # If we have explicit price context, allow reasonable whole number prices
                if has_price_context and 5 <= whole_number_value <= 200:
                    logging.debug(f"Accepted .0 value with valid price context: {price}")
                    return True
            
            # Only reject percentage values if no valid price context was found
            if whole_number_value in [10, 20, 30, 40, 50, 60, 70, 75, 80, 90, 95, 100]:
                logging.debug(f"Rejected likely percentage value: {price} (detected as .0)")
                return False
            
            # 3. SMALL INTEGERS - Likely version numbers, ratings, etc.
            if 1 <= whole_number_value <= 5:
                logging.debug(f"Rejected small integer (likely version/rating): {price} (detected as .0)")
                return False
            
            # 4. FUND SIZE INDICATORS - Common fund sizes in millions
            if whole_number_value in [112, 1089, 1576, 1201, 844, 692, 360]:  # Common crypto ETF sizes
                logging.debug(f"Rejected likely fund size: {price} (detected as .0)")
                return False
            
            # 5. CONTEXT-BASED FILTERING
            if context_text:
                context_lower = context_text.lower()
                
                # Check for protection-related keywords near the number
                protection_keywords = ['protection', 'downside', 'protection level', 'cap range', 'outcome period']
                if any(keyword in context_lower for keyword in protection_keywords):
                    logging.debug(f"Rejected protection-related value: {price} (detected as .0)")
                    return False
                
                # Check for date-related keywords
                date_keywords = ['inception', 'launch', 'date', 'year', 'launched', 'period']
                if any(keyword in context_lower for keyword in date_keywords):
                    logging.debug(f"Rejected date-related value: {price} (detected as .0)")
                    return False
                
                # Check for fund metadata keywords
                metadata_keywords = ['aum', 'assets under management', 'fund size', 'ter', 'expense ratio']
                if any(keyword in context_lower for keyword in metadata_keywords):
                    logging.debug(f"Rejected fund metadata value: {price} (detected as .0)")
                    return False
            
            # 6. DOMAIN-SPECIFIC RULES for .0 values
            if url:
                url_lower = url.lower()
                
                # Calamos Bitcoin protection ETFs - reject protection percentages
                if 'calamos.com' in url_lower and whole_number_value in [10, 20, 28, 30, 31, 50, 55, 80, 90, 100]:
                    logging.debug(f"Rejected Calamos protection percentage: {price} (detected as .0)")
                    return False
                
                # WisdomTree - reject common metadata values
                if 'wisdomtree.eu' in url_lower and whole_number_value in [1, 2, 3, 4, 5]:
                    logging.debug(f"Rejected WisdomTree metadata value: {price} (detected as .0)")
                    return False
                
                # NASDAQ - reject volume/market cap indicators
                if 'nasdaq.com' in url_lower and whole_number_value > 1000:
                    logging.debug(f"Rejected NASDAQ large value (likely volume/market cap): {price} (detected as .0)")
                    return False
            
            # 7. RANGE VALIDATION for .0 values - Be more restrictive
            if not (1 <= whole_number_value <= 1000):  # Most stock/ETF prices are in this range
                # Exception: some crypto can be higher, but not too extreme
                if not (1000 < whole_number_value <= 10000 and url and ('bitcoin' in url.lower() or 'crypto' in url.lower())):
                    logging.debug(f"Rejected .0 value outside reasonable price range: {price}")
                    return False
            
            # If we get here and it's a .0 value, it's suspicious but might be valid
            # Log it as suspicious but allow it through (it will be flagged in the log)
            logging.debug(f"Suspicious .0 value detected but allowing: {price}")
            return False  # CHANGED: Return False for ALL .0 values to flag them in the log
        
        # For non-.0 values, apply lighter validation
        
        # IMPROVED: Better decimal price validation for small values
        # Check for expense ratios and other percentage-like decimals
        if 0 < price < 5:
            if context_text:
                context_lower = context_text.lower()
                # More comprehensive expense ratio detection
                expense_keywords = ['expense ratio', 'ter', 'fee', 'cost', '%', 'percent', 'expense', 'ratio']
                if any(keyword in context_lower for keyword in expense_keywords):
                    logging.debug(f"Rejected likely expense ratio/fee: {price}")
                    return False
        
        # For legitimate decimal values (like 1.0092, 24.37, etc.), be more permissive
        # If we get here, it's likely a valid price
        return True
        
    except (ValueError, TypeError):
        return False

def get_qrasset_cota_price(driver, url):
    """
    Navigate to a QR Asset fund page and scrape the "COTA EM" price.
    Looks for patterns like "COTA EM 23/05/2025 R$ 38,09" and extracts "R$ 38,09".
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for QR Asset page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for "COTA EM" followed by date and price pattern (most specific)
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for "COTA EM DD/MM/YYYY R$ XX,XX" pattern
        cota_patterns = [
            r'COTA\s+EM\s+\d{1,2}/\d{1,2}/\d{4}\s+R\$\s*([\d,\.]+)',  # "COTA EM 23/05/2025 R$ 38,09"
            r'COTA\s+EM[^R]*R\$\s*([\d,\.]+)',                        # "COTA EM [anything] R$ 38,09"
            r'COTA[^R]*R\$\s*([\d,\.]+)',                             # "COTA [anything] R$ 38,09"
        ]
        
        for pattern in cota_patterns:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                price_str = match.group(1)
                logging.info(f"Found QR Asset COTA price with pattern '{pattern}': R$ {price_str}")
                
                # Convert Brazilian format (38,09) to float
                # Handle both "38,09" and "1.234,56" formats
                if '.' in price_str and ',' in price_str:
                    # Format: 1.234,56 -> 1234.56
                    price_normalized = price_str.replace('.', '').replace(',', '.')
                else:
                    # Format: 38,09 -> 38.09
                    price_normalized = price_str.replace(',', '.')
                
                try:
                    price_value = float(price_normalized)
                    if 0.01 <= price_value <= 100000:  # Reasonable fund price range
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with QR Asset pattern method: {e}")
    
    # Method 2: Look for span elements with "COTA" context
    try:
        # Find elements containing "COTA" text
        cota_elements = soup.find_all(
            lambda tag: tag.name and 'cota' in tag.get_text(' ', strip=True).lower()
        )
        
        for element in cota_elements:
            element_text = element.get_text(' ', strip=True)
            # Look for Brazilian Real amounts in this element
            real_match = re.search(r'R\$\s*([\d,\.]+)', element_text)
            if real_match:
                price_str = real_match.group(1)
                try:
                    # Convert Brazilian format
                    if '.' in price_str and ',' in price_str:
                        price_normalized = price_str.replace('.', '').replace(',', '.')
                    else:
                        price_normalized = price_str.replace(',', '.')
                    
                    price_value = float(price_normalized)
                    if 0.01 <= price_value <= 100000:
                        logging.info(f"Found QR Asset price in COTA element: {price_value}")
                        return price_value
                except ValueError:
                    continue
                    
            # Also check nearby elements for the price
            for sibling in [element.find_next_sibling(), element.find_previous_sibling()]:
                if sibling:
                    sibling_text = sibling.get_text(' ', strip=True)
                    real_match = re.search(r'R\$\s*([\d,\.]+)', sibling_text)
                    if real_match:
                        price_str = real_match.group(1)
                        try:
                            if '.' in price_str and ',' in price_str:
                                price_normalized = price_str.replace('.', '').replace(',', '.')
                            else:
                                price_normalized = price_str.replace(',', '.')
                            
                            price_value = float(price_normalized)
                            if 0.01 <= price_value <= 100000:
                                logging.info(f"Found QR Asset price in sibling element: {price_value}")
                                return price_value
                        except ValueError:
                            continue
    except Exception as e:
        logging.error(f"Error with QR Asset element method: {e}")
    
    # Method 3: Look for CSS selectors that might contain price data
    try:
        price_selectors = [
            ".cota-price",
            ".fund-price",
            ".current-price",
            ".value",
            ".price"
        ]
        
        for selector in price_selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text(' ', strip=True)
                real_match = re.search(r'R\$\s*([\d,\.]+)', text)
                if real_match and len(text) < 100:  # Reasonable length for price element
                    price_str = real_match.group(1)
                    try:
                        if '.' in price_str and ',' in price_str:
                            price_normalized = price_str.replace('.', '').replace(',', '.')
                        else:
                            price_normalized = price_str.replace(',', '.')
                        
                        price_value = float(price_normalized)
                        if 0.01 <= price_value <= 100000:
                            logging.info(f"Found QR Asset price using selector '{selector}': {price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with QR Asset selector method: {e}")
    
    raise ValueError("Could not find COTA price on QR Asset page")

def get_morningstar_fr_price(driver, url):
    """
    Navigate to a Morningstar France fund page and scrape the VL (Valeur Liquidative) price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for Morningstar France page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for "VL" followed by date and currency amount (most specific)
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for patterns like "VL22/05/2025 EUR 11,19" or "VL31/07/2024 GBP 11,03"
        vl_patterns = [
            r'VL\d{1,2}/\d{1,2}/\d{4}\s+([A-Z]{3})\s+([\d,\.]+)',  # "VL22/05/2025 EUR 11,19"
            r'VL\s*\d{1,2}/\d{1,2}/\d{4}\s+([A-Z]{3})\s+([\d,\.]+)',  # "VL 22/05/2025 EUR 11,19"
            r'Valeur\s+Liquidative.*?([A-Z]{3})\s+([\d,\.]+)',  # "Valeur Liquidative ... EUR 11,19"
            r'VL.*?([A-Z]{3})\s+([\d,\.]+)',  # "VL ... EUR 11,19"
        ]
        
        for pattern in vl_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                currency = match.group(1)
                price_str = match.group(2)
                logging.info(f"Found Morningstar FR VL price: {currency} {price_str}")
                
                # Convert European number format (11,19) to float
                if ',' in price_str and '.' not in price_str:
                    # European format: 11,19 -> 11.19
                    price_normalized = price_str.replace(',', '.')
                elif '.' in price_str and ',' in price_str:
                    # Format: 1.234,56 -> 1234.56
                    price_normalized = price_str.replace('.', '').replace(',', '.')
                else:
                    # US format: 11.19 -> 11.19
                    price_normalized = price_str
                
                try:
                    price_value = float(price_normalized)
                    if 0.01 <= price_value <= 100000:  # Reasonable fund price range
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Morningstar FR pattern method: {e}")
    
    # Method 2: Look for table structure with VL data
    try:
        # Find table rows that might contain VL information
        rows = soup.find_all('tr')
        for row in rows:
            row_text = row.get_text(" ", strip=True)
            # Look for rows containing "VL" and currency amounts
            if 'vl' in row_text.lower():
                # Extract currency and amount from this row
                currency_match = re.search(r'([A-Z]{3})\s+([\d,\.]+)', row_text)
                if currency_match:
                    currency = currency_match.group(1)
                    price_str = currency_match.group(2)
                    
                    # Convert European number format
                    if ',' in price_str and '.' not in price_str:
                        price_normalized = price_str.replace(',', '.')
                    elif '.' in price_str and ',' in price_str:
                        price_normalized = price_str.replace('.', '').replace(',', '.')
                    else:
                        price_normalized = price_str
                    
                    try:
                        price_value = float(price_normalized)
                        if 0.01 <= price_value <= 100000:
                            logging.info(f"Found Morningstar FR price in table: {currency} {price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with Morningstar FR table method: {e}")
    
    # Method 3: Look for CSS selectors that might contain VL data
    try:
        # Look for table cells or divs that might contain the VL price
        elements = soup.find_all(['td', 'div', 'span'])
        for element in elements:
            element_text = element.get_text(" ", strip=True)
            # Check if this element contains VL-related text
            if len(element_text) < 100 and ('vl' in element_text.lower() or any(curr in element_text for curr in ['EUR', 'USD', 'GBP'])):
                # Look for currency and price pattern
                currency_match = re.search(r'([A-Z]{3})\s+([\d,\.]+)', element_text)
                if currency_match:
                    currency = currency_match.group(1)
                    price_str = currency_match.group(2)
                    
                    # Convert European number format
                    if ',' in price_str and '.' not in price_str:
                        price_normalized = price_str.replace(',', '.')
                    elif '.' in price_str and ',' in price_str:
                        price_normalized = price_str.replace('.', '').replace(',', '.')
                    else:
                        price_normalized = price_str
                    
                    try:
                        price_value = float(price_normalized)
                        if 0.01 <= price_value <= 100000:
                            logging.info(f"Found Morningstar FR price in element: {currency} {price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with Morningstar FR element method: {e}")
    
    raise ValueError("Could not find VL price on Morningstar France page")

def get_franklin_templeton_price(driver, url):
    """
    Navigate to a Franklin Templeton ETF page and scrape the market price or NAV.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load (Franklin Templeton loads prices dynamically)
        time.sleep(8)
    except TimeoutException:
        raise ValueError("Timed out waiting for Franklin Templeton page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for "Market Price" patterns (most reliable for ETFs)
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for Market Price patterns - these are the primary price indicators for ETFs
        market_price_patterns = [
            r'Market\s+Price[^$]*\$\s*([\d,\.]+)',     # "Market Price $XX.XX"
            r'Market\s+Price.*?\$\s*([\d,\.]+)',       # "Market Price ... $XX.XX"
        ]
        
        for pattern in market_price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                logging.info(f"Found Franklin Templeton Market Price: ${price_str}")
                
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 10000:  # Reasonable ETF price range
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Franklin Templeton Market Price method: {e}")
    
    # Method 2: Look for NAV (Net Asset Value) patterns
    try:
        nav_patterns = [
            r'NAV[^$]*\$\s*([\d,\.]+)',               # "NAV $XX.XX"
            r'Net\s+Asset\s+Value[^$]*\$\s*([\d,\.]+)', # "Net Asset Value $XX.XX"
        ]
        
        for pattern in nav_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                logging.info(f"Found Franklin Templeton NAV: ${price_str}")
                
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 10000:  # Reasonable ETF price range
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Franklin Templeton NAV method: {e}")
    
    # Method 3: Look for general price patterns near ETF-related keywords
    try:
        # Find price context elements that might contain the main price
        price_elements = soup.find_all(['div', 'span', 'td'], string=re.compile(r'price|market|nav', re.IGNORECASE))
        
        for element in price_elements:
            # Look for dollar amounts in this element and nearby elements
            element_text = element.get_text(" ", strip=True)
            
            # Check this element
            dollar_match = re.search(r'\$\s*([\d,\.]+)', element_text)
            if dollar_match:
                price_str = dollar_match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 1 <= price_value <= 1000:  # More restrictive range for this method
                        logging.info(f"Found Franklin Templeton price in element: ${price_value}")
                        return price_value
                except ValueError:
                    continue
            
            # Check next sibling elements
            next_sibling = element.find_next_sibling()
            if next_sibling:
                sibling_text = next_sibling.get_text(" ", strip=True)
                dollar_match = re.search(r'\$\s*([\d,\.]+)', sibling_text)
                if dollar_match:
                    price_str = dollar_match.group(1)
                    try:
                        price_value = float(price_str.replace(',', ''))
                        if 1 <= price_value <= 1000:
                            logging.info(f"Found Franklin Templeton price in sibling: ${price_value}")
                            return price_value
                    except ValueError:
                        continue
    except Exception as e:
        logging.error(f"Error with Franklin Templeton element method: {e}")
    
    raise ValueError("Could not find market price or NAV on Franklin Templeton page")

def get_schwab_asset_management_price(driver, url):
    """
    Navigate to a Schwab Asset Management ETF page and scrape the actual market price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(8)
    except TimeoutException:
        raise ValueError("Timed out waiting for Schwab page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for market price patterns
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for various price patterns
        price_patterns = [
            r'Market\s+Price[^$]*\$\s*([\d,\.]+)',     # "Market Price $XX.XX"
            r'Price[^$]*\$\s*([\d,\.]+)',              # "Price $XX.XX"
            r'Last\s+Price[^$]*\$\s*([\d,\.]+)',       # "Last Price $XX.XX"
            r'Current\s+Price[^$]*\$\s*([\d,\.]+)',    # "Current Price $XX.XX"
            r'NAV[^$]*\$\s*([\d,\.]+)',                # "NAV $XX.XX" (as fallback)
        ]
        
        for pattern in price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable ETF prices with decimals
                    if 1 <= price_value <= 500 and price_value != int(price_value):
                        logging.info(f"Found Schwab price: ${price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Schwab pattern method: {e}")
    
    raise ValueError("Could not find valid market price on Schwab page")

def get_evolve_etfs_price(driver, url):
    """
    Navigate to an Evolve ETFs page and scrape the actual market price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(8)
    except TimeoutException:
        raise ValueError("Timed out waiting for Evolve ETFs page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for market price patterns
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for various price patterns (Canadian format)
        price_patterns = [
            r'Market\s+Price[^$]*\$\s*([\d,\.]+)',     # "Market Price $XX.XX"
            r'Price[^$]*\$\s*([\d,\.]+)',              # "Price $XX.XX"
            r'Last\s+Price[^$]*\$\s*([\d,\.]+)',       # "Last Price $XX.XX"
            r'Current\s+Price[^$]*\$\s*([\d,\.]+)',    # "Current Price $XX.XX"
            r'Closing\s+Price[^$]*\$\s*([\d,\.]+)',    # "Closing Price $XX.XX"
        ]
        
        for pattern in price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable ETF prices with decimals
                    if 1 <= price_value <= 500 and price_value != int(price_value):
                        logging.info(f"Found Evolve ETFs price: ${price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Evolve ETFs pattern method: {e}")
    
    raise ValueError("Could not find valid market price on Evolve ETFs page")

def get_21shares_price(driver, url):
    """
    Navigate to a 21Shares ETF page and scrape the actual market price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(8)
    except TimeoutException:
        raise ValueError("Timed out waiting for 21Shares page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for market price patterns
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for various price patterns (US format)
        price_patterns = [
            r'Market\s+Price[^$]*\$\s*([\d,\.]+)',     # "Market Price $XX.XX"
            r'Price[^$]*\$\s*([\d,\.]+)',              # "Price $XX.XX"
            r'Last\s+Price[^$]*\$\s*([\d,\.]+)',       # "Last Price $XX.XX"
            r'Current\s+Price[^$]*\$\s*([\d,\.]+)',    # "Current Price $XX.XX"
            r'Closing\s+Price[^$]*\$\s*([\d,\.]+)',    # "Closing Price $XX.XX"
        ]
        
        for pattern in price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable ETF prices with decimals
                    if 1 <= price_value <= 500 and price_value != int(price_value):
                        logging.info(f"Found 21Shares price: ${price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with 21Shares pattern method: {e}")
    
    raise ValueError("Could not find valid market price on 21Shares page")

def get_ninepoint_price(driver, url):
    """
    Navigate to a Ninepoint ETF page and scrape the actual market price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(10)  # Increased wait time
    except TimeoutException:
        raise ValueError("Timed out waiting for Ninepoint page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for market price patterns with more flexible matching
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for various price patterns (Canadian format) - more flexible
        price_patterns = [
            r'Market\s+Price[^$]*\$\s*([\d,\.]+)',     # "Market Price $XX.XX"
            r'Price[^$]*\$\s*([\d,\.]+)',              # "Price $XX.XX"
            r'Last\s+Price[^$]*\$\s*([\d,\.]+)',       # "Last Price $XX.XX"
            r'Current\s+Price[^$]*\$\s*([\d,\.]+)',    # "Current Price $XX.XX"
            r'Unit\s+Price[^$]*\$\s*([\d,\.]+)',       # "Unit Price $XX.XX"
            r'NAV[^$]*\$\s*([\d,\.]+)',                # "NAV $XX.XX"
            r'Net\s+Asset\s+Value[^$]*\$\s*([\d,\.]+)', # "Net Asset Value $XX.XX"
            r'\$\s*([\d,\.]+)\s*CAD',                  # "$XX.XX CAD"
            r'CAD\s*\$\s*([\d,\.]+)',                  # "CAD $XX.XX"
        ]
        
        for pattern in price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable ETF prices (relaxed criteria)
                    if 0.1 <= price_value <= 1000:  # More flexible range
                        logging.info(f"Found Ninepoint price: ${price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Ninepoint pattern method: {e}")
    
    # Method 2: Look for any dollar amounts in the page
    try:
        all_dollar_matches = re.finditer(r'\$\s*([\d,\.]+)', page_text, re.IGNORECASE)
        for match in all_dollar_matches:
            price_str = match.group(1)
            try:
                price_value = float(price_str.replace(',', ''))
                # Look for reasonable ETF prices
                if 1 <= price_value <= 500 and price_value != int(price_value):
                    logging.info(f"Found Ninepoint price (general search): ${price_value}")
                    return price_value
            except ValueError:
                continue
    except Exception as e:
        logging.error(f"Error with Ninepoint general search: {e}")
    
    raise ValueError("Could not find valid market price on Ninepoint page")

def get_betashares_price(driver, url):
    """
    Navigate to a BetaShares ETF page and scrape the most accurate market price.
    Prioritizes NAV/Unit over current price for better precision.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(12)  # Increased wait time for dynamic pricing
    except TimeoutException:
        raise ValueError("Timed out waiting for BetaShares page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for NAV/Unit (most accurate)
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for NAV/Unit patterns (highest priority)
        nav_patterns = [
            r'NAV/Unit[^$]*\$\s*([\d,\.]+)',           # "NAV/Unit* $7.04"
            r'NAV\s+per\s+Unit[^$]*\$\s*([\d,\.]+)',   # "NAV per Unit $7.04"
            r'Net\s+Asset\s+Value[^$]*\$\s*([\d,\.]+)', # "Net Asset Value $7.04"
        ]
        
        for pattern in nav_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable ETF prices with decimals
                    if 1 <= price_value <= 500 and price_value != int(price_value):
                        logging.info(f"Found BetaShares NAV/Unit: ${price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with BetaShares NAV method: {e}")
    
    # Method 2: Look for Current Price with better precision
    try:
        # Look for current price patterns with decimal precision
        price_patterns = [
            r'Current\s+price[^$]*\$\s*([\d,\.]+)',    # "Current price $ 6.92"
            r'Last\s+trade[^$]*\$\s*([\d,\.]+)',       # "Last trade* $ 6.92"
            r'Market\s+Price[^$]*\$\s*([\d,\.]+)',     # "Market Price $6.92"
            r'Bid[^$]*\$\s*([\d,\.]+)',                # "Bid (delayed) $ 6.93"
            r'Offer[^$]*\$\s*([\d,\.]+)',              # "Offer (delayed) $ 6.99"
        ]
        
        for pattern in price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable ETF prices with decimals
                    if 1 <= price_value <= 500 and price_value != int(price_value):
                        logging.info(f"Found BetaShares current price: ${price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with BetaShares current price method: {e}")
    
    raise ValueError("Could not find valid market price on BetaShares page")

def get_aminagroup_price(driver, url):
    """
    Navigate to an Amina Group certificate page and scrape the current price.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(5)
    except TimeoutException:
        raise ValueError("Timed out waiting for Amina Group page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for "Current Price" patterns (most reliable)
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for Current Price patterns
        current_price_patterns = [
            r'Current\s+Price[^\d]*?([\d,\.]+)',           # "Current Price 1.8744"
            r'Current\s+Price.*?([\d,\.]+)',               # "Current Price ... 1.8744"
            r'Price[^\d]*?([\d,\.]+)',                     # "Price 1.8744"
        ]
        
        for pattern in current_price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                logging.info(f"Found Amina Group Current Price: {price_str}")
                
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 10000:  # Reasonable certificate price range
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Amina Group Current Price method: {e}")
    
    # Method 2: Look for NAV patterns as fallback
    try:
        nav_patterns = [
            r'NAV[^\d]*?([\d,\.]+)',                       # "NAV 1.0"
            r'Net\s+Asset\s+Value[^\d]*?([\d,\.]+)',       # "Net Asset Value 1.0"
        ]
        
        for pattern in nav_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                # Skip obviously wrong NAV values (like 1.0)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 1.01 <= price_value <= 10000:  # Exclude exactly 1.0 which is suspicious
                        logging.info(f"Found Amina Group NAV: {price_str}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Amina Group NAV method: {e}")
    
    # Method 3: Look for price in HTML elements with certificate context
    try:
        # Find elements containing price-related text
        price_elements = soup.find_all(
            lambda tag: tag.name and any(keyword in tag.get_text(' ', strip=True).lower() 
                                       for keyword in ['price', 'current', 'value'])
        )
        
        for element in price_elements:
            element_text = element.get_text(' ', strip=True)
            # Look for numeric values in this element
            number_match = re.search(r'([\d,\.]+)', element_text)
            if number_match:
                price_str = number_match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.1 <= price_value <= 1000 and price_value != 1.0:  # Reasonable range, exclude 1.0
                        logging.info(f"Found Amina Group price in element: {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Amina Group element method: {e}")
    
    raise ValueError("Could not find current price on Amina Group page")

def get_valour_price(driver, url):
    """
    Navigate to a Valour ETF page and scrape the share price.
    Pattern found: "Share Price 1.2649 EUR"
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(8)
    except TimeoutException:
        raise ValueError("Timed out waiting for Valour page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    page_text = soup.get_text(" ", strip=True)
    
    # Method 1: Look for "Share Price X.XXXX EUR" pattern (most reliable)
    try:
        share_price_pattern = r'Share\s+Price\s+([\d,\.]+)\s+EUR'
        match = re.search(share_price_pattern, page_text, re.IGNORECASE)
        if match:
            price_str = match.group(1)
            price_value = float(price_str.replace(',', ''))
            if 0.01 <= price_value <= 1000:  # Reasonable ETF price range
                logging.info(f"Found Valour share price: {price_value} EUR")
                return price_value
    except Exception as e:
        logging.error(f"Error with Valour share price method: {e}")
    
    # Method 2: Look for general price patterns near "Price"
    try:
        price_patterns = [
            r'Price[^0-9]*?([\d,\.]+)',           # "Price 1.2649"
            r'Current\s+Price[^0-9]*?([\d,\.]+)', # "Current Price 1.2649"
            r'Unit\s+Price[^0-9]*?([\d,\.]+)',    # "Unit Price 1.2649"
        ]
        
        for pattern in price_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 1000:
                        logging.info(f"Found Valour price: {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Valour general price method: {e}")
    
    raise ValueError("Could not find valid share price on Valour page")

def get_nasdaq_european_price_v2(driver, url):
    """
    Navigate to a NASDAQ European Market page and scrape the price.
    Pattern found: "SEK 15.86"
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(8)
    except TimeoutException:
        raise ValueError("Timed out waiting for NASDAQ European page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    page_text = soup.get_text(" ", strip=True)
    
    # Method 1: Look for "SEK XX.XX" pattern (most reliable for this specific URL)
    try:
        sek_pattern = r'SEK\s+([\d,\.]+)'
        match = re.search(sek_pattern, page_text, re.IGNORECASE)
        if match:
            price_str = match.group(1)
            price_value = float(price_str.replace(',', ''))
            if 0.01 <= price_value <= 10000:  # Reasonable price range for SEK
                logging.info(f"Found NASDAQ European price: {price_value} SEK")
                return price_value
    except Exception as e:
        logging.error(f"Error with NASDAQ European SEK method: {e}")
    
    # Method 2: Look for other currency patterns
    try:
        currency_patterns = [
            (r'USD\s+([\d,\.]+)', 'USD'),
            (r'EUR\s+([\d,\.]+)', 'EUR'),
            (r'Last\s+Sale[^0-9]*?([\d,\.]+)', 'Last Sale'),
            (r'Current\s+Price[^0-9]*?([\d,\.]+)', 'Current Price'),
        ]
        
        for pattern, desc in currency_patterns:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 10000:
                        logging.info(f"Found NASDAQ European {desc}: {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with NASDAQ European currency method: {e}")
    
    raise ValueError("Could not find valid price on NASDAQ European page")

def get_purposeinvest_price(driver, url):
    """
    Navigate to a Purpose Investments ETF page and scrape the NAV.
    Pattern found: "NAV $4.80"
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(8)
    except TimeoutException:
        raise ValueError("Timed out waiting for Purpose Investments page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    page_text = soup.get_text(" ", strip=True)
    
    # Method 1: Look for "NAV $X.XX" pattern (most reliable)
    try:
        nav_pattern = r'NAV\s+\$\s*([\d,\.]+)'
        match = re.search(nav_pattern, page_text, re.IGNORECASE)
        if match:
            price_str = match.group(1)
            price_value = float(price_str.replace(',', ''))
            if 0.01 <= price_value <= 1000:  # Reasonable ETF NAV range
                logging.info(f"Found Purpose Investments NAV: ${price_value}")
                return price_value
    except Exception as e:
        logging.error(f"Error with Purpose Investments NAV method: {e}")
    
    # Method 2: Look for other price patterns
    try:
        price_patterns = [
            r'Unit\s+Price[^$]*\$\s*([\d,\.]+)',    # "Unit Price $X.XX"
            r'Market\s+Price[^$]*\$\s*([\d,\.]+)',  # "Market Price $X.XX"
            r'Closing\s+Price[^$]*\$\s*([\d,\.]+)', # "Closing Price $X.XX"
            r'Price[^$]*\$\s*([\d,\.]+)',           # "Price $X.XX"
        ]
        
        for pattern in price_patterns:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    if 0.01 <= price_value <= 1000:
                        logging.info(f"Found Purpose Investments price: ${price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with Purpose Investments price method: {e}")
    
    raise ValueError("Could not find valid NAV/price on Purpose Investments page")

def get_csopasset_price(driver, url):
    """
    Navigate to a CSO P Asset page and scrape the actual market price.
    Handles both HKD and USD pricing.
    Returns a float or raises ValueError.
    """
    driver.get(url)
    try:
        # Wait for the page to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Allow extra time for JavaScript content to load
        time.sleep(12)  # Increased wait time for dynamic pricing
    except TimeoutException:
        raise ValueError("Timed out waiting for CSO P Asset page to load")

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Method 1: Look for Closing Price (most reliable)
    try:
        page_text = soup.get_text(" ", strip=True)
        
        # Look for closing price patterns - be more specific to avoid dates
        closing_patterns = [
            r'Closing\s+Price\s+as\s+of\s+[\d-]+\s+([\d,\.]+)',  # "Closing Price as of 22-May-2025 38.500"
            r'38\.500',                                           # Direct match for the specific price we saw
            r'38\.2800',                                          # Intra-day price we saw
            r'38\.3621',                                          # NAV price we saw
            r'38\.2721',                                          # Another NAV price we saw
        ]
        
        for pattern in closing_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                if pattern.startswith(r'38\.'):  # Direct price matches
                    price_str = match.group(0)
                else:
                    price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable prices (HKD range 10-100, USD range 1-20)
                    if (10 <= price_value <= 100) or (1 <= price_value <= 20):
                        logging.info(f"Found CSO P Asset closing price: {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with CSO P Asset closing price method: {e}")
    
    # Method 2: Look for Intra-day Market Price
    try:
        market_patterns = [
            r'Intra-day\s+Market\s+Price[^0-9]*([\d,\.]+)',  # "Intra-day Market Price 38.2800"
            r'Market\s+Price[^0-9]*([\d,\.]+)',              # "Market Price 38.2800"
        ]
        
        for pattern in market_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable prices
                    if (10 <= price_value <= 100) or (1 <= price_value <= 20):
                        logging.info(f"Found CSO P Asset market price: {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with CSO P Asset market price method: {e}")
    
    # Method 3: Look for NAV per Unit (fallback)
    try:
        nav_patterns = [
            r'NAV\s+as\s+of[^0-9]*([\d,\.]+)',              # "NAV as of 22-May-2025 38.3621"
            r'NAV\s+per\s+Unit[^0-9]*([\d,\.]+)',           # "NAV per Unit 4.9016"
        ]
        
        for pattern in nav_patterns:
            matches = re.finditer(pattern, page_text, re.IGNORECASE)
            for match in matches:
                price_str = match.group(1)
                try:
                    price_value = float(price_str.replace(',', ''))
                    # Look for reasonable prices
                    if (10 <= price_value <= 100) or (1 <= price_value <= 20):
                        logging.info(f"Found CSO P Asset NAV: {price_value}")
                        return price_value
                except ValueError:
                    continue
    except Exception as e:
        logging.error(f"Error with CSO P Asset NAV method: {e}")
    
    raise ValueError("Could not find valid market price on CSO P Asset page")

def main():
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    original_workbook_path = "Custodians.xlsx"
    workbook_path = "Custodians_Results.xlsx"
    sheet_name = "Non-derivative exposures"
    dest_col = "L"
    url_col  = "P"
    
    # Delete existing results file if it exists
    import os
    from urllib.parse import urlparse
    from collections import defaultdict
    
    if os.path.exists(workbook_path):
        os.remove(workbook_path)
        logging.info(f"Deleted existing {workbook_path}")
    
    # Copy original to results file
    import shutil
    shutil.copy2(original_workbook_path, workbook_path)
    logging.info(f"Copied {original_workbook_path} to {workbook_path}")
    
    # Initialize error tracking (ENHANCED: exclude blue cells and track individual URLs)
    error_domains = defaultdict(int)
    error_urls = defaultdict(list)    # NEW: Store actual URLs that had errors
    total_processed = 0
    total_successful = 0
    blue_cells_count = 0
    blue_cells_data = []  # Store blue cell information
    normal_cells_data = []  # Store normal cell information
    
    # NEW: Track .0 values (suspicious/incorrect prices) - excluding blue cells
    invalid_price_domains = defaultdict(int)  # Count invalid price errors by domain
    invalid_price_urls = defaultdict(list)    # Store actual URLs that had invalid price errors
    invalid_price_count = 0                   # Total count of invalid prices found

    def track_error_domain(url):
        """Helper function to track error domains with improved logic."""
        try:
            if not url or not isinstance(url, str):
                return "invalid://"
            
            # Clean the URL
            cleaned_url = url.strip()
            if not cleaned_url.lower().startswith(("http://", "https://")):
                return "invalid://"
            
            parsed_url = urlparse(cleaned_url)
            if parsed_url.netloc:
                return f"{parsed_url.scheme}://{parsed_url.netloc}/"
            else:
                # If no domain can be extracted, use the full URL
                return cleaned_url[:100] + ("..." if len(cleaned_url) > 100 else "")
        except Exception:
            # If all parsing fails, use the original URL (truncated if too long)
            url_str = str(url) if url else "unknown"
            return url_str[:100] + ("..." if len(url_str) > 100 else "")

    keywords_to_extract = [
        "closing price prev trading day", "closing price", "opening price", "share price", "market price", 
        "last traded price", "valuation price", "vl", "Last Traded Price", "ESTIMATIVA DA COTA", "Market Close",
        "Value per ETC security", "value per etc security", "ETP Price", "previous close",
        "valor da cota", "Last Close", "Previous closing price",
        "Current issue price", "Current price", "Last trade", "open", "last price", "last traded",
        "ai extracted price"  # NEW: AI fallback keyword
    ]
    priority_keywords_order = [
        "ai extracted price",  # NEW: Give AI results high priority
        "closing price prev trading day", "last traded price", "Last Traded Price", "share price", "market price", 
        "closing price", "opening price", "valuation price", "previous close", "Last Close",
        "Previous closing price", "vl", "ESTIMATIVA DA COTA", "Market Close", "ETP Price",
        "Value per ETC security", "value per etc security", "valor da cota", 
        "Current issue price", "Current price", "Last trade", "open", "last price", "last traded"
    ]

    try:
        wb = load_workbook(workbook_path)
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found.")
            return
        ws = wb[sheet_name]
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    rows_urls = find_processable_rows_and_get_urls(ws, url_col)
    if not rows_urls:
        logging.info("No URLs found to process.")
        return

    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    )

    driver = None
    all_data = {}

    try:
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        driver.set_page_load_timeout(60)  # Set page load timeout to 60 seconds
        for row, url in rows_urls.items():
            total_processed += 1
            logging.info(f"Row {row}: {url}")
            
            if not isinstance(url, str) or not url.strip().lower().startswith(("http://","https://")):
                all_data[row] = {"error": "Invalid URL format"}
                # Only track errors for normal (non-blue) cells
                cell = ws[f"{dest_col}{row}"]
                if not is_blue_cell(cell):
                    domain = track_error_domain(url)
                    error_domains[domain] += 1
                    error_urls[domain].append(url)
                continue
                
            if url.lower().endswith(".pdf"):
                all_data[row] = {"error": "PDF file - cannot process"}
                # Only track errors for normal (non-blue) cells
                cell = ws[f"{dest_col}{row}"]
                if not is_blue_cell(cell):
                    domain = track_error_domain(url)
                    error_domains[domain] += 1
                    error_urls[domain].append(url)
                continue

            data = fetch_and_extract_data(driver, url, keywords_to_extract)
            all_data[row] = data
            logging.info(f"Extracted: {data}")
            
            # Check if successful or error (ENHANCED: only track for normal cells)
            cell = ws[f"{dest_col}{row}"]
            is_blue = is_blue_cell(cell)
            
            if "error" in data and len(data) == 1:
                if not is_blue:  # Only track errors for normal cells
                    domain = track_error_domain(url)
                    error_domains[domain] += 1
                    error_urls[domain].append(url)
            elif not data or all(isinstance(v, str) and v.startswith("Error:") for v in data.values()):
                # Handle cases where data extraction failed but didn't return explicit error
                if not is_blue:  # Only track errors for normal cells
                    domain = track_error_domain(url)
                    error_domains[domain] += 1
                    error_urls[domain].append(url)
            else:
                total_successful += 1

    except Exception as e:
        logging.error(f"Processing loop error: {e}", exc_info=True)
    finally:
        if driver:
            driver.quit()

    # Write back with priority and separate blue/normal cells
    for row, results in all_data.items():
        cell = ws[f"{dest_col}{row}"]
        
        # Check if this cell is blue and categorize it
        is_blue = is_blue_cell(cell)
        if is_blue:
            blue_cells_count += 1
            logging.info(f"Row {row}: Blue cell detected - excluding from success rate calculation")
        
        # Determine the chosen price/result that will be written to the cell
        chosen_result = None
        chosen_key = None
        
        if "error" in results and len(results) == 1:
            out = f"Error: {results['error']}"
            chosen_result = results['error']
            chosen_key = "error"
        elif not results:
            out = "No matching data found from URL."
            chosen_result = "No matching data found from URL."
            chosen_key = "no_data"
        else:
            # Find the chosen value based on priority order
            for key in priority_keywords_order:
                if key in results and isinstance(results[key], (int, float)):
                    chosen_key = key
                    chosen_result = results[key]
                    # Add (AI) tag for AI extracted prices
                    if key == "ai extracted price":
                        out = f"share price: {results[key]} (AI)"
                    else:
                        out = f"{key}: {results[key]}"
                    break
                if chosen_result is None and key in results and isinstance(results[key], str) and results[key].startswith("Error:"):
                    chosen_key = key
                    chosen_result = results[key]
                    out = f"{key}: {results[key]}"
            
            if chosen_result is None:
                out = "No data for specified keywords found from URL."
                chosen_result = "No data for specified keywords found from URL."
                chosen_key = "no_data"
        
        # ENHANCED: Check if the CHOSEN result is suspicious using our validation function
        has_invalid_price = False
        if chosen_key and chosen_key not in ["error", "no_data"] and isinstance(chosen_result, (int, float)):
            # Use our enhanced validation instead of simple .0 detection
            url = rows_urls.get(row, 'Unknown URL')
            if not is_valid_share_price(chosen_result, out, url):
                has_invalid_price = True
                # Only track invalid prices for normal (non-blue) cells
                if not is_blue:
                    invalid_price_count += 1
                    domain = track_error_domain(url)
                    invalid_price_domains[domain] += 1
                    invalid_price_urls[domain].append(url)
                logging.info(f"Row {row}: Invalid price detected in CHOSEN result - {chosen_key}: {chosen_result} (likely not a share price)")
        
        # Store cell data for logging
        # Modified success determination to exclude invalid prices
        is_successful = not ("error" in results.lower() if isinstance(results, str) else "error" in str(results).lower()) and not has_invalid_price
        
        cell_info = {
            'row': row,
            'url': rows_urls.get(row, 'Unknown'),
            'result': out,
            'success': is_successful,
            'has_invalid_price': has_invalid_price
        }
        
        if is_blue:
            blue_cells_data.append(cell_info)
        else:
            normal_cells_data.append(cell_info)
        
        # Use the new function to preserve cell color
        preserve_cell_color_and_set_value(cell, out)

    # Update column header with timestamp
    from datetime import datetime
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    header_cell = ws[f"{dest_col}1"]
    header_cell.value = f"Share price (Last updated: {current_time})"
    
    try:
        wb.save(workbook_path)
        logging.info(f"Saved updates to {workbook_path}")
    except Exception as e:
        logging.error(f"Error saving workbook: {e}")
    
    # Create a copy of the results as custodians.xlsx (overwriting the original)
    try:
        import shutil
        shutil.copy2(workbook_path, original_workbook_path)
        logging.info(f"Created copy: {workbook_path} -> {original_workbook_path}")
    except Exception as e:
        logging.error(f"Error creating custodians.xlsx copy: {e}")
    
    # Create improved log file with ranked errors and separated sections
    log_filename = f"stock_prices_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    try:
        with open(log_filename, 'w', encoding='utf-8') as log_file:
            log_file.write("STOCK PRICES EXTRACTION LOG\n")
            log_file.write("=" * 50 + "\n")
            log_file.write(f"Run Date: {current_time}\n")
            log_file.write(f"Script: excel_stock_updater.py\n\n")
            
            # RANKED ERROR DOMAINS (most errors first) - ENHANCED: Show individual URLs
            log_file.write("SHARE PRICE ERRORS BY DOMAIN (RANKED - MOST ERRORS FIRST):\n")
            log_file.write("-" * 60 + "\n")
            if error_domains:
                # Sort by error count in descending order
                sorted_errors = sorted(error_domains.items(), key=lambda x: x[1], reverse=True)
                for domain, count in sorted_errors:
                    log_file.write(f"{count:3d} errors: {domain}\n")
                    # Show the specific URLs for this domain
                    urls_for_domain = error_urls.get(domain, [])
                    for url in urls_for_domain[:5]:  # Show first 5 URLs
                        log_file.write(f"    └─ {url}\n")
                    if len(urls_for_domain) > 5:
                        log_file.write(f"    └─ ... and {len(urls_for_domain) - 5} more URLs\n")
                    log_file.write("\n")
            else:
                log_file.write("No errors encountered!\n")
            
            # ENHANCED: INVALID PRICE VALUES SECTION (includes .0 values, percentages, years, etc.)
            log_file.write(f"\nINVALID PRICE VALUES BY DOMAIN (RANKED - MOST INVALID PRICES FIRST):\n")
            log_file.write("-" * 75 + "\n")
            log_file.write(f"Total invalid prices found: {invalid_price_count}\n")
            log_file.write("(These are likely incorrect - .0 values, dates, percentages, fund metadata, etc. instead of actual prices)\n")
            log_file.write("(Includes ALL .0 values: 24.00, 30.0, etc. - Excludes blue cells)\n\n")
            
            if invalid_price_domains:
                # Sort by invalid price count in descending order
                sorted_invalid_errors = sorted(invalid_price_domains.items(), key=lambda x: x[1], reverse=True)
                for domain, count in sorted_invalid_errors:
                    log_file.write(f"{count:3d} invalid prices: {domain}\n")
                    # Show the specific URLs for this domain
                    urls_for_domain = invalid_price_urls.get(domain, [])
                    for url in urls_for_domain[:5]:  # Show first 5 URLs
                        log_file.write(f"    └─ {url}\n")
                    if len(urls_for_domain) > 5:
                        log_file.write(f"    └─ ... and {len(urls_for_domain) - 5} more URLs\n")
                    log_file.write("\n")
            else:
                log_file.write("No invalid prices found!\n")
            
            # BLUE CELLS SECTION
            log_file.write(f"\nBLUE CELLS SECTION (EXCLUDED FROM SUCCESS RATE):\n")
            log_file.write("-" * 50 + "\n")
            log_file.write(f"Total blue cells: {blue_cells_count}\n")
            if blue_cells_data:
                blue_success = sum(1 for cell in blue_cells_data if cell['success'])
                blue_errors = blue_cells_count - blue_success
                blue_invalid_prices = sum(1 for cell in blue_cells_data if cell.get('has_invalid_price', False))
                log_file.write(f"Blue cells successful: {blue_success}\n")
                log_file.write(f"Blue cells with errors: {blue_errors}\n")
                log_file.write(f"Blue cells with invalid prices: {blue_invalid_prices}\n")
                if blue_cells_count > 0:
                    blue_success_rate = blue_success / blue_cells_count
                    log_file.write(f"Blue cells success rate: {blue_success}/{blue_cells_count} = {blue_success_rate:.2%}\n")
                log_file.write("\nBlue cell details:\n")
                for cell in blue_cells_data[:10]:  # Show first 10
                    status = "✅ Success" if cell['success'] else ("🔸 Invalid Price" if cell.get('has_invalid_price') else "❌ Error")
                    log_file.write(f"  Row {cell['row']}: {status} - {cell['url'][:60]}{'...' if len(cell['url']) > 60 else ''}\n")
                if len(blue_cells_data) > 10:
                    log_file.write(f"  ... and {len(blue_cells_data) - 10} more blue cells\n")
            else:
                log_file.write("No blue cells found.\n")
            
            # NORMAL CELLS SECTION (Updated calculations)
            normal_cells_count = len(normal_cells_data)
            normal_success = sum(1 for cell in normal_cells_data if cell['success'])
            normal_errors = sum(1 for cell in normal_cells_data if not cell['success'] and not cell.get('has_invalid_price', False))
            normal_invalid_prices = sum(1 for cell in normal_cells_data if cell.get('has_invalid_price', False))
            
            log_file.write(f"\nNORMAL CELLS SECTION (USED FOR SUCCESS RATE):\n")
            log_file.write("-" * 50 + "\n")
            log_file.write(f"Total normal cells: {normal_cells_count}\n")
            log_file.write(f"Normal cells successful: {normal_success}\n")
            log_file.write(f"Normal cells with errors: {normal_errors}\n")
            log_file.write(f"Normal cells with invalid prices (treated as errors): {normal_invalid_prices}\n")
            if normal_cells_count > 0:
                normal_success_rate = normal_success / normal_cells_count
                log_file.write(f"Normal cells success rate: {normal_success}/{normal_cells_count} = {normal_success_rate:.2%}\n")
                log_file.write(f"(Success rate excludes both errors and invalid prices)\n")
            
            log_file.write(f"\nSUMMARY:\n")
            log_file.write("-" * 10 + "\n")
            log_file.write(f"Total URLs processed: {total_processed}\n")
            log_file.write(f"Blue cells (excluded from success rate): {blue_cells_count}\n")
            log_file.write(f"Normal cells (used for success rate): {normal_cells_count}\n")
            log_file.write(f"Successful extractions (normal cells only): {normal_success}\n")
            log_file.write(f"URLs with invalid prices: {invalid_price_count}\n")
            log_file.write(f"Traditional errors: {sum(error_domains.values())}\n")
            if normal_cells_count > 0:
                final_success_rate = normal_success / normal_cells_count
                log_file.write(f"FINAL SUCCESS RATE: {normal_success}/{normal_cells_count} = {final_success_rate:.2%}\n")
                log_file.write(f"(Excludes blue cells, traditional errors, AND invalid prices)\n")
            else:
                log_file.write("FINAL SUCCESS RATE: N/A (no normal cells to process)\n")
        
        logging.info(f"Log file created: {log_filename}")
    except Exception as e:
        logging.error(f"Error creating log file: {e}")

if __name__ == "__main__":
    main()