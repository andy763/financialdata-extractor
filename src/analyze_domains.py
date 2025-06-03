from openpyxl import load_workbook
import re
from urllib.parse import urlparse
from collections import defaultdict

def analyze_domains():
    """Analyze domains in the Excel file to identify patterns for custom functions"""
    
    # Load the workbook
    wb = load_workbook('Custodians.xlsx')
    ws = wb.active if 'Shares' not in wb.sheetnames else wb['Shares']
    
    # Extract URLs and group by domain
    domain_urls = defaultdict(list)
    error_domains = set()
    
    # Read the error log to identify failing domains
    try:
        with open('outstanding_shares_log_20250524_182204.txt', 'r') as f:
            content = f.read()
            for line in content.split('\n'):
                if ' errors: ' in line:
                    domain = line.split(' errors: ')[0].strip()
                    error_domains.add(domain)
    except:
        print("Could not read error log")
    
    # Extract all URLs from the Excel file
    for row in range(2, ws.max_row + 1):
        url_cell = ws[f'P{row}']
        if url_cell.value and isinstance(url_cell.value, str):
            url = url_cell.value.strip()
            if url.startswith(('http://', 'https://')):
                try:
                    parsed = urlparse(url)
                    domain = f'{parsed.scheme}://{parsed.netloc}/'
                    domain_urls[domain].append(url)
                except:
                    pass
    
    print('DOMAIN ANALYSIS FOR CUSTOM FUNCTIONS')
    print('=' * 50)
    
    # Find domains with multiple URLs that also have errors
    priority_domains = []
    for domain, urls in domain_urls.items():
        if len(urls) > 1 and domain in error_domains:
            priority_domains.append((domain, urls, len(urls)))
    
    # Sort by number of URLs (most URLs first)
    priority_domains.sort(key=lambda x: x[2], reverse=True)
    
    print(f'\nHIGH PRIORITY DOMAINS (Multiple URLs + Errors):')
    print('-' * 50)
    
    for domain, urls, count in priority_domains[:15]:  # Top 15
        print(f'\n{domain} ({count} URLs):')
        for url in urls[:5]:  # Show first 5 URLs
            print(f'  {url}')
        if count > 5:
            print(f'  ... and {count - 5} more')
    
    print(f'\n\nALL DOMAINS WITH MULTIPLE URLs:')
    print('-' * 30)
    
    multi_url_domains = [(domain, urls) for domain, urls in domain_urls.items() if len(urls) > 1]
    multi_url_domains.sort(key=lambda x: len(x[1]), reverse=True)
    
    for domain, urls in multi_url_domains[:20]:  # Top 20
        error_status = "❌ HAS ERRORS" if domain in error_domains else "✅ No errors"
        print(f'{domain} ({len(urls)} URLs) - {error_status}')

if __name__ == "__main__":
    analyze_domains() 