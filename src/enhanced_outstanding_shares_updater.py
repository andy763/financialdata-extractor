"""
Enhanced Outstanding Shares Updater with Custom Domain Extractors
This version includes custom extractors for high-error domains and improved reporting
"""

import re
import logging
import time
import traceback
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from .sixgroup_shares_extractor import extract_sixgroup_shares
from .custom_domain_extractors import extract_with_custom_function

# Import Groq AI functionality
try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False
    logging.warning("Groq library not available. AI fallback will be disabled.")

# Groq AI Configuration
GROQ_API_KEY = "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"
GROQ_MODELS = [
    "llama-3.3-70b-versatile",
    "llama-3.1-8b-instant",
    "gemma2-9b-it",
    "mixtral-8x7b-32768",
]

def make_groq_request_with_fallback(client, messages, max_retries=3):
    """Make a Groq API request with model fallback and rate limit handling"""
    import time
    import random
    
    for model_index, model in enumerate(GROQ_MODELS):
        for attempt in range(max_retries):
            try:
                logging.info(f"Trying Groq model: {model} (attempt {attempt + 1})")
                
                if model in ["llama-3.1-8b-instant", "gemma2-9b-it"]:
                    max_tokens = 50
                else:
                    max_tokens = 100
                
                response = client.chat.completions.create(
                    messages=messages,
                    model=model,
                    temperature=0.1,
                    max_tokens=max_tokens,
                )
                
                logging.info(f"✅ Success with {model}")
                return response, model
                
            except Exception as e:
                error_str = str(e).lower()
                
                if "rate limit" in error_str or "429" in error_str:
                    if attempt < max_retries - 1:
                        wait_time = (2 ** attempt) + random.uniform(0, 1)
                        logging.warning(f"Rate limit hit with {model}, waiting {wait_time:.1f}s before retry...")
                        time.sleep(wait_time)
                        continue
                    else:
                        logging.warning(f"Rate limit exhausted for {model}, trying next model...")
                        break
                
                elif any(err in error_str for err in ["unavailable", "overloaded", "timeout"]):
                    logging.warning(f"Model {model} unavailable: {e}")
                    break
                
                else:
                    logging.error(f"Unknown error with {model}: {e}")
                    break
    
    raise Exception("All Groq models failed or rate limited")

def clean_html_for_shares_ai(html_content, max_length=5000):
    """Clean and truncate HTML content for AI analysis"""
    try:
        soup = BeautifulSoup(html_content, "html.parser")
        
        for script in soup(["script", "style", "nav", "header", "footer"]):
            script.decompose()
        
        text = soup.get_text()
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = ' '.join(chunk for chunk in chunks if chunk)
        
        shares_keywords = [
            'outstanding', 'shares', 'units', 'issued', 'circulation',
            'share count', 'total shares', 'number of shares', 'shares issued',
            'shares in issue', 'number in issue', 'issued shares',
            'million', 'billion', 'thousand'
        ]
        
        sentences = text.split('.')
        shares_sentences = []
        other_sentences = []
        
        for sentence in sentences:
            if any(keyword in sentence.lower() for keyword in shares_keywords):
                shares_sentences.append(sentence.strip())
            else:
                other_sentences.append(sentence.strip())
        
        prioritized_text = '. '.join(shares_sentences)
        if len(prioritized_text) < max_length:
            remaining_space = max_length - len(prioritized_text)
            additional_text = '. '.join(other_sentences)[:remaining_space]
            prioritized_text = prioritized_text + '. ' + additional_text
        
        return prioritized_text[:max_length]
    
    except Exception as e:
        logging.error(f"Error cleaning HTML for shares AI: {e}")
        soup = BeautifulSoup(html_content, "html.parser")
        return soup.get_text()[:max_length]

def analyze_shares_with_groq(html_content, url):
    """Use Groq AI to analyze website content and extract outstanding shares"""
    if not GROQ_AVAILABLE:
        return {"error": "Groq AI not available"}
    
    try:
        client = Groq(api_key=GROQ_API_KEY)
        cleaned_content = clean_html_for_shares_ai(html_content, max_length=4000)
        
        prompt = f"""Extract outstanding shares from this webpage: {url}

RULES:
1. Find: outstanding shares, shares outstanding, total shares, shares issued
2. Extract NUMBER ONLY (no text)
3. Convert: "150.5 million" = "150500000" (full number)
4. If no shares found: "NO_SHARES_FOUND"

Content: {cleaned_content[:3000]}

Shares:"""

        messages = [{"role": "user", "content": prompt}]
        chat_completion, used_model = make_groq_request_with_fallback(client, messages)
        
        ai_response = chat_completion.choices[0].message.content.strip()
        logging.info(f"Groq AI shares response from {used_model} for {url}: {ai_response}")
        
        if "NO_SHARES_FOUND" in ai_response.upper():
            return {"error": "AI could not find valid outstanding shares"}
        
        number_match = re.search(r'(\d{1,15}(?:[,\.\s]\d{1,3})*)', ai_response)
        if number_match:
            shares_str = number_match.group(1)
            normalized = shares_str.replace(',', '').replace(' ', '').replace('.', '')
            
            try:
                shares_value = int(normalized)
                if 1000 <= shares_value <= 100000000000:
                    return {"outstanding_shares": f"{shares_value:,} (AI)"}
            except ValueError:
                pass
        
        return {"error": "AI response could not be parsed into valid shares number"}
        
    except Exception as e:
        logging.error(f"Groq AI analysis error: {e}")
        return {"error": f"AI analysis failed: {str(e)}"}

def try_shares_ai_fallback(driver, url):
    """Try AI fallback for shares extraction"""
    try:
        logging.info(f"Attempting AI fallback for shares extraction: {url}")
        page_source = driver.page_source
        
        if not page_source or len(page_source) < 100:
            return {"error": "Page source too short for AI analysis"}
        
        return analyze_shares_with_groq(page_source, url)
        
    except Exception as e:
        logging.error(f"AI fallback error for {url}: {e}")
        return {"error": f"AI fallback failed: {str(e)}"}

def find_processable_rows_and_get_urls(ws, url_column_letter):
    """Find all rows with URLs to process"""
    processable_rows_urls = {}
    
    for row in range(2, ws.max_row + 1):
        url_cell = ws[f"{url_column_letter}{row}"]
        if url_cell.value and isinstance(url_cell.value, str):
            url = url_cell.value.strip()
            if url.startswith(("http://", "https://")):
                processable_rows_urls[row] = url
    
    return processable_rows_urls

def extract_outstanding_shares(driver, url):
    """Traditional outstanding shares extraction method"""
    try:
        logging.info(f"Extracting outstanding shares from: {url}")
        driver.get(url)
        time.sleep(2)
        
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Traditional extraction logic here (simplified for this example)
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        text = soup.get_text()
        
        patterns = [
            r'outstanding\s+shares?[:\s]+(\d{1,15}(?:[,\.\s]\d{1,3})*)',
            r'shares?\s+outstanding[:\s]+(\d{1,15}(?:[,\.\s]\d{1,3})*)',
            r'total\s+shares?[:\s]+(\d{1,15}(?:[,\.\s]\d{1,3})*)',
        ]
        
        for pattern in patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                shares_str = match.group(1)
                normalized = shares_str.replace(',', '').replace(' ', '').replace('.', '')
                try:
                    shares_value = int(normalized)
                    if 1000 <= shares_value <= 100000000000:
                        return {"outstanding_shares": f"{shares_value:,}"}
                except ValueError:
                    continue
        
        return {"error": "Could not find outstanding shares using traditional methods"}
        
    except Exception as e:
        logging.error(f"Traditional extraction error: {e}")
        return {"error": f"Traditional extraction failed: {str(e)}"}

def extract_outstanding_shares_with_custom_and_ai_fallback(driver, url):
    """
    Enhanced extraction with custom domain extractors and AI fallback
    """
    extraction_method = "Unknown"
    
    try:
        # Step 1: Try custom domain-specific extractors first
        custom_result = extract_with_custom_function(driver, url)
        if "outstanding_shares" in custom_result:
            logging.info(f"✅ Custom extractor succeeded for {url}")
            extraction_method = "Custom"
            return {**custom_result, "method": extraction_method}
        
        # Step 2: Check for SIX Group URLs
        if "six-group.com" in url.lower():
            logging.info("Detected SIX Group URL, using specialized extractor")
            six_result = extract_sixgroup_shares(driver, url)
            if "outstanding_shares" in six_result:
                extraction_method = "SIX Group"
                return {**six_result, "method": extraction_method}
        
        # Step 3: Try traditional extraction
        traditional_result = extract_outstanding_shares(driver, url)
        if "outstanding_shares" in traditional_result:
            extraction_method = "Traditional"
            return {**traditional_result, "method": extraction_method}
        
        # Step 4: Try AI fallback
        logging.info(f"All standard methods failed for {url}, trying AI fallback...")
        ai_result = try_shares_ai_fallback(driver, url)
        if "outstanding_shares" in ai_result:
            extraction_method = "AI"
            return {**ai_result, "method": extraction_method}
        
        # All methods failed
        return {"error": "All extraction methods failed", "method": "None"}
        
    except Exception as e:
        logging.error(f"Extraction error for {url}: {e}")
        return {"error": f"Extraction exception: {str(e)}", "method": "Exception"}

def is_blue_cell(cell):
    """Check if a cell has blue-ish color"""
    try:
        if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
            color = cell.fill.fgColor
            if hasattr(color, 'rgb') and color.rgb:
                rgb_value = color.rgb
                if hasattr(rgb_value, 'value'):
                    rgb_value = rgb_value.value if rgb_value.value else str(rgb_value)
                elif not isinstance(rgb_value, str):
                    rgb_value = str(rgb_value)
                
                if isinstance(rgb_value, str) and len(rgb_value) == 8:
                    rgb_value = rgb_value[2:]
                
                if isinstance(rgb_value, str) and len(rgb_value) >= 6:
                    rgb_upper = rgb_value.upper()
                    blue_patterns = ['00B0F0', '0000FF', '4472C4', '5B9BD5', '2F75B5']
                    
                    for pattern in blue_patterns:
                        if pattern in rgb_upper:
                            return True
                    
                    try:
                        if len(rgb_upper) == 6:
                            red = int(rgb_upper[0:2], 16)
                            green = int(rgb_upper[2:4], 16)
                            blue = int(rgb_upper[4:6], 16)
                            if blue > 150 and blue > (red + green):
                                return True
                    except ValueError:
                        pass
                        
            elif hasattr(color, 'indexed') and color.indexed is not None:
                blue_indexed = [5, 41, 44, 49]
                return color.indexed in blue_indexed
    except Exception:
        pass
    return False

def preserve_cell_color_and_set_value(cell, value):
    """Set cell value while preserving existing fill color"""
    from openpyxl.styles import PatternFill
    
    current_fill_info = None
    try:
        if cell.fill and hasattr(cell.fill, 'patternType') and cell.fill.patternType:
            current_fill_info = {
                'patternType': cell.fill.patternType,
                'fgColor': None,
                'bgColor': None
            }
            
            if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                current_fill_info['fgColor'] = cell.fill.fgColor
            
            if hasattr(cell.fill, 'bgColor') and cell.fill.bgColor:
                current_fill_info['bgColor'] = cell.fill.bgColor
    except Exception:
        current_fill_info = None
    
    cell.value = value
    
    if current_fill_info and current_fill_info['patternType']:
        try:
            new_fill = PatternFill(
                patternType=current_fill_info['patternType'],
                fgColor=current_fill_info['fgColor'],
                bgColor=current_fill_info['bgColor']
            )
            cell.fill = new_fill
        except Exception:
            pass

def main():
    """Enhanced main function with custom extractors and better reporting"""
    print("Enhanced Outstanding Shares Updater Starting...")
    print("Features: Custom Domain Extractors + AI Fallback + Improved Reporting")
    
    driver = None
    original_filename = "Custodians.xlsx"
    results_filename = "Custodians_Results.xlsx"
    
    # Enhanced tracking
    from urllib.parse import urlparse
    from collections import defaultdict
    from datetime import datetime
    
    error_domains = defaultdict(int)
    method_stats = defaultdict(int)
    total_processed = 0
    total_successful = 0
    blue_cells_count = 0
    
    def track_error_domain(url):
        """Helper function to track error domains"""
        try:
            if not url or not isinstance(url, str):
                return "invalid://"
            
            cleaned_url = url.strip()
            if not cleaned_url.lower().startswith(("http://", "https://")):
                return "invalid://"
            
            parsed_url = urlparse(cleaned_url)
            if parsed_url.netloc:
                return f"{parsed_url.scheme}://{parsed_url.netloc}/"
            else:
                return cleaned_url[:100] + ("..." if len(cleaned_url) > 100 else "")
        except Exception:
            url_str = str(url) if url else "unknown"
            return url_str[:100] + ("..." if len(url_str) > 100 else "")
    
    # File management
    if not os.path.exists(results_filename):
        import shutil
        shutil.copy2(original_filename, results_filename)
        print(f"Copied {original_filename} to {results_filename}")
    else:
        print(f"Using existing {results_filename} (keeping previous results)")
    
    try:
        if not os.path.exists(results_filename):
            print(f"Error: File {results_filename} not found")
            return
        
        try:
            wb = load_workbook(results_filename)
            
            if "Shares" in wb.sheetnames:
                ws = wb["Shares"]
            else:
                ws = wb.active
                print(f"Worksheet 'Shares' not found, using '{ws.title}' instead")
        except PermissionError:
            print(f"Error: Cannot open {results_filename}. The file may be open in Excel or another program.")
            return
        except Exception as e:
            print(f"Error opening Excel file: {e}")
            return
            
        # Find processable rows
        processable_rows_urls = find_processable_rows_and_get_urls(ws, "P")
        print(f"Found {len(processable_rows_urls)} rows to process")
        
        # Set up WebDriver
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920x1080")
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs)
        
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        
        # Process each URL
        for row_idx, primary_url in processable_rows_urls.items():
            total_processed += 1
            print(f"Processing row {row_idx}: {primary_url}")
            shares_data = None
            used_url = primary_url
            
            try:
                # Try primary URL with enhanced extraction
                shares_data = extract_outstanding_shares_with_custom_and_ai_fallback(driver, primary_url)
                
                # If primary URL failed, try fallback URL from column Q
                if "error" in shares_data:
                    print(f"  Primary URL failed: {shares_data['error']}")
                    fallback_url = ws[f"Q{row_idx}"].value
                    
                    if fallback_url and isinstance(fallback_url, str) and fallback_url.startswith(("http://", "https://")):
                        print(f"  Trying fallback URL from column Q: {fallback_url}")
                        used_url = fallback_url
                        shares_data = extract_outstanding_shares_with_custom_and_ai_fallback(driver, fallback_url)
                        
            except Exception as e:
                print(f"  Error processing URLs for row {row_idx}: {str(e)}")
                traceback.print_exc()
                shares_data = {"error": f"Exception: {str(e)}", "method": "Exception"}
                
                # Try to reset the browser
                try:
                    print("  Attempting to reset the browser...")
                    driver.quit()
                    driver = webdriver.Chrome(
                        service=ChromeService(ChromeDriverManager().install()),
                        options=options
                    )
                except Exception as reset_error:
                    print(f"  Failed to reset browser: {str(reset_error)}")
            
            # Track statistics
            if "error" in shares_data:
                error_domains[track_error_domain(used_url)] += 1
            else:
                total_successful += 1
                method_used = shares_data.get("method", "Unknown")
                method_stats[method_used] += 1
            
            # Write result to Excel
            shares_cell = f"M{row_idx}"
            cell = ws[shares_cell]
            
            if is_blue_cell(cell):
                blue_cells_count += 1
                print(f"  Row {row_idx}: Blue cell detected - excluding from success rate calculation")
            
            if "error" in shares_data:
                print(f"  Error: {shares_data['error']}")
                preserve_cell_color_and_set_value(cell, f"Error: {shares_data['error']}")
            else:
                method_used = shares_data.get("method", "Unknown")
                print(f"  Found outstanding shares: {shares_data['outstanding_shares']} (using {used_url}) [Method: {method_used}]")
                preserve_cell_color_and_set_value(cell, shares_data['outstanding_shares'])
            
            # Save after each update
            try:
                wb.save(results_filename)
            except PermissionError:
                print(f"  Warning: Could not save to {results_filename}, it may be open in another program")
                alt_filename = f"Custodians_Results_{int(time.time())}.xlsx"
                print(f"  Attempting to save to alternative file: {alt_filename}")
                try:
                    wb.save(alt_filename)
                    results_filename = alt_filename
                    print(f"  Successfully saved to {alt_filename}")
                except Exception as e2:
                    print(f"  Could not save Excel file: {e2}")
            except Exception as e:
                print(f"  Error saving workbook: {e}")
        
        # Update column header with timestamp
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        header_cell = ws["M1"]
        header_cell.value = f"Outstanding shares (Last updated: {current_time})"
        
        # Save final results
        try:
            wb.save(results_filename)
            print(f"Processing completed and saved to {results_filename}")
        except Exception as e:
            print(f"Error saving final results: {e}")
        
        # Create enhanced log file
        log_filename = f"enhanced_outstanding_shares_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        try:
            with open(log_filename, 'w', encoding='utf-8') as log_file:
                log_file.write("ENHANCED OUTSTANDING SHARES EXTRACTION LOG\n")
                log_file.write("=" * 60 + "\n")
                log_file.write(f"Run Date: {current_time}\n")
                log_file.write(f"Script: enhanced_outstanding_shares_updater.py\n")
                log_file.write(f"Features: Custom Domain Extractors + AI Fallback\n\n")
                
                log_file.write("EXTRACTION METHOD STATISTICS:\n")
                log_file.write("-" * 40 + "\n")
                if method_stats:
                    for method, count in sorted(method_stats.items()):
                        log_file.write(f"{method} method: {count} successful extractions\n")
                else:
                    log_file.write("No successful extractions by method!\n")
                
                log_file.write(f"\nOUTSTANDING SHARES ERRORS BY DOMAIN:\n")
                log_file.write("-" * 40 + "\n")
                if error_domains:
                    for domain, count in sorted(error_domains.items()):
                        log_file.write(f"{domain} errors: {count}\n")
                else:
                    log_file.write("No errors encountered!\n")
                
                log_file.write(f"\nSUMMARY:\n")
                log_file.write("-" * 10 + "\n")
                log_file.write(f"Total URLs processed: {total_processed}\n")
                log_file.write(f"Blue cells (excluded from success rate): {blue_cells_count}\n")
                log_file.write(f"Successful extractions: {total_successful}\n")
                log_file.write(f"Effective URLs for success calculation: {total_processed - blue_cells_count}\n")
                if (total_processed - blue_cells_count) > 0:
                    success_rate = total_successful / (total_processed - blue_cells_count)
                    log_file.write(f"Adjusted success rate: {total_successful}/{total_processed - blue_cells_count} = {success_rate:.2%}\n")
                else:
                    log_file.write("Adjusted success rate: N/A (no valid URLs to process)\n")
                
            print(f"Enhanced log file created: {log_filename}")
        except Exception as e:
            print(f"Error creating log file: {e}")
        
    except Exception as e:
        print(f"Unexpected error: {e}")
        traceback.print_exc()
    finally:
        if driver:
            driver.quit()

if __name__ == "__main__":
    main() 