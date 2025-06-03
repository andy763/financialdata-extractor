from openpyxl import load_workbook
import logging
from pathlib import Path
import datetime

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Color mapping as specified
COLOUR_MAP = {
    # exact ARGB strings (keep the leading 'FF')
    "FF00B0F0": "no_data",        # bright-blue preset
    "FFFF0000": "error",          # red
    # theme-index tuples
    ("theme", 5): "dot_zero_error",   # orange
    ("theme", 6): "ai_wrong",         # dark green
    ("theme", 7): "no_data",          # blue (accent 4)
    ("theme", 8): "wrong_figure",     # purple
}

def classify_cell(cell):
    """Classify a cell based on its fill color."""
    fg = cell.fill.fgColor
    if fg.type == "rgb":
        return COLOUR_MAP.get(fg.rgb.upper(), "uncoloured")
    elif fg.type == "theme":
        return COLOUR_MAP.get(("theme", fg.theme), "uncoloured")
    return "uncoloured"

def analyze_excel_file(file_path):
    """Analyze the Excel file and calculate success rates based on cell colors."""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Initialize counters
        stats = {
            "total_rows": 0,
            "blue_cells": 0,  # cells to exclude from calculation
            "purple_cells": 0,  # wrong figure
            "red_cells": 0,  # formula/parse error
            "orange_cells": 0,  # .0 rounding error
            "dark_green_cells": 0,  # AI guess wrong
            "uncoloured_cells": 0,  # correct/white cells
        }
        
        # Analyze column L (skipping header row)
        for cell in ws["L"][1:]:
            if cell.value is not None:  # Only count cells with values
                stats["total_rows"] += 1
                status = classify_cell(cell)
                
                if status == "no_data":
                    stats["blue_cells"] += 1
                elif status == "wrong_figure":
                    stats["purple_cells"] += 1
                elif status == "error":
                    stats["red_cells"] += 1
                elif status == "dot_zero_error":
                    stats["orange_cells"] += 1
                elif status == "ai_wrong":
                    stats["dark_green_cells"] += 1
                elif status == "uncoloured":
                    stats["uncoloured_cells"] += 1
        
        # Calculate success rate
        incorrect_cells = (stats["purple_cells"] + stats["red_cells"] + 
                         stats["orange_cells"] + stats["dark_green_cells"])
        effective_total = stats["total_rows"] - stats["blue_cells"]
        success_rate = ((effective_total - incorrect_cells) / effective_total) * 100 if effective_total > 0 else 0
        
        # Generate log file
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        log_filename = f"stock_price_analysis_{timestamp}.txt"
        
        with open(log_filename, "w") as log_file:
            log_file.write("STOCK PRICE EXTRACTION ANALYSIS\n")
            log_file.write("=" * 50 + "\n")
            log_file.write(f"Analysis Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            log_file.write("DETAILED STATISTICS:\n")
            log_file.write("-" * 20 + "\n")
            log_file.write(f"Total rows analyzed: {stats['total_rows']}\n")
            log_file.write(f"Blue cells (excluded): {stats['blue_cells']}\n")
            log_file.write(f"Purple cells (wrong figure): {stats['purple_cells']}\n")
            log_file.write(f"Red cells (formula/parse error): {stats['red_cells']}\n")
            log_file.write(f"Orange cells (.0 rounding error): {stats['orange_cells']}\n")
            log_file.write(f"Dark green cells (AI wrong): {stats['dark_green_cells']}\n")
            log_file.write(f"White/uncoloured cells (correct): {stats['uncoloured_cells']}\n\n")
            
            log_file.write("SUCCESS RATE CALCULATION:\n")
            log_file.write("-" * 20 + "\n")
            log_file.write(f"Total incorrect cells: {incorrect_cells}\n")
            log_file.write(f"Effective total (excluding blue cells): {effective_total}\n")
            log_file.write(f"Success rate: {success_rate:.2f}%\n")
            log_file.write(f"Formula: (effective_total - incorrect_cells) / effective_total\n")
            log_file.write(f"       = ({effective_total} - {incorrect_cells}) / {effective_total}\n")
        
        logging.info(f"Analysis complete. Results written to {log_filename}")
        return stats, success_rate
        
    except Exception as e:
        logging.error(f"Error analyzing Excel file: {e}")
        raise

if __name__ == "__main__":
    try:
        file_path = "Custodians.xlsx"
        stats, success_rate = analyze_excel_file(file_path)
        print(f"Analysis complete. Success rate: {success_rate:.2f}%")
    except Exception as e:
        logging.error(f"Script execution failed: {e}") 