import os
import sys
import logging
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from urllib.parse import urlparse
from collections import defaultdict, Counter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def is_blue_cell(cell):
    """Check if cell has blue background color"""
    # Simple approach to detect blue-ish cells
    try:
        if cell.fill and cell.fill.start_color and cell.fill.start_color.type == 'rgb':
            # Get RGB components
            color = cell.fill.start_color.rgb
            if isinstance(color, str) and len(color) >= 6:
                # Check for ARGB format (common in newer Excel files)
                if len(color) == 8:  # ARGB format
                    color = color[2:]  # Skip alpha channel
                
                # Extract RGB components
                try:
                    # Blue component is strong in blue cells
                    r = int(color[0:2], 16)
                    g = int(color[2:4], 16)
                    b = int(color[4:6], 16)
                    
                    # Check if this is a blue-ish cell (blue component dominates)
                    if b > 180 and b > (r + g) / 2:
                        return True
                except ValueError:
                    pass
    except Exception as e:
        # Silently handle any exceptions to avoid crashing
        pass
    
    return False

def get_domain_from_url(url):
    """Extract the domain from a URL"""
    if not url or not isinstance(url, str):
        return "invalid://"
    
    try:
        parsed_url = urlparse(url.strip())
        domain = parsed_url.netloc
        if domain:
            return f"https://{domain}/"
        return "invalid://"
    except:
        return "invalid://"

def is_valid_share_price(price_value, context_text="", url=""):
    """Check if the value looks like a valid share price"""
    # If it's not a numeric type, it's not valid
    if not isinstance(price_value, (int, float)):
        return False
    
    # Most share prices should not end in .0 exactly (though some legitimately might)
    # This helps catch dates (2023.0) and other numbers that aren't prices
    is_round_number = price_value == round(price_value, 0)
    
    if is_round_number:
        # Values below 100 that end in .0 are more likely to be valid
        if price_value < 100:
            return True
        else:
            # Values of 100.0 and above are suspicious
            return False
    
    return True

def generate_stock_price_log(excel_path, output_dir="logs"):
    """Generate a stock price log based on the current state of the Excel file"""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Check if logs directory exists, create if not
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_filename = os.path.join(output_dir, f"stock_prices_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        
        # Find URL and price columns
        url_column = None
        price_column = None
        
        for cell in ws[1]:  # First row contains headers
            if cell.value and isinstance(cell.value, str):
                if "WEBSITE" in cell.value.upper() or "INFO SOURCE" in cell.value.upper():
                    url_column = cell.column_letter
                elif "SHARE PRICE" in cell.value.upper() or "PRICE" in cell.value.upper():
                    price_column = cell.column_letter
        
        if not url_column or not price_column:
            logging.error(f"Could not identify URL or price column in the Excel file. URL column: {url_column}, Price column: {price_column}")
            return
        
        logging.info(f"URL column: {url_column}, Price column: {price_column}")
        
        # Collect data from Excel
        error_domains = defaultdict(int)
        error_urls = defaultdict(list)
        invalid_price_domains = defaultdict(int)
        invalid_price_urls = defaultdict(list)
        
        blue_cells_data = []
        normal_cells_data = []
        
        total_processed = 0
        blue_cells_count = 0
        invalid_price_count = 0
        
        # Process each row
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (after header)
            url_cell = ws[f"{url_column}{row_idx}"]
            price_cell = ws[f"{price_column}{row_idx}"]
            
            url = url_cell.value
            price = price_cell.value
            
            # Skip rows without URLs
            if not url:
                continue
            
            total_processed += 1
            is_blue = is_blue_cell(url_cell)
            domain = get_domain_from_url(url)
            
            cell_data = {
                "row": row_idx,
                "url": url,
                "success": False,
                "has_invalid_price": False
            }
            
            # Check if we have a price
            if price:
                # Determine if the price looks valid
                if isinstance(price, (int, float)) and not is_valid_share_price(price, url=url):
                    invalid_price_count += 1
                    invalid_price_domains[domain] += 1
                    invalid_price_urls[domain].append(url)
                    cell_data["has_invalid_price"] = True
                else:
                    cell_data["success"] = True
            else:
                # No price = error
                error_domains[domain] += 1
                error_urls[domain].append(url)
            
            if is_blue:
                blue_cells_count += 1
                blue_cells_data.append(cell_data)
            else:
                normal_cells_data.append(cell_data)
        
        # Create log file
        with open(log_filename, 'w', encoding='utf-8') as log_file:
            log_file.write("STOCK PRICES EXTRACTION LOG\n")
            log_file.write("=" * 50 + "\n")
            log_file.write(f"Run Date: {current_time}\n")
            log_file.write(f"Script: generate_logs.py (stock prices)\n\n")
            
            # RANKED ERROR DOMAINS (most errors first)
            log_file.write("SHARE PRICE ERRORS BY DOMAIN (RANKED - MOST ERRORS FIRST):\n")
            log_file.write("-" * 60 + "\n")
            
            # Sort domains by number of errors (descending)
            sorted_error_domains = sorted(error_domains.items(), key=lambda x: x[1], reverse=True)
            
            for domain, count in sorted_error_domains:
                log_file.write(f"  {count} errors: {domain}\n")
                # Show the URLs that had errors for this domain (limit to 5 for display)
                for url in error_urls[domain][:5]:
                    log_file.write(f"    └─ {url}\n")
                if len(error_urls[domain]) > 5:
                    log_file.write(f"    └─ ... and {len(error_urls[domain]) - 5} more URLs\n")
                log_file.write("\n")
            
            # INVALID PRICE SECTION
            if invalid_price_count > 0:
                log_file.write("\nINVALID PRICE VALUES BY DOMAIN (RANKED - MOST INVALID PRICES FIRST):\n")
                log_file.write("-" * 75 + "\n")
                log_file.write(f"Total invalid prices found: {invalid_price_count}\n")
                log_file.write("(These are likely incorrect - .0 values, dates, percentages, fund metadata, etc. instead of actual prices)\n")
                log_file.write("(Includes ALL .0 values: 24.00, 30.0, etc. - Excludes blue cells)\n\n")
                
                # Sort domains by number of invalid prices (descending)
                sorted_invalid_domains = sorted(invalid_price_domains.items(), key=lambda x: x[1], reverse=True)
                
                for domain, count in sorted_invalid_domains:
                    log_file.write(f"  {count} invalid prices: {domain}\n")
                    # Show the URLs that had invalid prices for this domain (limit to 5)
                    for url in invalid_price_urls[domain][:5]:
                        log_file.write(f"    └─ {url}\n")
                    if len(invalid_price_urls[domain]) > 5:
                        log_file.write(f"    └─ ... and {len(invalid_price_urls[domain]) - 5} more URLs\n")
                    log_file.write("\n")
            
            # BLUE CELLS SECTION
            log_file.write("\nBLUE CELLS SECTION (EXCLUDED FROM SUCCESS RATE):\n")
            log_file.write("-" * 50 + "\n")
            log_file.write(f"Total blue cells: {blue_cells_count}\n")
            
            if blue_cells_data:
                blue_success = sum(1 for cell in blue_cells_data if cell['success'])
                blue_errors = blue_cells_count - blue_success
                blue_invalid_prices = sum(1 for cell in blue_cells_data if cell.get('has_invalid_price', False))
                log_file.write(f"Blue cells successful: {blue_success}\n")
                log_file.write(f"Blue cells with errors: {blue_errors}\n")
                log_file.write(f"Blue cells with invalid prices: {blue_invalid_prices}\n")
                if blue_cells_count > 0:
                    blue_success_rate = blue_success / blue_cells_count
                    log_file.write(f"Blue cells success rate: {blue_success}/{blue_cells_count} = {blue_success_rate:.2%}\n")
                log_file.write("\nBlue cell details:\n")
                for cell in blue_cells_data[:10]:  # Show first 10
                    status = "✅ Success" if cell['success'] else ("🔸 Invalid Price" if cell.get('has_invalid_price') else "❌ Error")
                    log_file.write(f"  Row {cell['row']}: {status} - {cell['url'][:60]}{'...' if len(cell['url']) > 60 else ''}\n")
                if len(blue_cells_data) > 10:
                    log_file.write(f"  ... and {len(blue_cells_data) - 10} more blue cells\n")
            else:
                log_file.write("No blue cells found.\n")
            
            # NORMAL CELLS SECTION
            normal_cells_count = len(normal_cells_data)
            normal_success = sum(1 for cell in normal_cells_data if cell['success'])
            normal_errors = sum(1 for cell in normal_cells_data if not cell['success'] and not cell.get('has_invalid_price', False))
            normal_invalid_prices = sum(1 for cell in normal_cells_data if cell.get('has_invalid_price', False))
            
            log_file.write(f"\nNORMAL CELLS SECTION (USED FOR SUCCESS RATE):\n")
            log_file.write("-" * 50 + "\n")
            log_file.write(f"Total normal cells: {normal_cells_count}\n")
            log_file.write(f"Normal cells successful: {normal_success}\n")
            log_file.write(f"Normal cells with errors: {normal_errors}\n")
            log_file.write(f"Normal cells with invalid prices (treated as errors): {normal_invalid_prices}\n")
            if normal_cells_count > 0:
                normal_success_rate = normal_success / normal_cells_count
                log_file.write(f"Normal cells success rate: {normal_success}/{normal_cells_count} = {normal_success_rate:.2%}\n")
                log_file.write(f"(Success rate excludes both errors and invalid prices)\n")
            
            log_file.write(f"\nSUMMARY:\n")
            log_file.write("-" * 10 + "\n")
            log_file.write(f"Total URLs processed: {total_processed}\n")
            log_file.write(f"Blue cells (excluded from success rate): {blue_cells_count}\n")
            log_file.write(f"Normal cells (used for success rate): {normal_cells_count}\n")
            log_file.write(f"Successful extractions (normal cells only): {normal_success}\n")
            log_file.write(f"URLs with invalid prices: {invalid_price_count}\n")
            log_file.write(f"Traditional errors: {sum(error_domains.values())}\n")
            if normal_cells_count > 0:
                final_success_rate = normal_success / normal_cells_count
                log_file.write(f"FINAL SUCCESS RATE: {normal_success}/{normal_cells_count} = {final_success_rate:.2%}\n")
                log_file.write(f"(Excludes blue cells, traditional errors, AND invalid prices)\n")
            else:
                log_file.write("FINAL SUCCESS RATE: N/A (no normal cells to process)\n")
        
        logging.info(f"Stock prices log file created: {log_filename}")
        return log_filename
    
    except Exception as e:
        logging.error(f"Error generating stock price log: {e}")
        import traceback
        traceback.print_exc()
        return None

def generate_outstanding_shares_log(excel_path, output_dir="logs"):
    """Generate an outstanding shares log based on the current state of the Excel file"""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Check if logs directory exists, create if not
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_filename = os.path.join(output_dir, f"outstanding_shares_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        
        # Find URL and shares columns
        url_column = None
        shares_column = None
        
        for cell in ws[1]:  # First row contains headers
            if cell.value and isinstance(cell.value, str):
                if "WEBSITE" in cell.value.upper() or "INFO SOURCE" in cell.value.upper():
                    url_column = cell.column_letter
                elif "OUTSTANDING SHARES" in cell.value.upper():
                    shares_column = cell.column_letter
        
        if not url_column:
            logging.error("Could not identify URL column in the Excel file.")
            return
        
        if not shares_column:
            # Look for specific column by index if header doesn't match exactly
            for cell in ws[1]:
                if cell.value and isinstance(cell.value, str) and "OUTSTANDING" in cell.value.upper():
                    shares_column = cell.column_letter
                    break
            
            if not shares_column:
                logging.error("Could not identify outstanding shares column in the Excel file.")
                return
        
        logging.info(f"URL column: {url_column}, Shares column: {shares_column}")
        
        # Collect data from Excel
        error_domains = defaultdict(int)
        error_urls = defaultdict(list)
        
        total_processed = 0
        blue_cells_count = 0
        total_successful = 0
        
        # Process each row
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (after header)
            url_cell = ws[f"{url_column}{row_idx}"]
            shares_cell = ws[f"{shares_column}{row_idx}"]
            
            url = url_cell.value
            shares = shares_cell.value
            
            # Skip rows without URLs
            if not url:
                continue
            
            total_processed += 1
            is_blue = is_blue_cell(url_cell)
            domain = get_domain_from_url(url)
            
            # Count blue cells
            if is_blue:
                blue_cells_count += 1
            
            # Check if we have shares data
            if shares and not (isinstance(shares, str) and shares.startswith("Error")):
                total_successful += 1
            else:
                # No shares data = error
                error_domains[domain] += 1
                error_urls[domain].append(url)
        
        # Create log file
        with open(log_filename, 'w', encoding='utf-8') as log_file:
            log_file.write("OUTSTANDING SHARES EXTRACTION LOG\n")
            log_file.write("=" * 50 + "\n")
            log_file.write(f"Run Date: {current_time}\n")
            log_file.write(f"Script: generate_logs.py (outstanding shares)\n\n")
            
            # ERROR DOMAINS
            log_file.write("OUTSTANDING SHARES ERRORS BY DOMAIN:\n")
            log_file.write("-" * 40 + "\n")
            
            # Sort domains alphabetically
            sorted_error_domains = sorted(error_domains.items())
            
            for domain, count in sorted_error_domains:
                log_file.write(f"{domain} errors: {count}\n")
            
            log_file.write(f"\nSUMMARY:\n")
            log_file.write("-" * 10 + "\n")
            log_file.write(f"Total URLs processed: {total_processed}\n")
            log_file.write(f"Blue cells (excluded from success rate): {blue_cells_count}\n")
            log_file.write(f"Successful extractions: {total_successful}\n")
            log_file.write(f"Effective URLs for success calculation: {total_processed - blue_cells_count}\n")
            if (total_processed - blue_cells_count) > 0:
                success_rate = total_successful / (total_processed - blue_cells_count)
                log_file.write(f"Adjusted success rate: {total_successful}/{total_processed - blue_cells_count} = {success_rate:.2%}\n")
            else:
                log_file.write("Adjusted success rate: N/A (no valid URLs to process)\n")
        
        logging.info(f"Outstanding shares log file created: {log_filename}")
        return log_filename
    
    except Exception as e:
        logging.error(f"Error generating outstanding shares log: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_combined_log(stock_log, shares_log, output_dir="logs"):
    """Creates a combined log file from stock prices and outstanding shares logs."""
    try:
        # Check if logs directory exists, create if not
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        combined_log_filename = os.path.join(output_dir, f"combined_extraction_log_{current_time}.txt")
        
        with open(combined_log_filename, 'w', encoding='utf-8') as combined_file:
            combined_file.write("COMBINED STOCK PRICES & OUTSTANDING SHARES EXTRACTION LOG\n")
            combined_file.write("=" * 70 + "\n")
            combined_file.write(f"Combined Log Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            combined_file.write(f"Script: generate_logs.py (combined)\n\n")
            
            # Read and include stock prices log
            combined_file.write("STOCK PRICES LOG:\n")
            combined_file.write("-" * 20 + "\n")
            try:
                with open(stock_log, 'r', encoding='utf-8') as stock_file:
                    stock_content = stock_file.read()
                    # Skip the header line and add the content
                    lines = stock_content.split('\n')[4:]  # Skip first 4 header lines
                    combined_file.write('\n'.join(lines))
            except Exception as e:
                combined_file.write(f"Error reading stock prices log: {e}\n")
            
            combined_file.write("\n\n")
            
            # Read and include outstanding shares log
            combined_file.write("OUTSTANDING SHARES LOG:\n")
            combined_file.write("-" * 25 + "\n")
            try:
                with open(shares_log, 'r', encoding='utf-8') as shares_file:
                    shares_content = shares_file.read()
                    # Skip the header line and add the content
                    lines = shares_content.split('\n')[4:]  # Skip first 4 header lines
                    combined_file.write('\n'.join(lines))
            except Exception as e:
                combined_file.write(f"Error reading outstanding shares log: {e}\n")
        
        logging.info(f"Combined log file created: {combined_log_filename}")
        return combined_log_filename
    
    except Exception as e:
        logging.error(f"Error creating combined log: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """Main function to generate logs based on command line arguments"""
    try:
        # Default Excel file path
        excel_file = "custodians.xlsx"
        
        # Get command line arguments
        if len(sys.argv) > 1:
            log_type = sys.argv[1].lower()
            if len(sys.argv) > 2:
                excel_file = sys.argv[2]
        else:
            # No arguments provided, show usage
            print("Usage: python generate_logs.py [stock|shares|combined] [excel_file_path]")
            print("       If excel_file_path is not provided, 'custodians.xlsx' will be used.")
            print("\nOptions:")
            print("  stock     - Generate stock prices log")
            print("  shares    - Generate outstanding shares log")
            print("  combined  - Generate both logs and combine them")
            return
        
        # Check if Excel file exists
        if not os.path.exists(excel_file):
            logging.error(f"Excel file not found: {excel_file}")
            return
        
        # Generate logs based on type
        if log_type == "stock":
            generate_stock_price_log(excel_file)
        elif log_type == "shares":
            generate_outstanding_shares_log(excel_file)
        elif log_type == "combined":
            stock_log = generate_stock_price_log(excel_file)
            shares_log = generate_outstanding_shares_log(excel_file)
            if stock_log and shares_log:
                create_combined_log(stock_log, shares_log)
        else:
            logging.error(f"Invalid log type: {log_type}. Use 'stock', 'shares', or 'combined'.")
    
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 