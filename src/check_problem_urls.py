#!/usr/bin/env python3
"""
Check specific URLs that have .0 errors to understand what needs to be fixed
"""

import openpyxl

def check_problem_urls():
    wb = openpyxl.load_workbook('Custodians.xlsx')
    ws = wb['Non-derivative exposures']

    # Check specific rows with .0 errors from the logs
    problem_rows = [211, 328, 376, 377, 378, 380]
    
    print("🔍 URLs with .0 errors that need investigation:")
    print("=" * 80)
    
    for row in problem_rows:
        url = ws[f'P{row}'].value
        current_result = ws[f'L{row}'].value
        print(f'Row {row}: {url}')
        print(f'  Current result: {current_result}')
        print()

if __name__ == "__main__":
    check_problem_urls() 