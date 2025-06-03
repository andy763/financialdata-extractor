#!/usr/bin/env python3
"""
Test the complete Excel update workflow to verify that successful 
custom extractions are properly written to Excel
"""

import logging
import os
import tempfile
import shutil
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def create_test_excel():
    """Create a small test Excel file with a single Valour URL"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shares"
    
    # Set up headers
    ws['P1'] = 'Primary URL'
    ws['M1'] = 'Outstanding shares'
    
    # Add a test row with Valour URL
    ws['P2'] = 'https://valour.com/en/products/physical-bitcoin-carbon-neutral-usd'
    
    # Save to temp file
    temp_file = "test_excel_update.xlsx"
    wb.save(temp_file)
    return temp_file

def test_excel_update_workflow():
    """Test the complete workflow from URL to Excel update"""
    
    print("📊 TESTING COMPLETE EXCEL UPDATE WORKFLOW")
    print("=" * 60)
    
    # Create test Excel file
    test_file = create_test_excel()
    print(f"✅ Created test Excel file: {test_file}")
    
    # Set up Chrome driver
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920x1080")
    
    driver = None
    
    try:
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        
        # Import required functions
        from outstanding_shares_updater import extract_outstanding_shares_with_ai_fallback
        from openpyxl import load_workbook
        
        # Load the test Excel file
        wb = load_workbook(test_file)
        ws = wb.active
        
        # Get the URL from the Excel file
        test_url = ws['P2'].value
        print(f"\n🌐 Testing URL from Excel: {test_url}")
        
        # Extract outstanding shares
        print("\n1️⃣ Extracting outstanding shares...")
        result = extract_outstanding_shares_with_ai_fallback(driver, test_url)
        print(f"   Result: {result}")
        
        # Check the result and update Excel
        print("\n2️⃣ Updating Excel cell...")
        shares_cell = 'M2'
        
        if result and "outstanding_shares" in result:
            # Success case
            shares_value = result['outstanding_shares']
            method = result.get('method', 'unknown')
            print(f"   ✅ SUCCESS: Found outstanding shares: {shares_value}")
            print(f"   🔧 Method used: {method}")
            
            # Update the Excel cell
            ws[shares_cell] = shares_value
            wb.save(test_file)
            print(f"   💾 Updated Excel cell {shares_cell} with: {shares_value}")
            
        elif result and "error" in result:
            # Error case
            error_message = result['error']
            method = result.get('method', 'unknown')
            print(f"   ❌ ERROR: {error_message}")
            print(f"   🔧 Method attempted: {method}")
            
            # Update the Excel cell with error
            ws[shares_cell] = f"Error: {error_message}"
            wb.save(test_file)
            print(f"   💾 Updated Excel cell {shares_cell} with error: {error_message}")
            
        else:
            # Unexpected case
            print(f"   ❓ UNEXPECTED: {result}")
            ws[shares_cell] = "Error: Unexpected result format"
            wb.save(test_file)
        
        # Verify the Excel file was updated correctly
        print("\n3️⃣ Verifying Excel update...")
        wb_verify = load_workbook(test_file)
        ws_verify = wb_verify.active
        final_value = ws_verify[shares_cell].value
        print(f"   📋 Final Excel cell value: {final_value}")
        
        # Analysis
        print("\n4️⃣ Analysis:")
        if final_value and "Error:" in str(final_value):
            print("   🚨 BUG CONFIRMED: Excel contains error despite successful extraction!")
            print("   🔍 This indicates the issue described in the conversation summary still exists")
            return False
        elif final_value and "120,000" in str(final_value):
            print("   ✅ SUCCESS: Excel correctly updated with extracted shares!")
            print("   🎉 The fix is working properly")
            return True
        else:
            print(f"   ❓ INCONCLUSIVE: Unexpected Excel value: {final_value}")
            return False
            
    except Exception as e:
        print(f"❌ Test error: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        if driver:
            driver.quit()
        
        # Clean up
        if os.path.exists(test_file):
            os.remove(test_file)
            print(f"\n🧹 Cleaned up test file: {test_file}")

if __name__ == "__main__":
    success = test_excel_update_workflow()
    if success:
        print("\n🎊 TEST PASSED: The fix is working correctly!")
    else:
        print("\n💀 TEST FAILED: The issue still exists!")
