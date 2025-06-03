import re
import logging
import time
import traceback
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from src.sixgroup_shares_extractor import extract_sixgroup_shares  # Import the SIX Group extractor

# Import improved custom domain extractors
IMPROVED_EXTRACTORS_AVAILABLE = False
CUSTOM_EXTRACTORS_AVAILABLE = False
extract_with_custom_function = None

try:
    # First try to import the improved extractors
    from src.improved_custom_domain_extractors import (
        extract_with_custom_function,
        extract_valour_shares,
        extract_grayscale_shares,
        extract_vaneck_shares,
        extract_vaneck_de_shares,
        extract_wisdomtree_shares,
        extract_proshares_shares,
        get_custom_extractor
    )
    IMPROVED_EXTRACTORS_AVAILABLE = True
    CUSTOM_EXTRACTORS_AVAILABLE = True
    logging.info("✅ Improved custom domain extractors loaded successfully")
    
    # Map domains to their extractors for direct access
    DOMAIN_EXTRACTORS = {
        'valour.com': extract_valour_shares,
        'grayscale.com': extract_grayscale_shares,
        'vaneck.com/de': extract_vaneck_de_shares,
        'vaneck.com': extract_vaneck_shares,
        'wisdomtree.eu': extract_wisdomtree_shares,
        'proshares.com': extract_proshares_shares,
    }
    
except ImportError as e:
    logging.warning(f"Could not import improved extractors: {e}")
    try:
        # Fall back to original extractors
        from src.custom_domain_extractors import extract_with_custom_function, get_custom_extractor
        IMPROVED_EXTRACTORS_AVAILABLE = False
        CUSTOM_EXTRACTORS_AVAILABLE = True
        logging.info("⚠️ Using original custom domain extractors")
        
        # We don't know exactly which extractors are available in the original module
        DOMAIN_EXTRACTORS = {}
        
    except ImportError as e2:
        logging.warning(f"Could not import original extractors: {e2}")
        extract_with_custom_function = None
        get_custom_extractor = None
        IMPROVED_EXTRACTORS_AVAILABLE = False
        CUSTOM_EXTRACTORS_AVAILABLE = False
        DOMAIN_EXTRACTORS = {}
        logging.warning("❌ No custom domain extractors available")

# NEW: Import for Groq AI analysis
try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False
    logging.warning("Groq library not available. AI fallback will be disabled.")

# Groq AI Configuration
GROQ_API_KEY = "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"

# Multiple models in order of preference (best to fastest/most available)
GROQ_MODELS = [
    "llama-3.3-70b-versatile",     # Best quality, higher rate limits
    "llama-3.1-8b-instant",       # Faster, more available
    "gemma2-9b-it",               # Alternative option
    "mixtral-8x7b-32768",         # Fallback option
]

def make_groq_request_with_fallback(client, messages, max_retries=1):  # Reduced from 3 to 1
    """
    Make a Groq API request with model fallback and rate limit handling
    OPTIMIZED: Reduced retries and wait times for speed
    """
    import time
    import random
    
    for model_index, model in enumerate(GROQ_MODELS):
        for attempt in range(max_retries):
            try:
                logging.info(f"Trying Groq model: {model} (attempt {attempt + 1})")
                
                # Shorter prompt for faster models to stay within token limits
                if model in ["llama-3.1-8b-instant", "gemma2-9b-it"]:
                    max_tokens = 30  # Further reduced for speed
                else:
                    max_tokens = 50  # Reduced from 100
                
                response = client.chat.completions.create(
                    messages=messages,
                    model=model,
                    temperature=0.1,
                    max_tokens=max_tokens,
                )
                
                logging.info(f"✅ Success with {model}")
                return response, model
                
            except Exception as e:
                error_str = str(e).lower()
                
                # Check if it's a rate limit error
                if "rate limit" in error_str or "429" in error_str:
                    if attempt < max_retries - 1:
                        # Much shorter wait time for speed
                        wait_time = 0.5 + random.uniform(0, 0.5)  # 0.5-1.0s instead of exponential
                        logging.warning(f"Rate limit hit with {model}, waiting {wait_time:.1f}s before retry...")
                        time.sleep(wait_time)
                        continue
                    else:
                        logging.warning(f"Rate limit exhausted for {model}, trying next model...")
                        break  # Try next model
                
                # Check for other errors that suggest model unavailability
                elif any(err in error_str for err in ["unavailable", "overloaded", "timeout"]):
                    logging.warning(f"Model {model} unavailable: {e}")
                    break  # Try next model immediately
                
                else:
                    # Unknown error, log and try next model
                    logging.error(f"Unknown error with {model}: {e}")
                    break
    
    # If all models failed
    raise Exception("All Groq models failed or rate limited")

def clean_html_for_shares_ai(html_content, max_length=2000):  # Further reduced for speed
    """
    Clean and truncate HTML content for AI analysis while preserving shares-relevant information.
    OPTIMIZED: Much shorter content and faster processing for speed.
    """
    try:
        soup = BeautifulSoup(html_content, "html.parser")
        
        # Remove script and style elements quickly
        for script in soup(["script", "style", "nav", "header", "footer"]):
            script.decompose()
        
        # Get text content
        text = soup.get_text()
        
        # Quick whitespace cleanup (simplified for speed)
        text = ' '.join(text.split())
        
        # Focus on shares content - simplified keyword search for speed
        shares_keywords = ['outstanding', 'shares', 'units', 'issued', 'million', 'billion']
        
        # Quick search for relevant sections
        relevant_parts = []
        words = text.split()
        
        for i, word in enumerate(words):
            if any(keyword in word.lower() for keyword in shares_keywords):
                # Take 20 words before and after the keyword for context
                start = max(0, i - 20)
                end = min(len(words), i + 21)
                relevant_parts.append(' '.join(words[start:end]))
        
        # Combine relevant parts
        if relevant_parts:
            prioritized_text = ' ... '.join(relevant_parts)
        else:
            # Fallback to first part of text
            prioritized_text = text
        
        return prioritized_text[:max_length]
    
    except Exception as e:
        logging.error(f"Error cleaning HTML for shares AI: {e}")
        # Fast fallback: just truncate raw text
        soup = BeautifulSoup(html_content, "html.parser")
        return soup.get_text()[:max_length]

def analyze_shares_with_groq(html_content, url):
    """
    Use Groq AI to analyze website content and extract outstanding shares.
    IMPROVED: Multiple model fallback and rate limit handling.
    """
    if not GROQ_AVAILABLE:
        return {"error": "Groq AI not available"}
    
    try:
        # Initialize Groq client
        client = Groq(api_key=GROQ_API_KEY)
        
        # Clean HTML content for AI analysis (further reduced for speed)
        cleaned_content = clean_html_for_shares_ai(html_content, max_length=2000)  # Reduced from 4000
        
        # Create a shorter, more efficient prompt to reduce token usage and speed up processing
        prompt = f"""Extract outstanding shares from: {url}

Find: outstanding shares, shares outstanding, total shares, shares issued
Return ONLY the full number (e.g., "56,839,000")
If not found: "NO_SHARES_FOUND"

Content: {cleaned_content[:1500]}

Number:"""

        # Create message format
        messages = [{"role": "user", "content": prompt}]
        
        # Make request with fallback handling
        chat_completion, used_model = make_groq_request_with_fallback(client, messages)
        
        # Extract the response
        ai_response = chat_completion.choices[0].message.content.strip()
        logging.info(f"Groq AI shares response from {used_model} for {url}: {ai_response}")
        
        # Parse the AI response
        if "NO_SHARES_FOUND" in ai_response.upper():
            return {"error": "AI could not find valid outstanding shares"}
        
        # Try multiple strategies to extract the shares number from AI response
        shares_value = None
        shares_str = None
        
        # Strategy 1: Look for "Shares Outstanding:" followed by a number
        outstanding_match = re.search(r'shares\s+outstanding\s*:?\s*([\d,\.\s]+)', ai_response, re.IGNORECASE)
        if outstanding_match:
            shares_str = outstanding_match.group(1).strip()
            logging.info(f"Found 'Shares Outstanding' pattern: {shares_str}")
        
        # Strategy 2: Look for any large number (6+ digits) that could be shares
        if not shares_str:
            large_number_matches = re.findall(r'(\d{1,3}(?:[,\.\s]\d{3})*(?:\d{3})*)', ai_response)
            for match in large_number_matches:
                # Clean the number and check if it's in reasonable shares range
                test_normalized = match.replace(',', '').replace(' ', '').replace('.', '')
                try:
                    test_value = int(test_normalized)
                    if 10000 <= test_value <= 100000000000:  # Reasonable shares range
                        shares_str = match
                        logging.info(f"Found large number pattern: {shares_str}")
                        break
                except ValueError:
                    continue
        
        # Strategy 2.5: Handle decimal numbers that might represent thousands (e.g., "56,839.00" = "56,839,000")
        if not shares_str:
            decimal_matches = re.findall(r'(\d{1,3}(?:,\d{3})*\.\d{2})', ai_response)
            for match in decimal_matches:
                # Check if this looks like a thousands representation
                base_number = match.split('.')[0].replace(',', '')
                try:
                    base_value = int(base_number)
                    if 10000 <= base_value <= 100000000:  # Could be thousands
                        # Assume it's in thousands and multiply by 1000
                        potential_shares = base_value * 1000
                        if 10000000 <= potential_shares <= 100000000000:  # Reasonable full shares range
                            shares_str = f"{potential_shares:,}"
                            logging.info(f"Found decimal thousands pattern: {match} -> {shares_str}")
                            break
                except ValueError:
                    continue
        
        # Strategy 3: Fallback to any number in the response
        if not shares_str:
            number_match = re.search(r'(\d{1,15}(?:[,\.\s]\d{1,3})*)', ai_response)
            if number_match:
                shares_str = number_match.group(1)
                logging.info(f"Found fallback number pattern: {shares_str}")
        
        if shares_str:
            # Clean and normalize the number
            normalized = shares_str.replace(',', '').replace(' ', '').replace('.', '')
            
            try:
                shares_value = int(normalized)
                if 1000 <= shares_value <= 100000000000:  # Reasonable shares range
                    # Format the number based on whether a unit was explicitly specified
                    if re.search(r'million$|^m$|m$|M$', shares_str, re.IGNORECASE) or re.search(r'billion$|^b$|b$', shares_str, re.IGNORECASE):
                        # Only format as millions/billions if explicitly specified
                        if shares_value >= 1000000000:
                            formatted_value = f"{shares_value/1000000000:.2f} billion"
                        else:
                            formatted_value = f"{shares_value/1000000:.2f} million"
                    else:
                        # Otherwise preserve the exact number, without commas for easier processing
                        formatted_value = f"{int(shares_value)}"
                    
                    logging.info(f"Successfully parsed shares: {shares_str} -> {shares_value} -> {formatted_value}")
                    return {"ai_extracted_shares": formatted_value}
                else:
                    return {"error": f"AI extracted unreasonable shares count: {shares_value}"}
            except ValueError:
                return {"error": f"Could not parse AI response as number: {shares_str}"}
        else:
            return {"error": f"AI response contains no recognizable number: {ai_response}"}
            
    except Exception as e:
        error_msg = str(e)
        if "rate limit" in error_msg.lower() or "429" in error_msg:
            logging.error(f"Groq rate limit exceeded for shares {url}: {e}")
            return {"error": f"AI rate limit exceeded - try again later"}
        else:
            logging.error(f"Groq AI shares analysis error for {url}: {e}")
            return {"error": f"AI shares analysis failed: {str(e)}"}

def try_shares_ai_fallback(driver, url):
    """
    Try AI analysis as a fallback when traditional shares scraping fails.
    """
    if not GROQ_AVAILABLE:
        return {"error": "AI fallback not available"}
    
    try:
        logging.info(f"Attempting AI fallback analysis for outstanding shares: {url}")
        
        # Get fresh page content for AI analysis - OPTIMIZED for speed
        driver.get(url)
        WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.TAG_NAME, "body")))  # Reduced from 15 to 8
        
        # Reduced wait time for dynamic content
        time.sleep(1)  # Reduced from 3 to 1 second
        
        # Get the page source
        html_content = driver.page_source
        
        # Analyze with AI
        result = analyze_shares_with_groq(html_content, url)
        
        if "ai_extracted_shares" in result:
            logging.info(f"AI successfully extracted shares: {result['ai_extracted_shares']}")
            return {"outstanding_shares": f"{result['ai_extracted_shares']} (AI)"}
        else:
            logging.info(f"AI shares fallback failed: {result.get('error', 'Unknown error')}")
            return result
            
    except Exception as e:
        logging.error(f"AI shares fallback error for {url}: {e}")
        return {"error": f"AI shares fallback failed: {str(e)}"}

def find_processable_rows_and_get_urls(ws, url_column_letter):
    """
    Dynamically finds ALL rows that have URLs in the specified column.
    No predetermined row limit - scans the entire worksheet.
    """
    processable_rows_urls = {}
    
    # Find the maximum row with data in the worksheet
    max_row = ws.max_row
    
    # Start from row 2 (skip header) and scan all rows
    for row_idx in range(2, max_row + 1):
        coord = f"{url_column_letter}{row_idx}"
        url_value = ws[coord].value
        
        # Check if the cell has a URL (string starting with http)
        if url_value and isinstance(url_value, str) and url_value.strip().lower().startswith(("http://", "https://")):
            processable_rows_urls[row_idx] = url_value.strip()
                
    return processable_rows_urls

def extract_outstanding_shares(driver, url):
    """
    Extract outstanding shares from a webpage using the keyword 'outstanding'
    """
    try:
        # Set a shorter page load timeout for speed
        driver.set_page_load_timeout(20)  # Reduced from 60 to 20
        
        # Special handling for Fidelity URLs which can be problematic
        if "fidelity.com" in url.lower():
            print(f"  Detected Fidelity URL: {url} - using special handling")
            try:
                driver.get(url)
                # Wait for the page to load with a shorter timeout
                WebDriverWait(driver, 15).until(  # Reduced from 45 to 15
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                # Reduced wait time for JavaScript content
                time.sleep(1)  # Reduced from 3 to 1
            except (TimeoutException, WebDriverException) as e:
                print(f"  Timeout or error with Fidelity URL: {str(e)}")
                return {"error": f"Timed out loading Fidelity page: {str(e)}"}
        else:
            driver.get(url)
            # Wait for the page to load with shorter timeout
            WebDriverWait(driver, 10).until(  # Reduced from 30 to 10
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            # Reduced wait time for JavaScript content
            time.sleep(2)  # Reduced from 5 to 2
    except (TimeoutException, WebDriverException) as e:
        return {"error": f"Timed out waiting for page to load: {str(e)}"}

    soup = BeautifulSoup(driver.page_source, "html.parser")
    
    # Get the text of the entire page
    page_text = soup.get_text(" ", strip=True)
    
    # Define regex patterns to find outstanding shares
    patterns = [
        # Common patterns for outstanding shares - include space-separated numbers
        r"(?:shares|units)\s+outstanding\s*(?:is|:|are|of)?\s*([\d,\.\s]+\s*(?:million|m|billion|b)?)",
        r"outstanding\s+(?:shares|units)\s*(?:is|:|are|of)?\s*([\d,\.\s]+\s*(?:million|m|billion|b)?)",
        r"(?:shares|units)\s+outstanding\s*(?:[^0-9]*?)([\d,\.\s]+\s*(?:million|m|billion|b)?)",
        r"outstanding\s+(?:shares|share count)(?:[^0-9]*?)([\d,\.\s]+\s*(?:million|m|billion|b)?)",
        r"total\s+(?:shares|units)\s+outstanding\s*(?:[^0-9]*?)([\d,\.\s]+\s*(?:million|m|billion|b)?)",
        r"outstanding[^0-9]{0,10}([\d,\.\s]+\s*(?:million|m|billion|b)?)"
    ]
    
    for pattern in patterns:
        matches = re.finditer(pattern, page_text, re.IGNORECASE)
        for match in matches:
            shares_text = match.group(1).strip()
            logging.info(f"Found potential outstanding shares (raw): {shares_text}")
            
            try:
                # Check for explicit magnitude indicators
                multiplier = 1
                original_text = shares_text  # Keep original for logging
                
                # Only apply multipliers if explicitly indicated
                if re.search(r'million$|^m$|m$', shares_text, re.IGNORECASE):
                    multiplier = 1000000
                    shares_text = re.sub(r'(?:million|m)$', '', shares_text, flags=re.IGNORECASE).strip()
                elif re.search(r'billion$|^b$|b$', shares_text, re.IGNORECASE):
                    multiplier = 1000000000
                    shares_text = re.sub(r'(?:billion|b)$', '', shares_text, flags=re.IGNORECASE).strip()
                elif re.search(r'k$', shares_text, re.IGNORECASE):
                    multiplier = 1000
                    shares_text = re.sub(r'k$', '', shares_text, flags=re.IGNORECASE).strip()
                elif re.search(r'M$', shares_text):  # Capital M suffix
                    multiplier = 1000000
                    shares_text = re.sub(r'M$', '', shares_text).strip()
                
                # Handle European number format with spaces as thousand separators
                if ' ' in shares_text:
                    shares_text = shares_text.replace(' ', '')
                
                # Remove any remaining non-numeric characters except periods and commas
                clean_text = re.sub(r'[^\d\.,]', '', shares_text)
                
                # Normalize decimal/thousand separators
                if ',' in clean_text and '.' in clean_text:
                    # Both comma and period exist - determine which is the decimal separator
                    if clean_text.rindex(',') > clean_text.rindex('.'):
                        # European format: 1.234,56
                        numeric_value = clean_text.replace('.', '').replace(',', '.')
                    else:
                        # US format: 1,234.56
                        numeric_value = clean_text.replace(',', '')
                elif ',' in clean_text:
                    # Only commas - treat as thousand separators unless it's clearly a decimal
                    if clean_text.count(',') == 1 and len(clean_text.split(',')[1]) <= 3:
                        # Likely decimal: "123,45" or "123,456"
                        numeric_value = clean_text.replace(',', '.')
                    else:
                        # Thousand separators: "1,234,567"
                        numeric_value = clean_text.replace(',', '')
                elif '.' in clean_text:
                    # Only periods - treat as decimal point if it appears once near the end
                    if clean_text.count('.') == 1 and len(clean_text.split('.')[1]) <= 3:
                        # Likely decimal: "123.45" or "123.456"
                        numeric_value = clean_text
                    else:
                        # Thousand separators: "1.234.567"
                        numeric_value = clean_text.replace('.', '')
                else:
                    # No separators
                    numeric_value = clean_text
                
                # Convert to float and apply multiplier only if explicitly indicated
                shares_value = float(numeric_value) * multiplier
                
                # Log the transformation for debugging
                if multiplier != 1:
                    logging.info(f"Applied explicit multiplier: {original_text} -> {shares_value:,.0f}")
                
                # Format the number based on whether a unit was explicitly specified
                if re.search(r'million$|^m$|m$|M$', original_text, re.IGNORECASE) or re.search(r'billion$|^b$|b$', original_text, re.IGNORECASE):
                    # Only format as millions/billions if explicitly specified
                    if shares_value >= 1000000000:
                        formatted_value = f"{shares_value/1000000000:.2f} billion"
                    else:
                        formatted_value = f"{shares_value/1000000:.2f} million"
                else:
                    # Otherwise preserve the exact number, without commas for easier processing
                    formatted_value = f"{int(shares_value)}"
                
                logging.info(f"Final shares value: {formatted_value}")
                return {"outstanding_shares": formatted_value}
            except (ValueError, TypeError) as e:
                logging.error(f"Error parsing shares value '{shares_text}': {e}")
                continue
    
    # Also specifically look for a table cell or div containing the word "outstanding"
    outstanding_elements = soup.find_all(
        lambda tag: tag.name in ('td', 'div', 'span', 'p') and 
                   re.search(r'\boutstanding\b', tag.get_text(strip=True), re.IGNORECASE)
    )
    
    for element in outstanding_elements:
        # Get text of the element and its siblings
        element_text = element.get_text(strip=True)
        
        # Look for numbers in the element itself - updated to handle space-separated numbers
        number_match = re.search(r'([\d,\.\s]+)\s*(?:million|m|billion|b|k)?', element_text)
        if number_match:
            shares_text = number_match.group(0).strip()
            logging.info(f"Found potential outstanding shares in element (raw): {shares_text}")
            
            try:
                # Check for explicit magnitude indicators
                multiplier = 1
                
                # Only apply multipliers if explicitly indicated
                if re.search(r'million$|^m$|m$', shares_text, re.IGNORECASE):
                    multiplier = 1000000
                    shares_text = re.sub(r'(?:million|m)$', '', shares_text, flags=re.IGNORECASE).strip()
                elif re.search(r'billion$|^b$|b$', shares_text, re.IGNORECASE):
                    multiplier = 1000000000
                    shares_text = re.sub(r'(?:billion|b)$', '', shares_text, flags=re.IGNORECASE).strip()
                # Handle K for thousands if present
                elif re.search(r'k$', shares_text, re.IGNORECASE):
                    multiplier = 1000
                    shares_text = re.sub(r'k$', '', shares_text, flags=re.IGNORECASE).strip()
                
                # Special case for 'M' suffix (for million)
                elif re.search(r'M$', shares_text):
                    multiplier = 1000000
                    shares_text = re.sub(r'M$', '', shares_text).strip()
                    
                # Handle European number format with spaces as thousand separators
                if ' ' in shares_text:
                    shares_text = shares_text.replace(' ', '')
                
                # Remove any remaining non-numeric characters except periods and commas
                clean_text = re.sub(r'[^\d\.,]', '', shares_text)
                
                # Normalize decimal/thousand separators
                if ',' in clean_text and '.' in clean_text:
                    # Both comma and period exist - determine which is the decimal separator
                    if clean_text.rindex(',') > clean_text.rindex('.'):
                        # European format: 1.234,56
                        numeric_value = clean_text.replace('.', '').replace(',', '.')
                    else:
                        # US format: 1,234.56
                        numeric_value = clean_text.replace(',', '')
                elif ',' in clean_text:
                    # Only commas - treat as thousand separators unless it's clearly a decimal
                    if clean_text.count(',') == 1 and len(clean_text.split(',')[1]) <= 3:
                        # Likely decimal: "123,45" or "123,456"
                        numeric_value = clean_text.replace(',', '.')
                    else:
                        # Thousand separators: "1,234,567"
                        numeric_value = clean_text.replace(',', '')
                elif '.' in clean_text:
                    # Only periods - treat as decimal point if it appears once near the end
                    if clean_text.count('.') == 1 and len(clean_text.split('.')[1]) <= 3:
                        # Likely decimal: "123.45" or "123.456"
                        numeric_value = clean_text
                    else:
                        # Thousand separators: "1.234.567"
                        numeric_value = clean_text.replace('.', '')
                else:
                    # No separators
                    numeric_value = clean_text
                
                # Convert to float and apply multiplier only if explicitly indicated
                shares_value = float(numeric_value) * multiplier
                
                # Log the transformation for debugging
                if multiplier != 1:
                    logging.info(f"Applied explicit multiplier: {shares_text} -> {shares_value:,.0f}")
                
                # Format the number based on whether a unit was explicitly specified
                if re.search(r'million$|^m$|m$|M$', shares_text, re.IGNORECASE) or re.search(r'billion$|^b$|b$', shares_text, re.IGNORECASE):
                    # Only format as millions/billions if explicitly specified
                    if shares_value >= 1000000000:
                        formatted_value = f"{shares_value/1000000000:.2f} billion"
                    else:
                        formatted_value = f"{shares_value/1000000:.2f} million"
                else:
                    # Otherwise preserve the exact number, without commas for easier processing
                    formatted_value = f"{int(shares_value)}"
                
                logging.info(f"Final shares value: {formatted_value}")
                return {"outstanding_shares": formatted_value}
            except (ValueError, TypeError) as e:
                logging.error(f"Error parsing element shares value '{shares_text}': {e}")
                continue
                
        # Look for next sibling or parent that might contain the number
        sibling = element.find_next_sibling()
        if sibling:
            sibling_text = sibling.get_text(strip=True)
            number_match = re.search(r'([\d,\.\s]+)\s*(?:million|m|billion|b|k)?', sibling_text)
            if number_match:
                shares_text = number_match.group(0).strip()
                logging.info(f"Found potential outstanding shares in sibling (raw): {shares_text}")
                
                try:
                    # Check for explicit magnitude indicators
                    multiplier = 1
                    
                    # Only apply multipliers if explicitly indicated
                    if re.search(r'million$|^m$|m$', shares_text, re.IGNORECASE):
                        multiplier = 1000000
                        shares_text = re.sub(r'(?:million|m)$', '', shares_text, flags=re.IGNORECASE).strip()
                    elif re.search(r'billion$|^b$|b$', shares_text, re.IGNORECASE):
                        multiplier = 1000000000
                        shares_text = re.sub(r'(?:billion|b)$', '', shares_text, flags=re.IGNORECASE).strip()
                    # Handle K for thousands if present
                    elif re.search(r'k$', shares_text, re.IGNORECASE):
                        multiplier = 1000
                        shares_text = re.sub(r'k$', '', shares_text, flags=re.IGNORECASE).strip()
                    
                    # Special case for 'M' suffix (for million)
                    elif re.search(r'M$', shares_text):
                        multiplier = 1000000
                        shares_text = re.sub(r'M$', '', shares_text).strip()
                        
                    # Handle European number format with spaces as thousand separators
                    if ' ' in shares_text:
                        shares_text = shares_text.replace(' ', '')
                    
                    # Remove any remaining non-numeric characters except periods and commas
                    clean_text = re.sub(r'[^\d\.,]', '', shares_text)
                    
                    # Normalize decimal/thousand separators
                    if ',' in clean_text and '.' in clean_text:
                        # Both comma and period exist - determine which is the decimal separator
                        if clean_text.rindex(',') > clean_text.rindex('.'):
                            # European format: 1.234,56
                            numeric_value = clean_text.replace('.', '').replace(',', '.')
                        else:
                            # US format: 1,234.56
                            numeric_value = clean_text.replace(',', '')
                    elif ',' in clean_text:
                        # Only commas - treat as thousand separators unless it's clearly a decimal
                        if clean_text.count(',') == 1 and len(clean_text.split(',')[1]) <= 3:
                            # Likely decimal: "123,45" or "123,456"
                            numeric_value = clean_text.replace(',', '.')
                        else:
                            # Thousand separators: "1,234,567"
                            numeric_value = clean_text.replace(',', '')
                    elif '.' in clean_text:
                        # Only periods - treat as decimal point if it appears once near the end
                        if clean_text.count('.') == 1 and len(clean_text.split('.')[1]) <= 3:
                            # Likely decimal: "123.45" or "123.456"
                            numeric_value = clean_text
                        else:
                            # Thousand separators: "1.234.567"
                            numeric_value = clean_text.replace('.', '')
                    else:
                        # No separators
                        numeric_value = clean_text
                    
                    # Convert to float and apply multiplier only if explicitly indicated
                    shares_value = float(numeric_value) * multiplier
                    
                    # Log the transformation for debugging
                    if multiplier != 1:
                        logging.info(f"Applied explicit multiplier: {shares_text} -> {shares_value:,.0f}")
                    
                    # Format the number based on whether a unit was explicitly specified
                    if re.search(r'million$|^m$|m$|M$', shares_text, re.IGNORECASE) or re.search(r'billion$|^b$|b$', shares_text, re.IGNORECASE):
                        # Only format as millions/billions if explicitly specified
                        if shares_value >= 1000000000:
                            formatted_value = f"{shares_value/1000000000:.2f} billion"
                        else:
                            formatted_value = f"{shares_value/1000000:.2f} million"
                    else:
                        # Otherwise preserve the exact number, without commas for easier processing
                        formatted_value = f"{int(shares_value)}"
                    
                    logging.info(f"Final shares value: {formatted_value}")
                    return {"outstanding_shares": formatted_value}
                except (ValueError, TypeError) as e:
                    logging.error(f"Error parsing sibling shares value '{shares_text}': {e}")
                    continue
    
    return {"error": "Could not find outstanding shares information"}

def extract_outstanding_shares_with_ai_fallback(driver, url):
    """
    Extract outstanding shares with improved multi-tier extraction strategy:
    1. Custom domain extractors (improved)
    2. SIX Group specialized extractor  
    3. Traditional extraction
    4. AI fallback
    """
    extraction_method = "unknown"
    
    # Tier 1: Try improved custom domain-specific extractors first
    if extract_with_custom_function:
        try:
            # Check if we have a domain-specific extractor for this URL
            custom_result = extract_with_custom_function(driver, url)
            
            # Handle the custom extractor result properly
            if custom_result:
                if isinstance(custom_result, str):
                    # Clean the result (remove commas, spaces, etc)
                    cleaned_result = custom_result.replace(',', '').replace(' ', '')
                    
                    # If the cleaned result is numeric, it's valid
                    if cleaned_result.isdigit() or (
                        cleaned_result.replace('.', '', 1).isdigit() and 
                        cleaned_result.count('.') <= 1
                    ):
                        extraction_method = "improved_custom" if IMPROVED_EXTRACTORS_AVAILABLE else "custom"
                        logging.info(f"✅ {extraction_method.title()} extractor succeeded for {url}: {custom_result}")
                        return {"outstanding_shares": custom_result, "method": extraction_method}
                elif isinstance(custom_result, dict) and "outstanding_shares" in custom_result:
                    # Some custom extractors might return a dictionary with the shares
                    extraction_method = "improved_custom" if IMPROVED_EXTRACTORS_AVAILABLE else "custom"
                    logging.info(f"✅ {extraction_method.title()} extractor succeeded for {url}: {custom_result['outstanding_shares']}")
                    custom_result["method"] = extraction_method
                    return custom_result
                
                # Log if result was returned but not in expected format
                logging.info(f"Custom extractor returned non-numeric result: {custom_result}")
        except Exception as e:
            logging.warning(f"Custom extractor error for {url}: {e}")
    
    # Tier 2: Try SIX Group specialized extractor
    try:
        sixgroup_result = extract_sixgroup_shares(driver, url)
        if "outstanding_shares" in sixgroup_result:
            extraction_method = "sixgroup"
            logging.info(f"✅ SIX Group extractor succeeded for {url}")
            sixgroup_result["method"] = extraction_method
            return sixgroup_result
    except Exception as e:
        logging.warning(f"SIX Group extractor error for {url}: {e}")
    
    # Tier 3: Try traditional extraction
    try:
        result = extract_outstanding_shares(driver, url)
        if "outstanding_shares" in result:
            extraction_method = "traditional"
            logging.info(f"✅ Traditional extractor succeeded for {url}")
            result["method"] = extraction_method
            return result
    except Exception as e:
        logging.warning(f"Traditional extractor error for {url}: {e}")
        result = {"error": f"Traditional extraction failed: {e}"}
    
    # Tier 4: Try AI fallback as last resort
    if "error" in result:
        logging.info(f"🤖 All standard methods failed for {url}, trying AI fallback...")
        try:
            ai_result = try_shares_ai_fallback(driver, url)
            
            # If AI succeeded, return AI result (already has "(AI)" tag)
            if "outstanding_shares" in ai_result:
                extraction_method = "ai_fallback"
                ai_result["method"] = extraction_method
                logging.info(f"✅ AI fallback succeeded for {url}")
                return ai_result
        except Exception as e:
            logging.error(f"AI fallback error for {url}: {e}")
    
    # All methods failed
    extraction_method = "failed"
    logging.error(f"❌ All extraction methods failed for {url}")
    return {"error": "All extraction methods failed", "method": extraction_method}

def is_blue_cell(cell):
    """
    Check if a cell has any blue-ish color (simplified detection)
    Returns True if the cell appears to be blue, False otherwise
    """
    try:
        if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
            color = cell.fill.fgColor
            # Check different color representations
            if hasattr(color, 'rgb') and color.rgb:
                # Handle both string and RGB object types
                rgb_value = color.rgb
                if hasattr(rgb_value, 'value'):  # RGB object case
                    rgb_value = rgb_value.value if rgb_value.value else str(rgb_value)
                elif not isinstance(rgb_value, str):  # Other object types
                    rgb_value = str(rgb_value)
                
                # Remove alpha channel if present (ARGB format)
                if isinstance(rgb_value, str) and len(rgb_value) == 8:  # ARGB format
                    rgb_value = rgb_value[2:]  # Remove first 2 characters (alpha)
                
                # Simplified blue detection - check if it contains blue-ish values
                # Look for various shades of blue (not just the exact #00B0F0)
                if isinstance(rgb_value, str) and len(rgb_value) >= 6:
                    # Convert to uppercase for consistent checking
                    rgb_upper = rgb_value.upper()
                    # Check for common blue patterns
                    blue_patterns = [
                        '00B0F0',  # Original target blue
                        '0000FF',  # Pure blue
                        '4472C4',  # Another common blue
                        '5B9BD5',  # Light blue
                        '2F75B5',  # Dark blue
                    ]
                    
                    # Check if it matches any blue pattern or if blue component is dominant
                    for pattern in blue_patterns:
                        if pattern in rgb_upper:
                            return True
                    
                    # Check if the blue component (last 2 digits) is significantly higher than red/green
                    try:
                        if len(rgb_upper) == 6:
                            red = int(rgb_upper[0:2], 16)
                            green = int(rgb_upper[2:4], 16)
                            blue = int(rgb_upper[4:6], 16)
                            # Consider it blue if blue component is much higher than red and green
                            if blue > 150 and blue > (red + green):
                                return True
                    except ValueError:
                        pass
                        
            elif hasattr(color, 'indexed') and color.indexed is not None:
                # Handle indexed colors - some indexed colors are blue
                # Common blue indexed colors: 5 (blue), 41 (light blue), etc.
                blue_indexed = [5, 41, 44, 49]
                return color.indexed in blue_indexed
    except Exception:
        # If anything fails, just return False
        pass
    return False

def preserve_cell_color_and_set_value(cell, value):
    """
    Set cell value while preserving existing fill color
    """
    from openpyxl.styles import PatternFill
    
    # Store the current fill properties before modifying
    current_fill_info = None
    try:
        if cell.fill and hasattr(cell.fill, 'patternType') and cell.fill.patternType:
            # Extract fill properties to recreate later
            current_fill_info = {
                'patternType': cell.fill.patternType,
                'fgColor': None,
                'bgColor': None
            }
            
            # Try to extract foreground color
            if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                current_fill_info['fgColor'] = cell.fill.fgColor
            
            # Try to extract background color  
            if hasattr(cell.fill, 'bgColor') and cell.fill.bgColor:
                current_fill_info['bgColor'] = cell.fill.bgColor
    except Exception:
        # If we can't extract fill info, we'll just set the value without preserving color
        current_fill_info = None
    
    # Set the value
    cell.value = value
    
    # Restore the fill if we extracted the info successfully
    if current_fill_info and current_fill_info['patternType']:
        try:
            # Create a new PatternFill with the same properties
            new_fill = PatternFill(
                patternType=current_fill_info['patternType'],
                fgColor=current_fill_info['fgColor'],
                bgColor=current_fill_info['bgColor']
            )
            cell.fill = new_fill
        except Exception:
            # If restoring fails, just leave the cell with its new value
            pass

def main():
    """
    Main function to find and extract outstanding shares
    """
    print("Outstanding Shares Updater Starting...")
    
    # Log which custom extractors are available
    if IMPROVED_EXTRACTORS_AVAILABLE:
        print("✅ Using improved custom domain extractors for:")
        for domain in DOMAIN_EXTRACTORS.keys():
            print(f"  - {domain}")
    elif CUSTOM_EXTRACTORS_AVAILABLE:
        print("⚠️ Using original custom domain extractors")
    else:
        print("❌ No custom domain extractors available")
    
    # Initialize driver to None first
    driver = None
    original_filename = "data/Custodians.xlsx"
    results_filename = "data/Custodians_Results.xlsx"
    
    # Initialize error tracking
    from urllib.parse import urlparse
    from collections import defaultdict
    from datetime import datetime
    
    error_domains = defaultdict(int)
    total_processed = 0
    total_successful = 0
    blue_cells_count = 0
    
    # Track extraction methods used
    method_stats = defaultdict(int)
    method_stats["improved_custom"] = 0
    method_stats["custom"] = 0
    method_stats["sixgroup"] = 0
    method_stats["traditional"] = 0
    method_stats["ai_fallback"] = 0
    method_stats["failed"] = 0
    
    def track_error_domain(url):
        """Helper function to track error domains with improved logic."""
        try:
            if not url or not isinstance(url, str):
                return "invalid://"
            
            # Clean the URL
            cleaned_url = url.strip()
            if not cleaned_url.lower().startswith(("http://", "https://")):
                return "invalid://"
            
            parsed_url = urlparse(cleaned_url)
            if parsed_url.netloc:
                return f"{parsed_url.scheme}://{parsed_url.netloc}/"
            else:
                # If no domain can be extracted, use the full URL
                return cleaned_url[:100] + ("..." if len(cleaned_url) > 100 else "")
        except Exception:
            # If all parsing fails, use the original URL (truncated if too long)
            url_str = str(url) if url else "unknown"
            return url_str[:100] + ("..." if len(url_str) > 100 else "")
    
    # Only create fresh copy if results file doesn't exist
    # This allows the batch file to run stock updater first, then shares updater
    if not os.path.exists(results_filename):
        # Copy original to results file
        import shutil
        shutil.copy2(original_filename, results_filename)
        print(f"Copied {original_filename} to {results_filename}")
    else:
        print(f"Using existing {results_filename} (keeping previous results)")
    
    try:
        # Check if the Excel file is accessible for reading
        if not os.path.exists(results_filename):
            print(f"Error: File {results_filename} not found")
            return
        
        try:
            # Load the Excel workbook and worksheet
            wb = load_workbook(results_filename)
            
            # Check if the "Shares" worksheet exists, if not use the first worksheet
            if "Shares" in wb.sheetnames:
                ws = wb["Shares"]
            else:
                # Use the first worksheet as fallback
                ws = wb.active
                print(f"Worksheet 'Shares' not found, using '{ws.title}' instead")
        except PermissionError:
            print(f"Error: Cannot open {results_filename}. The file may be open in Excel or another program.")
            return
        except Exception as e:
            print(f"Error opening Excel file: {e}")
            return
            
        # Find all rows to process (in column P) - dynamically detect all rows with URLs
        processable_rows_urls = find_processable_rows_and_get_urls(ws, "P")
        print(f"Found {len(processable_rows_urls)} rows to process")
        
        # Set up the WebDriver with improved options
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920x1080")
        # Add user agent to avoid detection
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        # Disable images to speed up page load
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs)
        
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        
        # Process each URL
        for row_idx, primary_url in processable_rows_urls.items():
            total_processed += 1
            print(f"Processing row {row_idx}: {primary_url}")
            shares_data = None
            used_url = primary_url
            
            try:
                # First try the URL from column P
                # Check if it's a SIX Group URL
                if "six-group.com" in primary_url.lower():
                    print("Detected SIX Group URL, using specialized extractor")
                    shares_data = extract_sixgroup_shares(driver, primary_url)
                else:
                    # Use the generic extractor for other URLs
                    shares_data = extract_outstanding_shares_with_ai_fallback(driver, primary_url)
                
                # If primary URL failed, try fallback URL from column Q (if it exists)
                if "error" in shares_data:
                    print(f"  Primary URL failed: {shares_data['error']}")
                    fallback_url = ws[f"Q{row_idx}"].value
                    
                    if fallback_url and isinstance(fallback_url, str) and fallback_url.startswith(("http://", "https://")):
                        print(f"  Trying fallback URL from column Q: {fallback_url}")
                        used_url = fallback_url
                        
                        # Check if it's a SIX Group URL
                        if "six-group.com" in fallback_url.lower():
                            print("  Detected SIX Group URL for fallback, using specialized extractor")
                            shares_data = extract_sixgroup_shares(driver, fallback_url)
                        else:
                            # Use the generic extractor for other URLs
                            shares_data = extract_outstanding_shares_with_ai_fallback(driver, fallback_url)
            except Exception as e:
                print(f"  Error processing URLs for row {row_idx}: {str(e)}")
                traceback.print_exc()
                shares_data = {"error": f"Exception: {str(e)}"}
                
                # Try to reset the browser if we encounter issues
                try:
                    print("  Attempting to reset the browser...")
                    driver.quit()
                    driver = webdriver.Chrome(
                        service=ChromeService(ChromeDriverManager().install()),
                        options=options
                    )
                except Exception as reset_error:
                    print(f"  Failed to reset browser: {str(reset_error)}")
            
            # Track success/failure, domain errors, and extraction methods
            if "error" in shares_data:
                error_domains[track_error_domain(used_url)] += 1
                method_stats["failed"] += 1
            else:
                total_successful += 1
                # Track which method was successful
                method_used = shares_data.get("method", "unknown")
                method_stats[method_used] += 1
            
            # Write result to Excel in column M
            shares_cell = f"M{row_idx}"
            cell = ws[shares_cell]
            
            # Check if this cell is blue and count it
            if is_blue_cell(cell):
                blue_cells_count += 1
                print(f"  Row {row_idx}: Blue cell detected - excluding from success rate calculation")
            
            if "error" in shares_data:
                print(f"  Error: {shares_data['error']}")
                preserve_cell_color_and_set_value(cell, f"Error: {shares_data['error']}")
            else:
                print(f"  Found outstanding shares: {shares_data['outstanding_shares']} (using {used_url})")
                preserve_cell_color_and_set_value(cell, shares_data['outstanding_shares'])
            
            # Save after each update in case of crash
            try:
                wb.save(results_filename)
            except PermissionError:
                print(f"  Warning: Could not save to {results_filename}, it may be open in another program")
                # Try with a different filename
                alt_filename = f"Custodians_Results_{int(time.time())}.xlsx"
                print(f"  Attempting to save to alternative file: {alt_filename}")
                try:
                    wb.save(alt_filename)
                    results_filename = alt_filename  # Use this filename for future saves
                    print(f"  Successfully saved to {alt_filename}")
                except Exception as e2:
                    print(f"  Could not save Excel file: {e2}")
            except Exception as e:
                print(f"  Error saving workbook: {e}")
        
        # Update column header with timestamp
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        header_cell = ws["M1"]
        header_cell.value = f"Outstanding shares (Last updated: {current_time})"
        
        # Save final results
        try:
            wb.save(results_filename)
            print(f"Processing completed and saved to {results_filename}")
            
            # Print method statistics summary
            print("\n" + "="*50)
            print("EXTRACTION METHOD STATISTICS:")
            print("="*50)
            for method, count in method_stats.items():
                if count > 0:
                    percentage = (count / total_processed) * 100 if total_processed > 0 else 0
                    print(f"{method.replace('_', ' ').title()}: {count} ({percentage:.1f}%)")
            
            print(f"\nSUCCESS SUMMARY:")
            print(f"Total URLs processed: {total_processed}")
            print(f"Successful extractions: {total_successful}")
            if total_processed > 0:
                success_rate = (total_successful / total_processed) * 100
                print(f"Overall success rate: {success_rate:.1f}%")
            print("="*50)
            
        except Exception as e:
            print(f"Error saving final results: {e}")
        
        # Create a copy of the results as custodians.xlsx (overwriting the original)
        try:
            import shutil
            shutil.copy2(results_filename, original_filename)
            print(f"Created copy: {results_filename} -> {original_filename}")
        except Exception as e:
            print(f"Error creating custodians.xlsx copy: {e}")
        
        # Create log file
        log_filename = f"outstanding_shares_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        try:
            with open(log_filename, 'w', encoding='utf-8') as log_file:
                log_file.write("OUTSTANDING SHARES EXTRACTION LOG\n")
                log_file.write("=" * 50 + "\n")
                log_file.write(f"Run Date: {current_time}\n")
                log_file.write(f"Script: outstanding_shares_updater.py\n\n")
                
                log_file.write("OUTSTANDING SHARES ERRORS BY DOMAIN:\n")
                log_file.write("-" * 40 + "\n")
                if error_domains:
                    for domain, count in sorted(error_domains.items()):
                        log_file.write(f"{domain} errors: {count}\n")
                else:
                    log_file.write("No errors encountered!\n")
                
                log_file.write(f"\nEXTRACTION METHOD STATISTICS:\n")
                log_file.write("-" * 30 + "\n")
                for method, count in method_stats.items():
                    if count > 0:
                        percentage = (count / total_processed) * 100 if total_processed > 0 else 0
                        log_file.write(f"{method.replace('_', ' ').title()}: {count} ({percentage:.1f}%)\n")
                
                log_file.write(f"\nSUMMARY:\n")
                log_file.write("-" * 10 + "\n")
                log_file.write(f"Total URLs processed: {total_processed}\n")
                log_file.write(f"Blue cells (excluded from success rate): {blue_cells_count}\n")
                log_file.write(f"Successful extractions: {total_successful}\n")
                log_file.write(f"Effective URLs for success calculation: {total_processed - blue_cells_count}\n")
                if (total_processed - blue_cells_count) > 0:
                    success_rate = total_successful / (total_processed - blue_cells_count)
                    log_file.write(f"Adjusted success rate: {total_successful}/{total_processed - blue_cells_count} = {success_rate:.2%}\n")
                else:
                    log_file.write("Adjusted success rate: N/A (no valid URLs to process)\n")
                
            print(f"Log file created: {log_filename}")
        except Exception as e:
            print(f"Error creating log file: {e}")
        
        # Run analysis after processing
        from src.run_analysis import run_analysis
        log_file = run_analysis(original_filename, "outstanding_shares")
        print(f"\nAnalysis complete! Check the detailed report at: {log_file}")
        
    except Exception as e:
        print(f"Unexpected error: {e}")
        traceback.print_exc()
    finally:
        # Close the WebDriver
        if driver:
            driver.quit()

def test_specific_urls():
    """Test function for specific URLs"""
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    test_urls = [
        "https://etf.dws.com/en-fi/etc/CH1315732268-xtrackers-galaxy-physical-ethereum-etc-securities/",
        "https://institutional.fidelity.com/prgw/digital/research/quote/dashboard/summary?symbol=FETH"
    ]
    
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    )

    driver = None
    
    try:
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        driver.set_page_load_timeout(60)
        
        for url in test_urls:
            logging.info(f"Testing URL: {url}")
            result = extract_outstanding_shares_with_ai_fallback(driver, url)
            logging.info(f"Result: {result}")
            print(f"URL: {url}")
            print(f"Result: {result}")
            print("-" * 50)
            
    except Exception as e:
        logging.error(f"Test error: {e}", exc_info=True)
    finally:
        if driver:
            driver.quit()

def test_fallback_urls():
    """
    Test function to verify the fallback URL functionality.
    Creates a test Excel file with primary and fallback URLs and runs the extraction process.
    """
    from openpyxl import Workbook
    import os
    
    print("Testing fallback URL functionality...")
    
    # Create a test workbook
    test_wb = Workbook()
    test_ws = test_wb.active
    test_ws.title = "Shares"
    
    # Add headers
    test_ws['L1'] = "Source"
    test_ws['P1'] = "Primary URL"
    test_ws['Q1'] = "Fallback URL"
    test_ws['M1'] = "Outstanding Shares"
    
    # Test case 1: Primary URL works (SIX Group)
    test_ws['L2'] = "Test 1: Primary URL works"
    test_ws['P2'] = "https://www.six-group.com/en/market-data/etp/etp-explorer/etp-detail.CH1146882316USD4.html#/"
    test_ws['Q2'] = "https://www.six-group.com/en/market-data/etp/etp-explorer/etp-detail.CH0496454155USD4.html#/"
    
    # Test case 2: Primary URL fails, fallback works (SIX Group)
    test_ws['L3'] = "Test 2: Primary fails, fallback works"
    test_ws['P3'] = "https://www.invalid-url.com/not-exists"
    test_ws['Q3'] = "https://www.six-group.com/en/market-data/etp/etp-explorer/etp-detail.CH0475552201USD4.html#/"
    
    # Test case 3: Both URLs fail
    test_ws['L4'] = "Test 3: Both URLs fail"
    test_ws['P4'] = "https://www.invalid-url1.com/not-exists"
    test_ws['Q4'] = "https://www.invalid-url2.com/not-exists"
    
    # Save test workbook
    test_filename = "test_fallback.xlsx"
    test_wb.save(test_filename)
    print(f"Created test file: {test_filename}")
    
    # Load the test workbook for processing
    wb = load_workbook(test_filename)
    ws = wb["Shares"]
    
    # Set up the WebDriver
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920x1080")
    
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=options
    )
    
    try:
        # Process each URL
        for row_idx in range(2, 5):  # Rows 2, 3, 4
            primary_url = ws[f"P{row_idx}"].value
            print(f"\nProcessing row {row_idx}: {ws[f'L{row_idx}'].value}")
            print(f"Primary URL: {primary_url}")
            
            shares_data = None
            used_url = primary_url
            
            # First try the URL from column P
            try:
                # Check if it's a SIX Group URL
                if primary_url and "six-group.com" in primary_url.lower():
                    print("Detected SIX Group URL, using specialized extractor")
                    shares_data = extract_sixgroup_shares(driver, primary_url)
                else:
                    # Use the generic extractor for other URLs
                    shares_data = extract_outstanding_shares_with_ai_fallback(driver, primary_url)
            except Exception as e:
                shares_data = {"error": f"Exception: {str(e)}"}
            
            # If primary URL failed, try fallback URL from column Q
            if "error" in shares_data:
                print(f"  Primary URL failed: {shares_data['error']}")
                fallback_url = ws[f"Q{row_idx}"].value
                
                if fallback_url and isinstance(fallback_url, str) and fallback_url.startswith(("http://", "https://")):
                    print(f"  Trying fallback URL: {fallback_url}")
                    used_url = fallback_url
                    
                    try:
                        # Check if it's a SIX Group URL
                        if "six-group.com" in fallback_url.lower():
                            print("  Detected SIX Group URL for fallback, using specialized extractor")
                            shares_data = extract_sixgroup_shares(driver, fallback_url)
                        else:
                            # Use the generic extractor for other URLs
                            shares_data = extract_outstanding_shares_with_ai_fallback(driver, fallback_url)
                    except Exception as e:
                        shares_data = {"error": f"Exception in fallback: {str(e)}"}
            
            # Write result to Excel in column M
            shares_cell = f"M{row_idx}"
            if "error" in shares_data:
                print(f"  Final result: Error: {shares_data['error']}")
                ws[shares_cell] = f"Error: {shares_data['error']}"
            else:
                print(f"  Final result: Found outstanding shares: {shares_data['outstanding_shares']} (using {used_url})")
                ws[shares_cell] = shares_data['outstanding_shares']
        
        # Save results
        result_filename = "test_fallback_results.xlsx"
        wb.save(result_filename)
        print(f"\nTest completed. Results saved to: {result_filename}")
        
    except Exception as e:
        print(f"Test error: {e}")
    finally:
        # Close the WebDriver
        if driver:
            driver.quit()

def test_custom_domain_extractors():
    """
    Test the custom domain extractors implementation to verify they work correctly.
    This function tests URLs that should be handled by the custom extractors in
    improved_custom_domain_extractors.py
    """
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # Test URLs for each custom domain extractor
    test_urls = [
        # Valour URLs
        "https://valour.com/digital-assets/bitcoin",
        # Grayscale URLs
        "https://grayscale.com/products/grayscale-bitcoin-trust/",
        # VanEck URLs
        "https://www.vaneck.com/us/en/investments/bitcoin-trust-gbtc/",
        # VanEck DE URLs
        "https://www.vaneck.com/de/en/bitcoin-etn/",
        # WisdomTree URLs 
        "https://www.wisdomtree.eu/en-gb/products/ucits-etfs-unleveraged-etps/cryptocurrencies/wisdomtree-bitcoin",
        # ProShares URLs
        "https://www.proshares.com/our-etfs/strategic/bito"
    ]
    
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    )

    driver = None
    
    try:
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=options
        )
        driver.set_page_load_timeout(60)
        
        print("=" * 50)
        print("TESTING CUSTOM DOMAIN EXTRACTORS")
        print("=" * 50)
        
        for url in test_urls:
            print(f"\nTesting URL: {url}")
            
            # 1. First test the direct custom extractor function
            if extract_with_custom_function:
                try:
                    print("Testing direct custom extractor function...")
                    direct_result = extract_with_custom_function(driver, url)
                    print(f"Direct result: {direct_result}")
                except Exception as e:
                    print(f"Direct extraction error: {e}")
            
            # 2. Then test the full extraction pipeline
            try:
                print("Testing full extraction pipeline...")
                full_result = extract_outstanding_shares_with_ai_fallback(driver, url)
                print(f"Full pipeline result: {full_result}")
                
                # Check if the custom extractor was used
                if full_result.get("method") in ["improved_custom", "custom"]:
                    print("✅ SUCCESS: Custom extractor was used!")
                else:
                    print(f"❌ FAILURE: Custom extractor was not used. Method: {full_result.get('method')}")
            except Exception as e:
                print(f"Full pipeline error: {e}")
            
            print("-" * 50)
            
    except Exception as e:
        logging.error(f"Test error: {e}", exc_info=True)
    finally:
        if driver:
            driver.quit()
    
    print("\nCustom domain extractor testing completed.")

if __name__ == "__main__":
    # Uncomment one of these based on what you want to run
    main()  # Process the Excel file
    # test_specific_urls()  # Test specific URLs only 
    # test_fallback_urls()  # Test fallback URL functionality
    # test_custom_domain_extractors()  # Test custom domain extractors 